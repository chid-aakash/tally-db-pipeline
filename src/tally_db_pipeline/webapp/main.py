"""FastAPI web app for production-entry data capture.

Flow:
  /                        -> pick company, date, voucher type (policy-driven)
  POST /entries            -> create draft entry, redirect to edit
  /entries/{id}/edit       -> render policy-allowed items as a table
  POST /entries/{id}/save  -> save lines
  /entries/{id}/review     -> voucher preview
  POST /entries/{id}/post  -> build voucher_dict, call TallyClient.import_voucher
  /entries/{id}            -> result view
  /entries                 -> list entries
"""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from fastapi import FastAPI, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..config import get_settings
from ..db import engine, get_session
from ..models import (
    Base,
    Company,
    CompanyAlias,
    ConsumptionReportSelection,
    DailyProductionReport,
    DPRHourlyCell,
    DPRHourModel,
    DPRIdleEvent,
    ProcessCatalogEntry,
    ProcessStage,
    LineDailyVoucherPost,
    ProductionEntry,
    ProductionEntryLine,
    ProductionModelHole,
    ProductionModelSpec,
    ProductionProcess,
    Shift,
    ShiftPresetSlot,
    SJPolicy,
    StockGroup,
    StockItem,
    Voucher,
    VoucherInventoryEntry,
)
from .. import daily_report as dpr
from ..policy import (
    default_godown_for_role,
    entry_to_voucher_dict,
    generate_remote_id,
    get_policy,
    list_policies,
    resolve_items_for_role,
)
from ..tally_client import TallyClient
from .. import line_voucher
from .. import onedrive_client


TEMPLATES_DIR = Path(__file__).parent / "templates"
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))


def _inr(value) -> str:
    """Format a number as Indian Rupee with lakh/crore-style commas."""
    try:
        n = float(value or 0)
    except (TypeError, ValueError):
        return ""
    negative = n < 0
    n = abs(n)
    whole, _, frac = f"{n:.2f}".partition(".")
    if len(whole) <= 3:
        grouped = whole
    else:
        last3 = whole[-3:]
        rest = whole[:-3]
        # Group the rest in pairs from the right
        pairs = []
        while len(rest) > 2:
            pairs.insert(0, rest[-2:])
            rest = rest[:-2]
        if rest:
            pairs.insert(0, rest)
        grouped = ",".join(pairs) + "," + last3
    result = f"₹{grouped}.{frac}"
    return f"-{result}" if negative else result


def _inum(value) -> str:
    """Format a number with Indian comma grouping (xx,xx,xxx.xx), no symbol."""
    try:
        n = float(value or 0)
    except (TypeError, ValueError):
        return ""
    negative = n < 0
    n = abs(n)
    whole, _, frac = f"{n:.2f}".partition(".")
    if len(whole) <= 3:
        grouped = whole
    else:
        last3 = whole[-3:]
        rest = whole[:-3]
        pairs = []
        while len(rest) > 2:
            pairs.insert(0, rest[-2:])
            rest = rest[:-2]
        if rest:
            pairs.insert(0, rest)
        grouped = ",".join(pairs) + "," + last3
    result = f"{grouped}.{frac}"
    return f"-{result}" if negative else result


def _nice_date(iso: str) -> str:
    """Format an ISO date (YYYY-MM-DD) as '24<sup>th</sup> April 26' (HTML)."""
    from markupsafe import Markup
    if not iso:
        return ""
    try:
        d = datetime.strptime(str(iso)[:10], "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return str(iso)
    n = d.day
    if 10 <= n % 100 <= 20:
        suf = "th"
    else:
        suf = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return Markup(
        f'<span style="white-space:nowrap">{n}<sup>{suf}</sup> {d.strftime("%B")} {d.strftime("%y")}</span>'
    )


def _dmy(value) -> str:
    """Format a date (or YYYY-MM-DD string) as DD-MM-YYYY for display.
    Returns the input unchanged if it can't be parsed, so non-date strings
    pass through harmlessly."""
    if value in (None, ""):
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%d-%m-%Y")
    s = str(value)[:10]
    try:
        return datetime.strptime(s, "%Y-%m-%d").strftime("%d-%m-%Y")
    except (TypeError, ValueError):
        return str(value)


def _co_short(name: str | None) -> str:
    """Map a full company name to its short code for print/PDF output.

    Why: the registered names ("Avinash Appliances Private Limited (April 26 - March 27)",
    "Shanke Enterprise ...") are too long for letterheads; users want AAPL / SEPL.
    """
    n = (name or "").lower()
    if "avinash" in n:
        return "AAPL"
    if "shanke" in n:
        return "SEPL"
    return (name or "").strip()


templates.env.filters["inr"] = _inr
templates.env.filters["inum"] = _inum
templates.env.filters["nice_date"] = _nice_date
templates.env.filters["dmy"] = _dmy
templates.env.filters["co_short"] = _co_short

app = FastAPI(title="Tally Production Entry")


def _self_heal_production_processes() -> None:
    """Drop the production_processes table if it carries the obsolete unique
    constraint on (company, line, section, label). That constraint blocked
    legitimate repeated rejection labels under different group headings.
    Safe because the table only holds DPR config rows and is regenerated by
    the seed flow on the config page."""
    from sqlalchemy import inspect, text
    insp = inspect(engine)
    if "production_processes" not in insp.get_table_names():
        return
    constraints = insp.get_unique_constraints("production_processes")
    if any(c.get("name") == "uq_pp_company_line_section_label" for c in constraints):
        with engine.begin() as conn:
            conn.execute(text("DROP TABLE production_processes"))


_self_heal_production_processes()


def _self_heal_production_process_role() -> None:
    """Add the `role` column to production_processes / process_catalog if
    they predate the Input/Output split, then backfill production rows by
    inferring from the label ('input' substring → 'input', else 'output')."""
    from sqlalchemy import inspect, text
    insp = inspect(engine)
    for table in ("production_processes", "process_catalog"):
        if table not in insp.get_table_names():
            continue
        cols = {c["name"] for c in insp.get_columns(table)}
        if "role" in cols:
            continue
        with engine.begin() as conn:
            conn.execute(text(f"ALTER TABLE {table} ADD COLUMN role VARCHAR(10)"))
            conn.execute(text(
                f"UPDATE {table} SET role = "
                "CASE WHEN section = 'production' AND lower(label) LIKE '%input%' THEN 'input' "
                "     WHEN section = 'production' THEN 'output' "
                "     ELSE NULL END"
            ))


_self_heal_production_process_role()


def _self_heal_validate_count_column() -> None:
    """Add the `validate_count` column where missing, defaulting to TRUE
    (1) so existing production rows keep their pre-flag behaviour."""
    from sqlalchemy import inspect, text
    insp = inspect(engine)
    for table in ("production_processes", "process_catalog"):
        if table not in insp.get_table_names():
            continue
        cols = {c["name"] for c in insp.get_columns(table)}
        if "validate_count" in cols:
            continue
        with engine.begin() as conn:
            conn.execute(text(f"ALTER TABLE {table} ADD COLUMN validate_count BOOLEAN NOT NULL DEFAULT 1"))


_self_heal_validate_count_column()


def _self_heal_drop_legacy_spec_columns() -> None:
    """Drop the original single-hole columns from production_model_specs.

    The first cut of the model-spec feature stored one (radius, count) pair
    per model. Multi-diameter holes now live in production_model_holes, so
    these columns are dead weight — and worse, they are NOT NULL so any
    INSERT that omits them fails. Drop them on startup if present (SQLite
    3.35+ supports DROP COLUMN)."""
    from sqlalchemy import inspect, text
    insp = inspect(engine)
    if "production_model_specs" not in insp.get_table_names():
        return
    cols = {c["name"] for c in insp.get_columns("production_model_specs")}
    legacy = [c for c in ("hole_radius_mm", "hole_count") if c in cols]
    if not legacy:
        return
    with engine.begin() as conn:
        for col in legacy:
            try:
                conn.execute(text(f"ALTER TABLE production_model_specs DROP COLUMN {col}"))
            except Exception:
                # If SQLite is too old, leave the column — the ORM model
                # below now declares it as nullable so inserts still work.
                pass


_self_heal_drop_legacy_spec_columns()


def _self_heal_add_price_columns() -> None:
    """Add blank/drilled/printed price columns to production_model_specs if missing."""
    from sqlalchemy import inspect, text
    insp = inspect(engine)
    if "production_model_specs" not in insp.get_table_names():
        return
    cols = {c["name"] for c in insp.get_columns("production_model_specs")}
    with engine.begin() as conn:
        for col in ("blank_price", "drilled_price", "printed_price"):
            if col not in cols:
                conn.execute(text(f"ALTER TABLE production_model_specs ADD COLUMN {col} FLOAT"))


_self_heal_add_price_columns()


def _self_heal_shift_preset_times() -> None:
    """Add `from_time` / `to_time` columns to shift_preset_slots if missing.
    Pre-existing rows keep `label` only; from/to stay NULL until re-saved."""
    from sqlalchemy import inspect, text
    insp = inspect(engine)
    if "shift_preset_slots" not in insp.get_table_names():
        return
    cols = {c["name"] for c in insp.get_columns("shift_preset_slots")}
    with engine.begin() as conn:
        for col in ("from_time", "to_time"):
            if col not in cols:
                conn.execute(text(f"ALTER TABLE shift_preset_slots ADD COLUMN {col} VARCHAR(5)"))


_self_heal_shift_preset_times()
# Ensure newly introduced tables exist (idempotent). Sync also does this, but the
# webapp may be started before a full sync on a fresh machine.
Base.metadata.create_all(bind=engine)


CONSUMPTION_DEFAULT_VOUCHER_TYPES = [
    "Stock Journal",
    "Delivery Note",
    "Rejections Out",
    "Physical Stock",
    "Material Out",
]


def _session() -> Session:
    return get_session()


def _tally_client() -> TallyClient:
    settings = get_settings()
    return TallyClient(
        host=settings.tally_host,
        port=settings.tally_port,
        timeout=settings.tally_timeout_seconds,
        request_delay_ms=settings.tally_request_delay_ms,
        max_retries=settings.tally_max_retries,
        retry_backoff_ms=settings.tally_retry_backoff_ms,
        lock_file=settings.tally_lock_file,
        lock_stale_seconds=settings.tally_lock_stale_seconds,
    )


def _primary_uom(raw: str | None) -> str:
    """Strip any compound / conversion suffix from a UOM string.
    e.g. 'Mtrs = 2136.364 SQM' -> 'Mtrs'; 'No.' -> 'No.'."""
    if not raw:
        return ""
    s = str(raw).strip()
    for sep in [" = ", "=", " of "]:
        if sep in s:
            s = s.split(sep, 1)[0].strip()
            break
    return s


def _build_summary(session: Session, summary_date: str, company: str | None):
    """Group posted entries for (date) by company → voucher_type → item.
    Returns (companies, meta) where:
      companies = [{name, groups: [{voucher_type, rows, ...}], subtotal, voucher_count}]
    """
    sstmt = select(ProductionEntry).where(
        ProductionEntry.entry_date == summary_date,
        ProductionEntry.status == "posted",
    )
    if company:
        sstmt = sstmt.where(ProductionEntry.company_name == company)
    posted = list(session.scalars(sstmt))

    # company -> voucher_type -> item_name -> agg
    by_co: dict[str, dict[str, dict[str, dict]]] = {}
    vch_by_co_vt: dict[tuple[str, str], set] = {}
    count_by_co: dict[str, int] = {}
    total_amount = 0.0
    for e in posted:
        co = e.company_name
        vt = e.voucher_type
        bucket = by_co.setdefault(co, {}).setdefault(vt, {})
        vch_by_co_vt.setdefault((co, vt), set())
        if e.tally_voucher_number:
            vch_by_co_vt[(co, vt)].add(str(e.tally_voucher_number))
        count_by_co[co] = count_by_co.get(co, 0) + 1
        for l in e.lines:
            if not l.quantity or l.role != "consume":
                continue
            r = bucket.setdefault(l.item_name, {
                "item_name": l.item_name,
                "uom": _primary_uom(l.uom),
                "quantity": 0.0,
                "amount": 0.0,
                "godowns": set(),
                "vchs": set(),
            })
            r["quantity"] += float(l.quantity or 0)
            r["amount"] += float(l.amount or 0)
            total_amount += float(l.amount or 0)
            if l.godown: r["godowns"].add(l.godown)
            if e.tally_voucher_number: r["vchs"].add(str(e.tally_voucher_number))

    companies = []
    for co in sorted(by_co.keys()):
        groups = []
        for vt in sorted(by_co[co].keys()):
            items = [
                {
                    "item_name": r["item_name"],
                    "uom": r["uom"],
                    "quantity": r["quantity"],
                    "amount": r["amount"],
                    "rate_avg": (r["amount"] / r["quantity"]) if r["quantity"] else 0.0,
                    "godowns": ", ".join(sorted(r["godowns"])),
                    "vchs": ", ".join(sorted(r["vchs"])),
                }
                for r in sorted(by_co[co][vt].values(), key=lambda x: x["item_name"])
            ]
            groups.append({
                "voucher_type": vt,
                "voucher_numbers": ", ".join(sorted(vch_by_co_vt[(co, vt)])),
                "rows": items,
                "subtotal": sum(r["amount"] for r in items),
                "item_count": len(items),
            })
        companies.append({
            "name": co,
            "groups": groups,
            "subtotal": sum(g["subtotal"] for g in groups),
            "voucher_count": count_by_co[co],
        })
    meta = {
        "total_amount": total_amount,
        "voucher_count": len(posted),
        "company": company or (posted[0].company_name if posted else ""),
        "multi_company": len(by_co) > 1,
    }
    return companies, meta


def _list_companies(session: Session) -> list[str]:
    # Prefer companies that have policies loaded
    with_policy = set(
        session.scalars(select(SJPolicy.company_name).distinct())
    )
    all_companies = list(session.scalars(select(Company.name).order_by(Company.name)))
    if with_policy:
        return [c for c in all_companies if c in with_policy] or all_companies
    return all_companies


@app.get("/")
def root():
    return RedirectResponse("/app", status_code=307)


@app.get("/legacy", response_class=HTMLResponse)
def home(request: Request):
    with _session() as session:
        companies = _list_companies(session)
        company = companies[0] if companies else ""
        policies = list_policies(session, company) if company else []
    return templates.TemplateResponse(
        request,
        "home.html",
        {
            "companies": companies,
            "selected_company": company,
            "policies": policies,
            "today": date.today().isoformat(),
        },
    )


@app.post("/entries")
def create_entry(
    company: str = Form(...),
    entry_date: str = Form(...),
    voucher_type: str = Form(...),
    narration: str = Form(""),
):
    with _session() as session:
        policy = get_policy(session, company, voucher_type)
        if policy is None:
            raise HTTPException(400, f"No active policy for {company} / {voucher_type}")

        # Insert entry without remote_id first so we can derive it from id.
        entry = ProductionEntry(
            remote_id="pending",
            company_name=company,
            entry_date=entry_date,
            voucher_type=voucher_type,
            status="draft",
            narration=narration or None,
        )
        session.add(entry)
        session.flush()
        entry.remote_id = generate_remote_id(entry_date, voucher_type, entry.id)

        # Seed lines for all policy-allowed items (consume + produce) with qty=0
        for role in ("consume", "produce"):
            items = resolve_items_for_role(session, policy, role)
            godown = default_godown_for_role(policy, role)
            for item in items:
                session.add(
                    ProductionEntryLine(
                        entry_id=entry.id,
                        role=role,
                        item_name=item.name,
                        quantity=0.0,
                        uom=item.closing_uom or item.base_units or "No.",
                        rate=item.closing_rate or 0.0,
                        amount=0.0,
                        godown=godown,
                        opening_stock_snapshot=item.closing_quantity or 0.0,
                    )
                )
        session.commit()
        return RedirectResponse(f"/entries/{entry.id}/edit", status_code=303)


def _load_entry(session: Session, entry_id: int) -> ProductionEntry:
    entry = session.get(ProductionEntry, entry_id)
    if entry is None:
        raise HTTPException(404, "Entry not found")
    return entry


@app.get("/entries/{entry_id}/edit", response_class=HTMLResponse)
def edit_entry(request: Request, entry_id: int):
    with _session() as session:
        entry = _load_entry(session, entry_id)
        consume_lines = sorted(
            [l for l in entry.lines if l.role == "consume"], key=lambda l: l.item_name
        )
        produce_lines = sorted(
            [l for l in entry.lines if l.role == "produce"], key=lambda l: l.item_name
        )
    return templates.TemplateResponse(
        request,
        "entry.html",
        {
            "entry": entry,
            "consume_lines": consume_lines,
            "produce_lines": produce_lines,
        },
    )


@app.post("/entries/{entry_id}/save")
async def save_entry(entry_id: int, request: Request):
    form = await request.form()
    with _session() as session:
        entry = _load_entry(session, entry_id)
        if entry.status not in ("draft", "failed"):
            raise HTTPException(400, f"Cannot edit entry in status {entry.status}")

        narration = form.get("narration", "")
        entry.narration = narration or None

        for line in entry.lines:
            qty_raw = form.get(f"qty_{line.id}", "")
            rate_raw = form.get(f"rate_{line.id}", "")
            godown_raw = form.get(f"godown_{line.id}", "")
            try:
                qty = float(qty_raw) if qty_raw.strip() else 0.0
            except ValueError:
                qty = 0.0
            try:
                rate = float(rate_raw) if rate_raw.strip() else line.rate
            except ValueError:
                rate = line.rate
            line.quantity = qty
            line.rate = rate
            line.amount = round(qty * rate, 2)
            if godown_raw:
                line.godown = godown_raw

        session.commit()

        action = form.get("action", "save")
        if action == "review":
            return RedirectResponse(f"/entries/{entry_id}/review", status_code=303)
        return RedirectResponse(f"/entries/{entry_id}/edit", status_code=303)


@app.get("/entries/{entry_id}/review", response_class=HTMLResponse)
def review_entry(request: Request, entry_id: int):
    with _session() as session:
        entry = _load_entry(session, entry_id)
        nonzero = [l for l in entry.lines if l.quantity != 0]
        voucher_preview = entry_to_voucher_dict(entry)
    return templates.TemplateResponse(
        request,
        "review.html",
        {
            "entry": entry,
            "lines": sorted(nonzero, key=lambda l: (l.role, l.item_name)),
            "voucher_preview": voucher_preview,
        },
    )


@app.post("/entries/{entry_id}/post")
def post_entry(entry_id: int):
    with _session() as session:
        entry = _load_entry(session, entry_id)
        if entry.status == "posted":
            return RedirectResponse(f"/entries/{entry_id}", status_code=303)

        nonzero = [l for l in entry.lines if l.quantity != 0]
        if not nonzero:
            raise HTTPException(400, "No non-zero lines to post")

        voucher_dict = entry_to_voucher_dict(entry)
        entry.status = "submitted"
        entry.submitted_at = datetime.utcnow()
        session.commit()

        try:
            client = _tally_client()
            result = client.import_voucher(entry.company_name, voucher_dict, dry_run=False)
        except Exception as exc:
            entry.status = "failed"
            entry.tally_error = f"{type(exc).__name__}: {exc}"
            session.commit()
            return RedirectResponse(f"/entries/{entry_id}", status_code=303)

        if result.get("ok"):
            entry.status = "posted"
            entry.posted_at = datetime.utcnow()
            vch_id = result.get("last_vch_id")
            entry.tally_master_id = str(vch_id) if vch_id is not None else None
            vch_no = None
            if vch_id is not None:
                try:
                    vch_no = client.fetch_voucher_number_by_master_id(entry.company_name, vch_id)
                except Exception:
                    vch_no = None
            entry.tally_voucher_number = vch_no or (str(vch_id) if vch_id is not None else None)
            entry.tally_error = None
        else:
            entry.status = "failed"
            parts = []
            if result.get("line_error"):
                parts.append(f"LINEERROR: {result['line_error']}")
            if result.get("exception"):
                parts.append(f"EXCEPTIONS: {result['exception']}")
            parts.append(
                f"created={result.get('created')} altered={result.get('altered')} "
                f"ignored={result.get('ignored')} errors={result.get('errors')}"
            )
            entry.tally_error = " | ".join(parts)
        session.commit()
        return RedirectResponse(f"/entries/{entry_id}", status_code=303)


@app.get("/entries/{entry_id}", response_class=HTMLResponse)
def view_entry(request: Request, entry_id: int):
    with _session() as session:
        entry = _load_entry(session, entry_id)
        nonzero = sorted(
            [l for l in entry.lines if l.quantity != 0],
            key=lambda l: (l.role, l.item_name),
        )
    return templates.TemplateResponse(
        request,
        "result.html",
        {"entry": entry, "lines": nonzero},
    )


# ---------------------------------------------------------------------------
# JSON API (consumed by React SPA at /app)
# ---------------------------------------------------------------------------

class ApiLineIn(BaseModel):
    role: str  # 'consume' | 'produce'
    item_name: str
    quantity: float
    uom: str | None = None
    rate: float | None = None
    godown: str | None = None
    narration: str | None = None


class ApiEntryIn(BaseModel):
    company: str
    entry_date: str
    voucher_type: str
    narration: str | None = None
    lines: list[ApiLineIn]


@app.get("/api/companies")
def api_companies():
    with _session() as session:
        return {"companies": _list_companies(session)}


@app.get("/api/voucher-types")
def api_voucher_types(company: str):
    with _session() as session:
        policies = list_policies(session, company)
    return {
        "voucher_types": [
            {
                "voucher_type": p.voucher_type,
                "description": p.description,
                "strict": p.strict,
                "rate_policy": p.rate_policy,
            }
            for p in policies
        ]
    }


@app.get("/api/groups")
def api_groups(company: str, voucher_type: str, role: str = "consume"):
    """Return policy-allowed groups for a voucher type as a recursive tree.

    Each node shape:
      { stock_group, items: [...all items in subtree...], children: [node, ...] }

    Items at every level are the items in that node's full subtree, so the UI
    can scope the [+ Add item] dropdown at any level it chooses.
    """
    from ..models import StockGroup
    with _session() as session:
        policy = get_policy(session, company, voucher_type)
        if policy is None:
            raise HTTPException(404, f"No policy for {voucher_type}")

        all_sg = list(
            session.scalars(select(StockGroup).where(StockGroup.company_name == company))
        )
        children_of: dict[str, list[str]] = {}
        for sg in all_sg:
            if sg.parent and sg.parent != sg.name:
                children_of.setdefault(sg.parent, []).append(sg.name)

        # Pre-build direct-items index: parent_group_name -> [StockItem, ...]
        direct_items: dict[str, list[StockItem]] = {}
        for it in session.scalars(
            select(StockItem)
            .where(StockItem.company_name == company)
            .order_by(StockItem.name)
        ):
            if it.parent:
                direct_items.setdefault(it.parent, []).append(it)

        def build(node: str) -> dict:
            kids = [build(c) for c in sorted(children_of.get(node, []))]
            # subtree items = direct items at this node + every descendant's items
            subtree = [_item_json(i) for i in direct_items.get(node, [])]
            for k in kids:
                subtree.extend(k["items"])
            return {
                "stock_group": node,
                "items": subtree,
                "children": kids,
            }

        role_rows = [g for g in policy.groups if g.role == role]
        groups_out = []
        for g in role_rows:
            tree = build(g.stock_group)
            groups_out.append(
                {
                    "stock_group": g.stock_group,
                    "default_godown": g.default_godown,
                    "tree": tree,
                }
            )
    return {"voucher_type": voucher_type, "role": role, "groups": groups_out}


def _item_json(it: StockItem) -> dict:
    return {
        "name": it.name,
        "uom": it.closing_uom or it.base_units or "No.",
        "rate": it.closing_rate or 0.0,
        "opening_stock": it.closing_quantity or 0.0,
        "parent_group": it.parent,
    }


@app.post("/api/entries")
def api_create_entry(payload: ApiEntryIn):
    with _session() as session:
        policy = get_policy(session, payload.company, payload.voucher_type)
        if policy is None:
            raise HTTPException(400, f"No policy for {payload.voucher_type}")
        if not payload.lines:
            raise HTTPException(400, "At least one line is required")

        entry = ProductionEntry(
            remote_id="pending",
            company_name=payload.company,
            entry_date=payload.entry_date,
            voucher_type=payload.voucher_type,
            status="draft",
            narration=payload.narration,
        )
        session.add(entry)
        session.flush()
        entry.remote_id = generate_remote_id(payload.entry_date, payload.voucher_type, entry.id)

        # Lookup stock items by name for defaults (rate, uom)
        by_name = {
            it.name: it
            for it in session.scalars(
                select(StockItem).where(StockItem.company_name == payload.company)
            )
        }
        for line in payload.lines:
            item = by_name.get(line.item_name)
            uom = line.uom or (item.closing_uom or item.base_units or "No." if item else "No.")
            rate = line.rate if line.rate is not None else (item.closing_rate if item else 0.0)
            amount = round(line.quantity * (rate or 0.0), 2)
            godown = line.godown or default_godown_for_role(policy, line.role)
            # Per-line narration is stored in the shared narration by appending; Tally voucher
            # only has a single narration, so we collect per-line notes into the header.
            session.add(
                ProductionEntryLine(
                    entry_id=entry.id,
                    role=line.role,
                    item_name=line.item_name,
                    quantity=line.quantity,
                    uom=uom,
                    rate=rate or 0.0,
                    amount=amount,
                    godown=godown,
                    opening_stock_snapshot=item.closing_quantity if item else 0.0,
                    description=line.narration or None,
                )
            )
        # Keep entry.narration as the voucher-level narration only.
        entry.narration = payload.narration or None

        session.commit()
        return _entry_to_json(entry)


@app.get("/api/entries")
def api_list_entries(
    company: str | None = None,
    entry_date: str | None = None,
    status: str | None = None,
    limit: int = 100,
):
    with _session() as session:
        stmt = select(ProductionEntry).order_by(ProductionEntry.created_at.desc())
        if company:
            stmt = stmt.where(ProductionEntry.company_name == company)
        if entry_date:
            stmt = stmt.where(ProductionEntry.entry_date == entry_date)
        if status:
            stmt = stmt.where(ProductionEntry.status == status)
        entries = list(session.scalars(stmt.limit(limit)))
        return {"entries": [_entry_to_json(e) for e in entries]}


@app.get("/api/entries/{entry_id}")
def api_get_entry(entry_id: int):
    with _session() as session:
        entry = _load_entry(session, entry_id)
        return _entry_to_json(entry)


@app.delete("/api/entries/{entry_id}")
def api_delete_entry(entry_id: int):
    """Delete a staged entry (draft / failed). Posted entries are not deletable
    from here — the voucher is already in Tally."""
    with _session() as session:
        entry = _load_entry(session, entry_id)
        if entry.status == "posted":
            raise HTTPException(
                400,
                "Cannot delete a posted entry — the voucher is already in Tally.",
            )
        session.delete(entry)
        session.commit()
        return {"ok": True, "deleted_id": entry_id}


@app.post("/api/entries/{entry_id}/post")
def api_post_entry(entry_id: int):
    with _session() as session:
        entry = _load_entry(session, entry_id)
        if entry.status == "posted":
            return _entry_to_json(entry)

        nonzero = [l for l in entry.lines if l.quantity != 0]
        if not nonzero:
            raise HTTPException(400, "No non-zero lines to post")

        voucher_dict = entry_to_voucher_dict(entry)
        entry.status = "submitted"
        entry.submitted_at = datetime.utcnow()
        session.commit()

        try:
            client = _tally_client()
            result = client.import_voucher(entry.company_name, voucher_dict, dry_run=False)
        except Exception as exc:
            entry.status = "failed"
            entry.tally_error = f"{type(exc).__name__}: {exc}"
            session.commit()
            return _entry_to_json(entry)

        if result.get("ok"):
            entry.status = "posted"
            entry.posted_at = datetime.utcnow()
            vch_id = result.get("last_vch_id")
            entry.tally_master_id = str(vch_id) if vch_id is not None else None
            vch_no = None
            if vch_id is not None:
                try:
                    vch_no = client.fetch_voucher_number_by_master_id(entry.company_name, vch_id)
                except Exception:
                    vch_no = None
            entry.tally_voucher_number = vch_no or (str(vch_id) if vch_id is not None else None)
            entry.tally_error = None
        else:
            entry.status = "failed"
            parts = []
            if result.get("line_error"):
                parts.append(f"LINEERROR: {result['line_error']}")
            if result.get("exception"):
                parts.append(f"EXCEPTIONS: {result['exception']}")
            parts.append(
                f"created={result.get('created')} altered={result.get('altered')} "
                f"ignored={result.get('ignored')} errors={result.get('errors')}"
            )
            entry.tally_error = " | ".join(parts)
        session.commit()
        return _entry_to_json(entry)


def _entry_to_json(entry: ProductionEntry) -> dict:
    return {
        "id": entry.id,
        "remote_id": entry.remote_id,
        "company": entry.company_name,
        "voucher_type": entry.voucher_type,
        "entry_date": entry.entry_date,
        "status": entry.status,
        "narration": entry.narration,
        "tally_voucher_number": entry.tally_voucher_number,
        "tally_master_id": entry.tally_master_id,
        "tally_error": entry.tally_error,
        "lines": [
            {
                "id": l.id,
                "role": l.role,
                "item_name": l.item_name,
                "quantity": l.quantity,
                "uom": l.uom,
                "rate": l.rate,
                "amount": l.amount,
                "godown": l.godown,
                "description": l.description,
            }
            for l in entry.lines
        ],
        "created_at": entry.created_at.isoformat() if entry.created_at else None,
        "posted_at": entry.posted_at.isoformat() if entry.posted_at else None,
    }


def resolve_items_for_role_single_group(session: Session, company: str, group_root: str) -> list[StockItem]:
    """Resolve items under a single group (walks descendants)."""
    from ..policy import _descendant_group_names
    expanded = _descendant_group_names(session, company, [group_root])
    return list(
        session.scalars(
            select(StockItem)
            .where(StockItem.company_name == company, StockItem.parent.in_(expanded))
            .order_by(StockItem.name)
        )
    )


# ---------------------------------------------------------------------------
# React SPA at /app
# ---------------------------------------------------------------------------

STATIC_DIR = Path(__file__).parent / "static"
if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


@app.get("/app", response_class=HTMLResponse)
def spa_react(request: Request):
    response = templates.TemplateResponse(request, "app.html")
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


@app.get("/entries", response_class=HTMLResponse)
def list_entries(
    request: Request,
    company: str | None = None,
    entry_date: str | None = None,
    summary_date: str | None = None,
):
    with _session() as session:
        stmt = select(ProductionEntry).order_by(ProductionEntry.created_at.desc())
        if company:
            stmt = stmt.where(ProductionEntry.company_name == company)
        if entry_date:
            stmt = stmt.where(ProductionEntry.entry_date == entry_date)
        entries = list(session.scalars(stmt))
        companies = _list_companies(session)

        # Distinct entry dates (across filters so sidebar reflects visible scope)
        dates_stmt = select(ProductionEntry.entry_date).distinct().order_by(ProductionEntry.entry_date.desc())
        if company:
            dates_stmt = dates_stmt.where(ProductionEntry.company_name == company)
        all_dates = [d for d in session.scalars(dates_stmt).all() if d]

        summary_companies: list[dict] = []
        summary_meta: dict = {}
        if summary_date:
            summary_companies, summary_meta = _build_summary(session, summary_date, company)

    return templates.TemplateResponse(
        request,
        "list.html",
        {
            "entries": entries,
            "companies": companies,
            "selected_company": company or "",
            "selected_date": entry_date or "",
            "all_dates": all_dates,
            "summary_date": summary_date or "",
            "summary_companies": summary_companies,
            "summary_meta": summary_meta,
        },
    )


def _logo_for_company(name: str) -> str | None:
    if not name:
        return None
    n = name.lower()
    if "avinash" in n:
        return "/static/logos/aapl.png"
    if "shanke" in n:
        return "/static/logos/sepl.png"
    return None


@app.get("/policies", response_class=HTMLResponse)
def policies_page(request: Request):
    """Serve the policy management SPA page."""
    response = templates.TemplateResponse(request, "policies.html")
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return response


@app.get("/api/all-companies")
def api_all_companies():
    """All companies in the DB, including those without policies (for the policy page)."""
    with _session() as session:
        return {"companies": list(session.scalars(select(Company.name).order_by(Company.name)))}


@app.get("/api/company-default-godown")
def api_get_company_default_godown(company: str):
    with _session() as session:
        c = session.scalar(select(Company).where(Company.name == company))
        return {"default_godown": getattr(c, "default_godown", None) if c else None}


class CompanyDefaultGodownPayload(BaseModel):
    company: str
    default_godown: str | None = None


@app.put("/api/company-default-godown")
def api_put_company_default_godown(payload: CompanyDefaultGodownPayload):
    with _session() as session:
        c = session.scalar(select(Company).where(Company.name == payload.company))
        if c is None:
            raise HTTPException(404, f"Company not found: {payload.company}")
        c.default_godown = payload.default_godown or None
        session.commit()
        return {"ok": True, "default_godown": c.default_godown}


@app.get("/api/tally-status")
def api_tally_status():
    """Quick Tally-connectivity probe used by the UI status indicator."""
    import time, re
    from ..config import get_settings
    settings = get_settings()
    # Use a tight timeout so the UI stays responsive.
    client = TallyClient(
        host=settings.tally_host,
        port=settings.tally_port,
        timeout=3,
        request_delay_ms=0,
        max_retries=0,
        retry_backoff_ms=0,
        lock_file=settings.tally_lock_file,
        lock_stale_seconds=settings.tally_lock_stale_seconds,
    )
    xml = (
        "<ENVELOPE><HEADER><VERSION>1</VERSION><TALLYREQUEST>Export</TALLYREQUEST>"
        "<TYPE>Collection</TYPE><ID>List of Companies</ID></HEADER><BODY><DESC>"
        "<STATICVARIABLES><SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT></STATICVARIABLES>"
        "<TDL><TDLMESSAGE><COLLECTION NAME=\"List of Companies\" ISMODIFY=\"No\">"
        "<TYPE>Company</TYPE><NATIVEMETHOD>Name</NATIVEMETHOD></COLLECTION>"
        "</TDLMESSAGE></TDL></DESC></BODY></ENVELOPE>"
    )
    started = time.monotonic()
    try:
        resp = client.post(xml)
    except Exception as exc:
        return {
            "ok": False,
            "host": f"{settings.tally_host}:{settings.tally_port}",
            "error": f"{type(exc).__name__}: {exc}",
            "latency_ms": int((time.monotonic() - started) * 1000),
            "companies": [],
        }
    latency = int((time.monotonic() - started) * 1000)
    companies = re.findall(r'<COMPANY[^>]*NAME="([^"]+)"', resp)
    return {
        "ok": True,
        "host": f"{settings.tally_host}:{settings.tally_port}",
        "latency_ms": latency,
        "companies": companies,
    }


@app.get("/api/godowns")
def api_godowns(company: str):
    from ..models import Godown
    with _session() as session:
        rows = session.scalars(
            select(Godown).where(Godown.company_name == company).order_by(Godown.name)
        )
        return {"godowns": [{"name": g.name, "parent": g.parent} for g in rows]}


@app.get("/api/policy-group-usage")
def api_policy_group_usage(company: str):
    """Returns { stock_group: [voucher_type, ...] } — groups already assigned
    to any policy in this company. Used by the policy editor to grey-out
    groups that belong to a different voucher type."""
    from ..models import SJPolicyGroup
    with _session() as session:
        rows = session.execute(
            select(SJPolicy.voucher_type, SJPolicyGroup.stock_group, SJPolicyGroup.role)
            .join(SJPolicyGroup, SJPolicyGroup.policy_id == SJPolicy.id)
            .where(SJPolicy.company_name == company)
        ).all()
        usage: dict[str, list[dict]] = {}
        for vt, sg, role in rows:
            usage.setdefault(sg, []).append({"voucher_type": vt, "role": role})
        return {"usage": usage}


@app.get("/api/tally-voucher-types")
def api_tally_voucher_types(company: str):
    """All voucher types synced from Tally for the company, with a flag indicating
    whether each already has a policy."""
    from ..models import VoucherType
    with _session() as session:
        vts = list(session.scalars(
            select(VoucherType).where(VoucherType.company_name == company).order_by(VoucherType.name)
        ))
        policy_names = set(session.scalars(
            select(SJPolicy.voucher_type).where(SJPolicy.company_name == company)
        ))
        return {
            "voucher_types": [
                {
                    "name": v.name,
                    "parent": v.parent,
                    "has_policy": v.name in policy_names,
                }
                for v in vts
            ]
        }


@app.get("/api/stock-group-tree")
def api_stock_group_tree(company: str):
    from ..models import StockGroup
    with _session() as session:
        all_sg = list(session.scalars(select(StockGroup).where(StockGroup.company_name == company)))
        children_of: dict[str, list[str]] = {}
        names = set()
        for sg in all_sg:
            names.add(sg.name)
            if sg.parent and sg.parent != sg.name:
                children_of.setdefault(sg.parent, []).append(sg.name)
        # Roots = groups whose parent is None or not in the set (e.g. "Primary")
        roots = sorted(sg.name for sg in all_sg if not sg.parent or sg.parent not in names)

        def build(name: str) -> dict:
            return {
                "name": name,
                "children": [build(c) for c in sorted(children_of.get(name, []))],
            }
        return {"tree": [build(r) for r in roots]}


class PolicyUpsertPayload(BaseModel):
    company: str
    voucher_type: str
    description: str | None = None
    strict: bool = True
    rate_policy: str = "stock_master"
    active: bool = True
    groups: list[dict]  # [{stock_group, role, default_godown?}]


@app.get("/api/policy")
def api_get_policy(company: str, voucher_type: str):
    with _session() as session:
        p = get_policy(session, company, voucher_type)
        if p is None:
            return {"exists": False, "voucher_type": voucher_type, "groups": []}
        return {
            "exists": True,
            "company": p.company_name,
            "voucher_type": p.voucher_type,
            "description": p.description,
            "strict": p.strict,
            "rate_policy": p.rate_policy,
            "active": p.active,
            "groups": [
                {"stock_group": g.stock_group, "role": g.role, "default_godown": g.default_godown}
                for g in p.groups
            ],
        }


@app.put("/api/policy")
def api_put_policy(payload: PolicyUpsertPayload):
    from ..models import SJPolicy, SJPolicyGroup
    with _session() as session:
        p = session.scalar(
            select(SJPolicy).where(
                SJPolicy.company_name == payload.company,
                SJPolicy.voucher_type == payload.voucher_type,
            )
        )
        if p is None:
            p = SJPolicy(company_name=payload.company, voucher_type=payload.voucher_type)
            session.add(p)
            session.flush()
        p.description = payload.description
        p.strict = payload.strict
        p.rate_policy = payload.rate_policy
        p.active = payload.active
        for row in list(p.groups):
            session.delete(row)
        session.flush()
        for g in payload.groups:
            sg = (g.get("stock_group") or "").strip()
            role = (g.get("role") or "").strip()
            if not sg or role not in ("consume", "produce"):
                continue
            session.add(SJPolicyGroup(
                policy_id=p.id,
                stock_group=sg,
                role=role,
                default_godown=(g.get("default_godown") or None),
            ))
        session.commit()
        return {"ok": True, "voucher_type": p.voucher_type, "group_count": len(payload.groups)}


@app.delete("/api/policy")
def api_delete_policy(company: str, voucher_type: str):
    from ..models import SJPolicy
    with _session() as session:
        p = session.scalar(
            select(SJPolicy).where(
                SJPolicy.company_name == company,
                SJPolicy.voucher_type == voucher_type,
            )
        )
        if p is None:
            raise HTTPException(404, "Policy not found")
        session.delete(p)
        session.commit()
        return {"ok": True}


@app.get("/summary/print", response_class=HTMLResponse)
def print_summary(request: Request, summary_date: str, company: str | None = None):
    with _session() as session:
        companies, meta = _build_summary(session, summary_date, company)
    return templates.TemplateResponse(
        request,
        "print_summary.html",
        {
            "summary_date": summary_date,
            "companies": companies,
            "meta": meta,
            "logo_left": "/static/logos/aapl.png",
            "logo_right": "/static/logos/sepl.png",
            "generated_at": datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"),
        },
    )


# ---------------------------------------------------------------------------
# Daily stock consumption pivot
# ---------------------------------------------------------------------------

def _get_selection(session: Session, company: str) -> dict[str, set[str]]:
    rows = session.scalars(
        select(ConsumptionReportSelection).where(
            ConsumptionReportSelection.company_name == company
        )
    )
    sel: dict[str, set[str]] = {"group": set(), "item": set(), "vtype": set()}
    for r in rows:
        sel.setdefault(r.kind, set()).add(r.name)
    return sel


def _resolve_items_for_selection(
    session: Session, company: str, sel: dict[str, set[str]]
) -> list[StockItem]:
    """Expand selected groups (+ descendants) to concrete items and union with explicit items."""
    if not sel["group"] and not sel["item"]:
        return []

    # Walk group descendants
    all_sg = list(session.scalars(select(StockGroup).where(StockGroup.company_name == company)))
    children_of: dict[str, list[str]] = {}
    for sg in all_sg:
        if sg.parent and sg.parent != sg.name:
            children_of.setdefault(sg.parent, []).append(sg.name)
    expanded_groups: set[str] = set()
    stack = list(sel["group"])
    while stack:
        g = stack.pop()
        if g in expanded_groups:
            continue
        expanded_groups.add(g)
        stack.extend(children_of.get(g, []))

    stmt = select(StockItem).where(StockItem.company_name == company)
    items = list(session.scalars(stmt.order_by(StockItem.name)))
    chosen: list[StockItem] = []
    for it in items:
        if it.name in sel["item"] or (it.parent and it.parent in expanded_groups):
            chosen.append(it)
    return chosen


def _daterange(start: str, end: str) -> list[str]:
    from datetime import date as _d, timedelta
    s = datetime.strptime(start, "%Y-%m-%d").date()
    e = datetime.strptime(end, "%Y-%m-%d").date()
    if e < s:
        s, e = e, s
    out = []
    cur = s
    while cur <= e:
        out.append(cur.isoformat())
        cur += timedelta(days=1)
    return out


def _build_consumption_pivot(
    session: Session,
    company: str,
    start: str,
    end: str,
    voucher_types: list[str],
    items: list[StockItem],
) -> dict:
    dates = _daterange(start, end)
    # dd-mm label + short weekday + Sunday flag for table headers.
    _wd = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    date_headers = []
    for d in dates:
        dt = datetime.strptime(d, "%Y-%m-%d").date()
        date_headers.append({
            "iso": d,
            "label": dt.strftime("%d-%m"),
            "weekday": _wd[dt.weekday()],
            "is_sunday": dt.weekday() == 6,
        })
    item_names = [it.name for it in items]
    # name -> {date -> qty, uom, parent, total}
    by_item: dict[str, dict] = {
        it.name: {
            "name": it.name,
            "uom": it.closing_uom or it.base_units or "",
            "parent": it.parent or "",
            "by_date": {d: 0.0 for d in dates},
            "total": 0.0,
        }
        for it in items
    }
    if not item_names or not voucher_types:
        return {
            "dates": dates,
            "date_headers": date_headers,
            "by_item": by_item,
            "rows": _tree_rows(session, company, by_item, dates),
            "column_totals": {d: 0.0 for d in dates},
            "grand_total": 0.0,
        }

    q = (
        select(Voucher.voucher_date, VoucherInventoryEntry.item_name, VoucherInventoryEntry.quantity)
        .join(VoucherInventoryEntry, VoucherInventoryEntry.voucher_id == Voucher.id)
        .where(
            Voucher.company_name == company,
            Voucher.is_cancelled == False,  # noqa: E712
            Voucher.voucher_type_name.in_(voucher_types),
            Voucher.voucher_date >= start,
            Voucher.voucher_date <= end,
            VoucherInventoryEntry.item_name.in_(item_names),
        )
    )
    column_totals = {d: 0.0 for d in dates}
    grand_total = 0.0
    for vdate, item_name, qty in session.execute(q):
        if not vdate or vdate not in column_totals:
            continue
        row = by_item.get(item_name)
        if row is None:
            continue
        # Consumption is the outward flow; Tally stores consume as negative qty
        # in Stock Journal (DEEMEDPOSITIVE="No"). Display absolute quantity.
        amt = abs(float(qty or 0))
        row["by_date"][vdate] = row["by_date"].get(vdate, 0.0) + amt
        row["total"] += amt
        column_totals[vdate] += amt
        grand_total += amt

    tree_rows = _tree_rows(session, company, by_item, dates)
    return {
        "dates": dates,
        "date_headers": date_headers,
        "by_item": by_item,
        "rows": tree_rows,
        "column_totals": column_totals,
        "grand_total": grand_total,
    }


def _tree_rows(
    session: Session,
    company: str,
    by_item: dict[str, dict],
    dates: list[str],
) -> list[dict]:
    """Render the selection as a hierarchical, depth-ordered row list.

    Walks StockGroup parent → child relationships for the company, so the
    table mirrors the Tally stock-group tree (root → sub-groups → items),
    not a flat alphabetical bucket. Each group row aggregates subtotals over
    *all* descendant items, not just its direct children. Groups with no
    selected descendants are omitted. Items whose parent is not a known
    stock group fall under a synthetic `(ungrouped)` node."""
    all_sg = list(session.scalars(select(StockGroup).where(StockGroup.company_name == company))) if company else []
    group_names = {sg.name for sg in all_sg}
    children_of: dict[str, list[str]] = {}
    for sg in all_sg:
        if sg.parent and sg.parent in group_names and sg.parent != sg.name:
            children_of.setdefault(sg.parent, []).append(sg.name)
    roots = sorted(
        sg.name for sg in all_sg
        if not sg.parent or sg.parent not in group_names or sg.parent == sg.name
    )

    items_by_group: dict[str, list[dict]] = {}
    for row in by_item.values():
        key = row["parent"] if (row["parent"] and row["parent"] in group_names) else "(ungrouped)"
        items_by_group.setdefault(key, []).append(row)

    def descendants_have_items(g: str) -> bool:
        if items_by_group.get(g):
            return True
        return any(descendants_have_items(c) for c in children_of.get(g, []))

    def accumulate(g: str, bucket: dict[str, float]) -> float:
        total = 0.0
        for r in items_by_group.get(g, []):
            for d, q in r["by_date"].items():
                bucket[d] = bucket.get(d, 0.0) + q
            total += r["total"]
        for c in children_of.get(g, []):
            total += accumulate(c, bucket)
        return total

    out: list[dict] = []
    counter = [0]

    def walk(name: str, depth: int, ancestor_ids: list[str]) -> None:
        if not descendants_have_items(name):
            return
        counter[0] += 1
        gid = f"g{counter[0]}"
        subtotal_by_date: dict[str, float] = {d: 0.0 for d in dates}
        subtotal = accumulate(name, subtotal_by_date)
        out.append({
            "kind": "group",
            "id": gid,
            "name": name,
            "depth": depth,
            "ancestors": list(ancestor_ids),
            "subtotal_by_date": subtotal_by_date,
            "subtotal": subtotal,
        })
        new_ancestors = ancestor_ids + [gid]
        for it in sorted(items_by_group.get(name, []), key=lambda r: r["name"]):
            counter[0] += 1
            out.append({
                "kind": "item",
                "id": f"i{counter[0]}",
                "name": it["name"],
                "uom": it["uom"],
                "depth": depth + 1,
                "ancestors": new_ancestors,
                "by_date": it["by_date"],
                "total": it["total"],
            })
        for c in sorted(children_of.get(name, [])):
            walk(c, depth + 1, new_ancestors)

    for r in roots:
        walk(r, 0, [])
    if items_by_group.get("(ungrouped)"):
        walk("(ungrouped)", 0, [])
    return out


def _parse_csv(raw: str | None) -> list[str]:
    if not raw:
        return []
    return [x.strip() for x in raw.split(",") if x.strip()]


@app.get("/consumption", response_class=HTMLResponse)
def consumption_report(
    request: Request,
    company: str | None = None,
    start: str | None = None,
    end: str | None = None,
    voucher_types: list[str] | None = Query(None),
    configure: int = 0,
):
    from datetime import date as _d
    with _session() as session:
        companies = sorted(
            c for c in session.scalars(select(Voucher.company_name).distinct()) if c
        )
        # Fall back to all companies if none have vouchers yet
        if not companies:
            companies = _list_companies(session)
        company = company or (companies[0] if companies else "")

        today = _d.today()
        if not start:
            start = today.replace(day=1).isoformat()
        if not end:
            end = today.isoformat()

        # Voucher-type universe for this company, grouped by Tally base type.
        # We only keep (base, voucher_type_name) pairs that actually carry
        # inventory entries — strips Payment / Receipt / Journal / Contra / etc.
        base_alias_rows: list[tuple[str | None, str]] = []
        if company:
            base_alias_rows = list(session.execute(
                select(Voucher.base_voucher_type, Voucher.voucher_type_name)
                .join(VoucherInventoryEntry, VoucherInventoryEntry.voucher_id == Voucher.id)
                .where(Voucher.company_name == company)
                .distinct()
            ).all())

        def _canonical_base(base: str | None, name: str | None) -> str:
            """Re-classify user-defined voucher types that didn't resolve to a
            Tally default. Some companies create types like 'SJ - DRILL' whose
            stored base is the same user-defined name (because the parent chain
            wasn't fully synced); we still want them listed under 'Stock Journal'."""
            b = (base or "").strip()
            n = (name or "").strip()
            if b in CONSUMPTION_DEFAULT_VOUCHER_TYPES:
                return b
            nl = n.lower()
            bl = b.lower()
            # Stock Journal family
            if (
                nl.startswith("sj ") or nl.startswith("sj-") or nl.startswith("sj ")
                or nl.startswith("sj_") or "stock journal" in nl
                or "manufacturing journal" in nl
                or bl.startswith("sj") or "stock journal" in bl
            ):
                return "Stock Journal"
            if "delivery note" in nl or nl.startswith("dn ") or nl.startswith("dn-") or "delivery note" in bl:
                return "Delivery Note"
            if "physical stock" in nl or "physical stock" in bl:
                return "Physical Stock"
            if "rejections out" in nl or "rejections out" in bl:
                return "Rejections Out"
            if "material out" in nl or "material out" in bl:
                return "Material Out"
            return b or "(unclassified)"

        aliases_by_base: dict[str, list[str]] = {}
        for base, name in base_alias_rows:
            if not name:
                continue
            key = _canonical_base(base, name)
            aliases_by_base.setdefault(key, []).append(name)
        for k in aliases_by_base:
            aliases_by_base[k] = sorted(set(aliases_by_base[k]))

        # Order bases: Tally defaults first (in preferred order), then any others.
        base_groups: list[dict] = []
        seen_bases: set[str] = set()
        for base in CONSUMPTION_DEFAULT_VOUCHER_TYPES:
            if base in aliases_by_base:
                base_groups.append({
                    "base": base,
                    "aliases": aliases_by_base[base],
                    "is_tally_default": True,
                })
                seen_bases.add(base)
        for base in sorted(aliases_by_base.keys()):
            if base in seen_bases:
                continue
            base_groups.append({
                "base": base,
                "aliases": aliases_by_base[base],
                "is_tally_default": False,
            })

        all_voucher_types = sorted({a for group in base_groups for a in group["aliases"]})

        sel = _get_selection(session, company) if company else {"group": set(), "item": set(), "vtype": set()}

        if voucher_types is not None:
            # URL override (e.g. shareable link) wins.
            vtypes = [vt for vt in voucher_types if vt]
        elif sel["vtype"]:
            # Persisted selection from the Configure Items screen.
            vtypes = [vt for vt in sorted(sel["vtype"]) if vt in all_voucher_types]
        else:
            # First-ever load for this company: auto-tick every alias whose base
            # is a Tally default outward-stock voucher type.
            vtypes = [
                a for group in base_groups
                if group["is_tally_default"]
                for a in group["aliases"]
            ]
            if not vtypes:
                vtypes = list(all_voucher_types)

        # Data for configure panel: full stock group tree + direct items
        groups_tree: list[dict] = []
        direct_items: dict[str, list[str]] = {}
        expanded_groups: set[str] = set()
        if configure and company:
            all_sg = list(session.scalars(select(StockGroup).where(StockGroup.company_name == company)))
            group_names = {sg.name for sg in all_sg}
            children_of: dict[str, list[str]] = {}
            parent_of: dict[str, str] = {}
            for sg in all_sg:
                if sg.parent and sg.parent != sg.name:
                    children_of.setdefault(sg.parent, []).append(sg.name)
                    parent_of[sg.name] = sg.parent
            roots = sorted(sg.name for sg in all_sg if not sg.parent or sg.parent not in group_names)
            for it in session.scalars(
                select(StockItem).where(StockItem.company_name == company).order_by(StockItem.name)
            ):
                if it.parent:
                    direct_items.setdefault(it.parent, []).append(it.name)

            def build(n: str) -> dict:
                return {
                    "name": n,
                    "children": [build(c) for c in sorted(children_of.get(n, []))],
                    "items": direct_items.get(n, []),
                }
            groups_tree = [build(r) for r in roots]
            # Tree starts fully collapsed — keeps the Configure Items screen
            # compact. Users can use Expand all / Collapse all or click rows.

        items = _resolve_items_for_selection(session, company, sel) if company else []
        pivot = _build_consumption_pivot(session, company, start, end, vtypes, items)

    return templates.TemplateResponse(
        request,
        "consumption.html",
        {
            "companies": companies,
            "selected_company": company,
            "start": start,
            "end": end,
            "voucher_types": vtypes,
            "voucher_types_csv": ",".join(vtypes),
            "all_voucher_types": all_voucher_types,
            "base_groups": base_groups,
            "selected_vtypes_set": set(vtypes),
            "pivot": pivot,
            "configure": bool(configure),
            "groups_tree": groups_tree,
            "selected_groups": sorted(sel["group"]),
            "selected_items": sorted(sel["item"]),
            "expanded_groups": expanded_groups,
            "selection_count": len(sel["group"]) + len(sel["item"]),
        },
    )


@app.post("/consumption/selection")
async def save_consumption_selection(request: Request):
    from urllib.parse import urlencode
    form = await request.form()
    company = (form.get("company") or "").strip()
    if not company:
        raise HTTPException(400, "company required")
    selected_groups = [v for v in form.getlist("group") if v]
    selected_items = [v for v in form.getlist("item") if v]
    selected_vtypes = [v for v in form.getlist("voucher_types") if v]
    with _session() as session:
        # Replace selection for this company
        for row in session.scalars(
            select(ConsumptionReportSelection).where(
                ConsumptionReportSelection.company_name == company
            )
        ):
            session.delete(row)
        session.flush()
        for g in selected_groups:
            session.add(ConsumptionReportSelection(company_name=company, kind="group", name=g))
        for it in selected_items:
            session.add(ConsumptionReportSelection(company_name=company, kind="item", name=it))
        for vt in selected_vtypes:
            session.add(ConsumptionReportSelection(company_name=company, kind="vtype", name=vt))
        session.commit()
    # Redirect back to the main (non-configure) screen so the user sees the
    # report with the new selection applied.
    params: dict[str, str] = {"company": company}
    for k in ("start", "end"):
        v = form.get(k)
        if v:
            params[k] = v
    return RedirectResponse(f"/consumption?{urlencode(params)}", status_code=303)


@app.get("/consumption/export.xlsx")
def consumption_export_xlsx(
    company: str | None = None,
    start: str | None = None,
    end: str | None = None,
):
    """Excel export of the daily stock-consumption pivot.

    Formula-enabled: per-item row totals and per-date column totals are live
    SUM() formulas, and group subtotal rows sum their descendant item cells —
    so edits in Excel recalculate automatically. Uses the same selection +
    voucher-type config persisted for the company.
    """
    import io
    from datetime import date as _d
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import Response

    with _session() as session:
        if not company:
            companies = sorted(
                c for c in session.scalars(select(Voucher.company_name).distinct()) if c
            )
            company = companies[0] if companies else ""
        today = _d.today()
        start = start or today.replace(day=1).isoformat()
        end = end or today.isoformat()
        sel = _get_selection(session, company) if company else {"group": set(), "item": set(), "vtype": set()}
        vtypes = sorted(sel["vtype"]) if sel["vtype"] else []
        items = _resolve_items_for_selection(session, company, sel) if company else []
        pivot = _build_consumption_pivot(session, company, start, end, vtypes, items)

    wb = Workbook()
    ws = wb.active
    ws.title = "Consumption"

    headers = pivot["date_headers"]
    n_dates = len(headers)
    first_date_col = 3  # A=Item, B=UOM, C..=dates, last=Total
    last_date_col = first_date_col + n_dates - 1
    total_col = last_date_col + 1

    navy = PatternFill("solid", fgColor="E8EEF9")
    group_fill = PatternFill("solid", fgColor="FCE7C4")
    sunday_fill = PatternFill("solid", fgColor="FFF7ED")
    bold = Font(bold=True)
    thin = Side(style="thin", color="D5DBE5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center")
    right = Alignment(horizontal="right")

    ws.cell(row=1, column=1, value=f"Daily Stock Consumption — {company}").font = Font(bold=True, size=13)
    ws.cell(row=2, column=1, value=f"{start} to {end}").font = Font(italic=True, color="666666")

    hdr_row1 = 4
    hdr_row2 = 5
    ws.cell(row=hdr_row1, column=1, value="Item name").font = bold
    ws.cell(row=hdr_row2, column=1, value="").font = bold
    ws.merge_cells(start_row=hdr_row1, start_column=1, end_row=hdr_row2, end_column=1)
    ws.cell(row=hdr_row1, column=2, value="UOM").font = bold
    ws.merge_cells(start_row=hdr_row1, start_column=2, end_row=hdr_row2, end_column=2)
    for i, h in enumerate(headers):
        c1 = ws.cell(row=hdr_row1, column=first_date_col + i, value=h["weekday"])
        c2 = ws.cell(row=hdr_row2, column=first_date_col + i, value=h["label"])
        for c in (c1, c2):
            c.font = bold
            c.alignment = center
            c.fill = sunday_fill if h["is_sunday"] else navy
    ws.cell(row=hdr_row1, column=total_col, value="Total").font = bold
    ws.merge_cells(start_row=hdr_row1, start_column=total_col, end_row=hdr_row2, end_column=total_col)

    # Emit data rows. Track per-group the row range of its *direct* item rows;
    # group subtotals use SUMIF-style formulas that sum over the direct-descendant
    # item rows in each date column, so editing an item cell updates the group row.
    r = hdr_row2 + 1
    group_item_rows: dict[str, list[int]] = {}  # group_id -> list of item excel row numbers under it (incl descendants)
    group_row_num: dict[str, int] = {}

    for row in pivot["rows"]:
        if row["kind"] == "group":
            group_row_num[row["id"]] = r
            indent = "    " * row["depth"]
            cell = ws.cell(row=r, column=1, value=f"{indent}📁 {row['name']}")
            cell.font = bold
            ws.cell(row=r, column=2, value="—").alignment = center
            for i, h in enumerate(headers):
                c = ws.cell(row=r, column=first_date_col + i)
                c.font = bold
                c.alignment = right
                if h["is_sunday"]:
                    c.fill = sunday_fill
                else:
                    c.fill = group_fill
            ws.cell(row=r, column=total_col).font = bold
            # Fill color for item-less cells (still mark the row)
            ws.cell(row=r, column=1).fill = group_fill
            ws.cell(row=r, column=2).fill = group_fill
            ws.cell(row=r, column=total_col).fill = group_fill
            r += 1
        else:
            indent = "    " * row["depth"]
            ws.cell(row=r, column=1, value=f"{indent}{row['name']}")
            ws.cell(row=r, column=2, value=row["uom"] or "")
            for i, h in enumerate(headers):
                v = row["by_date"].get(h["iso"], 0)
                cell = ws.cell(row=r, column=first_date_col + i, value=(round(v, 4) if v else None))
                cell.number_format = "#,##0.00;-#,##0.00;—"
                cell.alignment = right
                if h["is_sunday"]:
                    cell.fill = sunday_fill
            # Row total as a live SUM formula across date columns
            first_letter = get_column_letter(first_date_col)
            last_letter = get_column_letter(last_date_col)
            tot = ws.cell(row=r, column=total_col, value=f"=SUM({first_letter}{r}:{last_letter}{r})")
            tot.number_format = "#,##0.00;-#,##0.00;—"
            tot.alignment = right
            tot.font = bold
            for gid in row["ancestors"]:
                group_item_rows.setdefault(gid, []).append(r)
            r += 1

    # Fill in group subtotal formulas now that we know which item rows belong to each group.
    for gid, grow in group_row_num.items():
        item_rows = group_item_rows.get(gid, [])
        for i in range(n_dates):
            col = first_date_col + i
            col_letter = get_column_letter(col)
            if item_rows:
                # Build refs like C7,C8,C12 — explicit non-contiguous cell list
                refs = ",".join(f"{col_letter}{ir}" for ir in item_rows)
                ws.cell(row=grow, column=col, value=f"=SUM({refs})")
            else:
                ws.cell(row=grow, column=col, value=0)
            ws.cell(row=grow, column=col).number_format = "#,##0.00;-#,##0.00;—"
        # Group row total = SUM of its date cells on the same row
        first_letter = get_column_letter(first_date_col)
        last_letter = get_column_letter(last_date_col)
        tc = ws.cell(row=grow, column=total_col, value=f"=SUM({first_letter}{grow}:{last_letter}{grow})")
        tc.number_format = "#,##0.00;-#,##0.00;—"

    # Column totals row — live SUM over every row in each date column (groups are
    # omitted to avoid double counting; we sum only item rows).
    if r > hdr_row2 + 1:
        total_row = r + 1
        ws.cell(row=total_row, column=1, value="Column total").font = bold
        ws.cell(row=total_row, column=1).fill = navy
        ws.cell(row=total_row, column=2).fill = navy
        all_item_rows = sorted({ir for rows in group_item_rows.values() for ir in rows})
        for i in range(n_dates):
            col = first_date_col + i
            col_letter = get_column_letter(col)
            if all_item_rows:
                refs = ",".join(f"{col_letter}{ir}" for ir in all_item_rows)
                cell = ws.cell(row=total_row, column=col, value=f"=SUM({refs})")
            else:
                cell = ws.cell(row=total_row, column=col, value=0)
            cell.font = bold
            cell.alignment = right
            cell.fill = sunday_fill if headers[i]["is_sunday"] else navy
            cell.number_format = "#,##0.00;-#,##0.00;—"
        first_letter = get_column_letter(first_date_col)
        last_letter = get_column_letter(last_date_col)
        grand = ws.cell(
            row=total_row, column=total_col,
            value=f"=SUM({first_letter}{total_row}:{last_letter}{total_row})",
        )
        grand.font = bold
        grand.fill = navy
        grand.number_format = "#,##0.00;-#,##0.00;—"
        grand.alignment = right

    # Column widths + freeze panes for usability
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 10
    for i in range(n_dates):
        ws.column_dimensions[get_column_letter(first_date_col + i)].width = 10
    ws.column_dimensions[get_column_letter(total_col)].width = 12
    ws.freeze_panes = ws.cell(row=hdr_row2 + 1, column=first_date_col)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_co = "".join(c if c.isalnum() else "_" for c in (company or "company"))
    filename = f"consumption_{safe_co}_{start}_to_{end}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Daily Production Report
# ---------------------------------------------------------------------------
from dataclasses import dataclass as _dpr_dataclass


@_dpr_dataclass(frozen=True)
class _DprRow:
    """Lightweight row adapter — DPR templates only read `.key` and `.label`,
    so this stands in for the legacy static `daily_report.Row` while pulling
    from `ProductionProcess` instead. `key` is the process_id as a string
    (the value persisted in `dpr_hourly_cells.row_key`)."""
    key: str
    label: str
    stage: str
    group: str | None = None
    role: str | None = None  # 'input'|'output' for production rows; None elsewhere
    validate_count: bool = True


def _dpr_rows_for_report(
    session: Session, report: DailyProductionReport
) -> tuple[list[_DprRow], list[tuple[str, list[_DprRow]]], list[_DprRow]]:
    """Load this report's section row catalog from ProductionProcess.

    Returns (production_rows, rejection_groups, rework_rows) where
    rejection_groups preserves on-form sub-headings as
    [(group_label, [row, ...]), ...]. Empty lists if no config exists for
    the report's (company, line)."""
    pp_rows = list(session.scalars(
        select(ProductionProcess)
        .where(
            ProductionProcess.company_name == report.company_name,
            ProductionProcess.line == report.line,
            ProductionProcess.active.is_(True),
        )
        .order_by(ProductionProcess.section, ProductionProcess.sort_order, ProductionProcess.id)
    ))
    by_section: dict[str, list[_DprRow]] = {"production": [], "rejection": [], "rework": []}
    for r in pp_rows:
        if r.section not in by_section:
            continue
        by_section[r.section].append(_DprRow(
            key=str(r.id), label=r.label, stage=r.stage, group=r.group_label,
            role=r.role,
            validate_count=bool(getattr(r, "validate_count", True)),
        ))
    grouped: list[tuple[str, list[_DprRow]]] = []
    seen: dict[str, int] = {}
    for r in by_section["rejection"]:
        g = r.group or ""
        if g not in seen:
            seen[g] = len(grouped)
            grouped.append((g, []))
        grouped[seen[g]][1].append(r)
    return by_section["production"], grouped, by_section["rework"]


def _dpr_empty_matrix_dynamic(
    sections_rows: dict[str, list[_DprRow]],
    hour_slots: list[tuple[str, str]],
) -> dict[str, dict[str, dict[str, float]]]:
    hour_keys = [dpr.CUMULATIVE_KEY] + [h[0] for h in hour_slots]
    return {
        sec: {r.key: {hk: 0.0 for hk in hour_keys} for r in rows}
        for sec, rows in sections_rows.items()
    }


def _dpr_load_cells(
    session: Session,
    report: DailyProductionReport,
    hour_slots: list[tuple[str, str]],
    sections_rows: dict[str, list[_DprRow]] | None = None,
) -> dict[str, dict[str, dict[str, float]]]:
    """Return a fully-populated matrix `{section: {row_key: {hour_key: value}}}`.

    Starts from an empty matrix so missing DB cells render as 0 and overlays
    whatever cells the report has persisted. Templates can iterate without
    key-missing guards."""
    if sections_rows is None:
        # Legacy / fallback path — use static daily_report layout.
        matrix = dpr.empty_cells_matrix(hour_slots)
    else:
        matrix = _dpr_empty_matrix_dynamic(sections_rows, hour_slots)
    for cell in report.cells:
        section = matrix.get(cell.section)
        if section is None:
            continue
        row = section.get(cell.row_key)
        if row is None:
            continue
        if cell.hour_key in row:
            row[cell.hour_key] = cell.value
    return matrix


def _dpr_load_hour_models(report: DailyProductionReport) -> dict[str, str]:
    """Return `{hour_key: model}` for hours whose model differs from the
    report's primary model. Hours absent from the dict use `report.model`."""
    return {hm.hour_key: hm.model for hm in report.hour_models if hm.model}


def _dpr_resolve_hour_model(hour_models: dict[str, str], primary: str, hour_key: str) -> str:
    return (hour_models.get(hour_key) or primary or "").strip()


def _dpr_resolve_hour_model_chain(
    hour_models: dict[str, str],
    primary: str,
    hour_slots: list[tuple[str, str]],
) -> dict[str, str]:
    """Walk hour columns left→right, propagating each override forward
    until the next change point. Returns `{hour_key: effective_model}`.

    A model entered at hour N is the model running for hours N, N+1, ...
    until another override appears."""
    chain: dict[str, str] = {}
    current = (primary or "").strip()
    for key, _label in hour_slots:
        override = (hour_models.get(key) or "").strip()
        if override:
            current = override
        chain[key] = current
    return chain


def _dpr_production_by_model(
    cells: dict[str, dict[str, dict[str, float]]],
    hour_slots: list[tuple[str, str]],
    production_rows: list[_DprRow],
    hour_models: dict[str, str],
    primary_model: str,
) -> list[dict]:
    """Sum production output per model across hour columns.

    Picks the row identified as 'output' role per stage (or falls back to the
    last production row if no role flag is set). Returns
    `[{"model": str, "qty": float, "hours": [labels]}]` ordered by appearance."""
    output_keys = [r.key for r in production_rows if r.role == "output"] or [r.key for r in production_rows]
    chain = _dpr_resolve_hour_model_chain(hour_models, primary_model, hour_slots)
    by_model: dict[str, dict] = {}
    order: list[str] = []
    for hkey, hlabel in hour_slots:
        model = chain.get(hkey, "")
        if not model:
            continue
        qty = 0.0
        for rkey in output_keys:
            qty += cells.get("production", {}).get(rkey, {}).get(hkey, 0.0) or 0.0
        if model not in by_model:
            by_model[model] = {"model": model, "qty": 0.0, "hours": []}
            order.append(model)
        by_model[model]["qty"] += qty
        by_model[model]["hours"].append(hlabel)
    return [by_model[m] for m in order]


def _dpr_compute_totals(
    cells: dict[str, dict[str, dict[str, float]]],
    hour_slots: list[tuple[str, str]],
    production_rows: list[_DprRow] | None = None,
) -> dict:
    """Aggregate totals the summary page and KPI bar need in one pass."""
    prod = cells["production"]
    rej = cells["rejection"]
    rw = cells["rework"]
    prod_col = dpr.column_totals(prod, hour_slots)
    rej_col = dpr.column_totals(rej, hour_slots)
    rw_col = dpr.column_totals(rw, hour_slots)
    prod_rows = dpr.row_totals(prod)
    rej_rows = dpr.row_totals(rej)
    rw_rows = dpr.row_totals(rw)
    # "Output" benchmark for rejection/rework % is the sum of all production
    # rows tagged stage=washing (the last process stage, including rework).
    # If unset or zero, fall back to the largest row total so the percentages
    # stay meaningful during a partial shift.
    washing = 0.0
    if production_rows:
        washing = sum(prod_rows.get(r.key, 0.0) for r in production_rows if r.stage == "washing")
    else:
        washing = prod_rows.get("washing_incl_rework", 0.0)
    if washing <= 0:
        washing = max(prod_rows.values()) if prod_rows else 0.0
    # Line-input total drives the rejection/rework % denominators. Prefer the
    # explicit role='input' production row(s); fall back to the largest row
    # total so percentages stay meaningful before role flags are filled in.
    line_input = 0.0
    if production_rows:
        line_input = sum(
            prod_rows.get(r.key, 0.0)
            for r in production_rows if r.role == "input"
        )
    if line_input <= 0:
        line_input = max(prod_rows.values()) if prod_rows else 0.0
    total_rej = dpr.grand_total(rej)
    total_rw = dpr.grand_total(rw)
    return {
        "production": {"rows": prod_rows, "cols": prod_col, "total": dpr.grand_total(prod)},
        "rejection": {"rows": rej_rows, "cols": rej_col, "total": total_rej},
        "rework": {"rows": rw_rows, "cols": rw_col, "total": total_rw},
        "output_for_pct": washing,
        "input_for_pct": line_input,
        "rejection_pct": dpr.safe_pct(total_rej, line_input),
        "rework_pct": dpr.safe_pct(total_rw, line_input),
    }


def _review_clock_display(from_t: str | None, to_t: str | None, label: str) -> str:
    """Friendly cutoff label for the review dropdown. Prefer the slot's `to_time`
    (the end of the interval — what the user picks "until"). Fall back to
    `from_time`, then the legacy paper label with a leading '<'."""
    def fmt(hhmm: str) -> str:
        try:
            h, mm = hhmm.split(":")
            hi = int(h)
        except (ValueError, AttributeError):
            return hhmm
        ampm = "AM" if hi < 12 else "PM"
        h12 = hi % 12 or 12
        return f"{h12}:{mm} {ampm}"
    if to_t:
        return fmt(to_t)
    if from_t:
        return fmt(from_t)
    return f"< {label}"


@app.get("/production/review", response_class=HTMLResponse)
def production_review(
    request: Request,
    date: str | None = None,
    segments: int = 1,
    from1: str | None = None, until1: str | None = None,
    from2: str | None = None, until2: str | None = None,
    from3: str | None = None, until3: str | None = None,
    # Legacy single-cutoff URL (older bookmarks). Maps onto until1.
    cutoff: str | None = None,
):
    from datetime import date as _d
    today = _d.today().isoformat()
    review_date = date or today
    segments = max(1, min(3, segments or 1))

    # The picker is a fixed half-hour clock grid (06:00 → 23:30) — independent
    # of any shift's preset rows. Each option's `key` is the clock time as
    # "HH:MM"; aggregation maps every report cell to its slot's end-time and
    # includes the cell when window_from < end_time <= window_until.
    company = _tally_company_name()

    def _half_hour_grid(start: str = "06:00", end: str = "23:30") -> list[str]:
        out: list[str] = []
        h, m = int(start[:2]), int(start[3:])
        eh, em = int(end[:2]), int(end[3:])
        while (h, m) <= (eh, em):
            out.append(f"{h:02d}:{m:02d}")
            m += 30
            if m >= 60:
                h += 1
                m = 0
        return out

    def _fmt_clock(hhmm: str) -> str:
        h, m = int(hhmm[:2]), int(hhmm[3:])
        ampm = "AM" if h < 12 else "PM"
        h12 = h % 12 or 12
        return f"{h12}:{m:02d} {ampm}"

    grid_keys = _half_hour_grid("00:00", "23:30")
    slot_options = [
        {"key": t, "label": _fmt_clock(t), "display": _fmt_clock(t), "index": idx}
        for idx, t in enumerate(grid_keys)
    ]
    valid_keys = list(grid_keys)

    def _default_range(seg_index: int, total: int) -> tuple[str, str]:
        """Spread default segment ranges evenly across the day's data window
        so 2 segments = halves, 3 = thirds. The window is anchored on the
        first/last clock-time we have report data for (set below); when there
        are no reports yet we fall back to the full grid (06:00 → 23:30)."""
        if not valid_keys:
            return ("", "")
        try:
            lo = valid_keys.index(_default_window[0])
            hi = valid_keys.index(_default_window[1])
        except ValueError:
            lo, hi = 0, len(valid_keys) - 1
        n = (hi - lo) + 1
        start_pos = lo + (seg_index * n) // total
        end_pos = lo + (((seg_index + 1) * n) // total) - 1
        start_pos = max(lo, min(hi, start_pos))
        end_pos = max(start_pos, min(hi, end_pos))
        return (valid_keys[start_pos], valid_keys[end_pos])

    # Filled in once the data load completes — used by `_default_range`.
    _default_window: tuple[str, str] = (valid_keys[0], valid_keys[-1])

    # Aggregate cells once, then derive per-segment totals from the in-memory
    # rows. Each cell is anchored on its slot's clock end-time (fallback start)
    # so window inclusion compares clock times directly.
    report_id_by_line: dict[str, int] = {}
    with _session() as session:
        reports = list(session.scalars(
            select(DailyProductionReport).where(
                DailyProductionReport.report_date == review_date,
            )
        ))
        for r in reports:
            line_label = f"LINE-{r.line}"
            if line_label not in report_id_by_line or (r.shift or "").upper() == "S1":
                report_id_by_line[line_label] = r.id
        report_ids = [r.id for r in reports]
        line_by_report = {r.id: r.line for r in reports}
        shift_by_report = {r.id: (r.shift or "").strip().upper() for r in reports}
        # Per-line "input" row: the active production_processes row whose
        # role='input'. row_key in dpr_hourly_cells stores its id as a string.
        input_row_key_by_line: dict[int, str] = {}
        if reports and company:
            for p in session.scalars(
                select(ProductionProcess).where(
                    ProductionProcess.company_name == company,
                    ProductionProcess.section == "production",
                    ProductionProcess.role == "input",
                    ProductionProcess.active.is_(True),
                )
            ):
                input_row_key_by_line.setdefault(p.line, str(p.id))
        # Each hour-bucket lives at one anchor on the half-hour picker grid:
        #   • normal slot [from, to]          → anchor = to
        #   • "Before X" (only to_time set)   → anchor = to − 30 min
        #   • "After X"  (only from_time set) → anchor = from + 30 min
        # Inclusion is `From < anchor ≤ Until` (left-exclusive) so adjacent
        # windows never double-count. Hourly data is hourly, so each bucket
        # belongs to exactly one window even though the picker is half-hourly.
        def _shift_clock(t: str, minutes: int) -> str:
            h, m = int(t[:2]), int(t[3:])
            total = max(0, min(23 * 60 + 59, h * 60 + m + minutes))
            return f"{total // 60:02d}:{total % 60:02d}"

        slot_anchor: dict[tuple[str, str], str] = {}
        shifts_present = {s for s in shift_by_report.values() if s}
        for sh in shifts_present:
            for s in (_shift_load(session, company, sh) if company else []):
                ft = (s.from_time or "").strip()
                tt = (s.to_time or "").strip()
                if ft and tt:
                    anchor = tt
                elif tt:
                    anchor = _shift_clock(tt, -30)
                elif ft:
                    anchor = _shift_clock(ft, 30)
                else:
                    continue
                slot_anchor[(sh, s.key)] = anchor
        all_cells: list[tuple[str, str, str, float]] = []  # (line, anchor, kind, value)
        if report_ids:
            for c in session.scalars(
                select(DPRHourlyCell).where(DPRHourlyCell.report_id.in_(report_ids))
            ):
                if c.hour_key == "cumulative":
                    continue
                rline = line_by_report.get(c.report_id)
                if not rline:
                    continue
                anchor = slot_anchor.get(
                    (shift_by_report.get(c.report_id, ""), c.hour_key)
                )
                if not anchor:
                    continue
                if (
                    c.section == "production"
                    and str(c.row_key) == input_row_key_by_line.get(rline)
                ):
                    kind = "input"
                elif c.section == "rejection":
                    kind = "rej"
                elif c.section == "rework":
                    kind = "rework"
                else:
                    continue
                all_cells.append(
                    (f"LINE-{rline}", anchor, kind, float(c.value or 0))
                )

    # Snap defaults to the actual data window — earliest and latest anchors
    # rounded onto the half-hour grid. From is one tick earlier so the
    # earliest anchor is included by the left-exclusive comparison below.
    if all_cells:
        anchors = sorted({a for _, a, _, _ in all_cells})

        def _snap(t: str, *, up: bool) -> str:
            for k in (grid_keys if up else reversed(grid_keys)):
                if (k >= t) if up else (k <= t):
                    return k
            return grid_keys[-1] if up else grid_keys[0]

        lo = _snap(anchors[0], up=False)
        # Step From back one half-hour so `from < anchor` still passes for
        # the earliest bucket (otherwise its anchor equals From and is missed).
        try:
            lo_idx = grid_keys.index(lo)
            if lo_idx > 0:
                lo = grid_keys[lo_idx - 1]
        except ValueError:
            pass
        _default_window = (lo, _snap(anchors[-1], up=True))

    lines = sorted(
        {f"LINE-{r.line}" for r in reports},
        key=lambda s: (len(s), s),
    )

    raw_pairs = [(from1, until1), (from2, until2), (from3, until3)]
    # Map legacy ?cutoff=K onto segment 1's Until when from1/until1 are blank.
    if cutoff and not raw_pairs[0][1]:
        raw_pairs[0] = (raw_pairs[0][0], cutoff)

    chosen_ranges: list[tuple[str, str]] = []
    for i in range(segments):
        df, dt = _default_range(i, segments)
        f_raw = (raw_pairs[i][0] or "").strip()
        u_raw = (raw_pairs[i][1] or "").strip()
        f_key = f_raw if f_raw in valid_keys else df
        u_key = u_raw if u_raw in valid_keys else dt
        # Guarantee From precedes (or equals) Until — swap if user inverted.
        if f_key and u_key:
            f_idx = valid_keys.index(f_key)
            u_idx = valid_keys.index(u_key)
            if f_idx > u_idx:
                f_key, u_key = u_key, f_key
        chosen_ranges.append((f_key, u_key))

    seg_totals: list[dict[str, dict[str, float]]] = []
    seg_displays: list[dict[str, str]] = []
    for f_key, u_key in chosen_ranges:
        per_line: dict[str, dict[str, float]] = {}
        if f_key and u_key:
            for line, anchor, kind, val in all_cells:
                # Left-exclusive: From < anchor ≤ Until. Each hour-bucket sits
                # at one anchor, so adjacent windows partition the day cleanly.
                if not (f_key < anchor <= u_key):
                    continue
                b = per_line.setdefault(line, {"input": 0.0, "rej": 0.0, "rework": 0.0})
                b[kind] += val
        for line in lines:
            b = per_line.setdefault(line, {"input": 0.0, "rej": 0.0, "rework": 0.0})
            b["actual"] = b["input"] - b["rej"]
        seg_totals.append(per_line)
        seg_displays.append({"from": _fmt_clock(f_key) if f_key else "",
                             "until": _fmt_clock(u_key) if u_key else ""})

    return templates.TemplateResponse(
        request,
        "production_review.html",
        {
            "review_date": review_date,
            "today": today,
            "lines": lines,
            "slot_options": slot_options,
            "segments": segments,
            "chosen_ranges": chosen_ranges,
            "seg_totals": seg_totals,
            "seg_displays": seg_displays,
            "report_id_by_line": report_id_by_line,
        },
    )


@app.get("/production/recent-reports", response_class=HTMLResponse)
def dpr_recent(
    request: Request,
    date_from: str | None = None,
    date_to: str | None = None,
):
    with _session() as session:
        company = _tally_company_name()
        q = select(DailyProductionReport).order_by(
            DailyProductionReport.report_date.desc(),
            DailyProductionReport.line,
            DailyProductionReport.shift,
            DailyProductionReport.model,
        )
        if company:
            q = q.where(DailyProductionReport.company_name == company)
        if date_from:
            q = q.where(DailyProductionReport.report_date >= date_from)
        if date_to:
            q = q.where(DailyProductionReport.report_date <= date_to)
        reports = list(session.scalars(q))
        draft_count = sum(1 for r in reports if (r.status or "").lower() == "draft")
    return templates.TemplateResponse(
        request,
        "production_recent.html",
        {
            "tally_company_name": company,
            "filter_date_from": date_from or "",
            "filter_date_to": date_to or "",
            "reports": reports,
            "draft_count": draft_count,
        },
    )


@app.get("/production/daily-report", response_class=HTMLResponse)
def dpr_list(
    request: Request,
    company: str | None = None,
    date: str | None = None,
):
    from datetime import date as _d
    with _session() as session:
        # Company is fixed to TALLY_COMPANY_NAME — no longer a user-chosen
        # dropdown. The query-param is still accepted (older bookmarks) but
        # only filters when it matches the configured company.
        company = _tally_company_name()
        q = select(DailyProductionReport).order_by(
            DailyProductionReport.report_date.desc(),
            DailyProductionReport.line,
            DailyProductionReport.shift,
            DailyProductionReport.model,
        )
        if company:
            q = q.where(DailyProductionReport.company_name == company)
        if date:
            q = q.where(DailyProductionReport.report_date == date)
        reports = list(session.scalars(q))
        draft_count = sum(1 for r in reports if (r.status or "").lower() == "draft")
        # Customer / Model cascading dropdowns for the New report form. Only
        # models whose BLANK GLASS / DRILLED GLASS stock items both exist in
        # Tally are loaded — the BG-NF / DG-NF / both-missing flags exclude
        # the rest, mirroring the per-hour entry picker.
        all_specs = list(session.scalars(
            select(ProductionModelSpec).order_by(
                ProductionModelSpec.company_name, ProductionModelSpec.model
            )
        ))
        status_map = _model_missing_status(session, [s.model for s in all_specs])
        models_by_customer: dict[str, list[str]] = {}
        for s in all_specs:
            if status_map.get(s.model, "ok") != "ok":
                continue
            cust = (s.company_name or "").strip()
            if not cust:
                continue
            models_by_customer.setdefault(cust, []).append(s.model)
        shift_rows = _shifts_load(session, company) if company else []
    return templates.TemplateResponse(
        request,
        "production_list.html",
        {
            "tally_company_name": company,
            "filter_date": date or "",
            "reports": reports,
            "draft_count": draft_count,
            "today": _d.today().isoformat(),
            "shift_options": [(s.key, s.name) for s in shift_rows],
            "line_options": dpr.LINE_OPTIONS,
            "models_by_customer": models_by_customer,
        },
    )


@app.post("/production/daily-report")
async def dpr_create(request: Request):
    from urllib.parse import urlencode
    form = await request.form()
    company = (form.get("company") or "").strip()
    report_date = (form.get("report_date") or "").strip()
    shift = (form.get("shift") or "").strip()
    line = (form.get("line") or "").strip()
    model = (form.get("model") or "").strip()
    if not all([company, report_date, shift, line]):
        raise HTTPException(400, "company, date, shift and line are required")
    import json as _json
    with _session() as session:
        preset_rows = _shift_load(session, company, shift)
        slots = [(r.key, r.label) for r in preset_rows]
    if not slots:
        slots = list(dpr.SHIFT_PRESETS.get(shift) or dpr.HOUR_SLOTS)
    hour_slots_json = _json.dumps([{"key": k, "label": l} for k, l in slots])
    with _session() as session:
        # Upsert: if a report already exists for this key, reuse it so users
        # don't accidentally create two.
        existing = session.scalar(
            select(DailyProductionReport).where(
                DailyProductionReport.company_name == company,
                DailyProductionReport.report_date == report_date,
                DailyProductionReport.shift == shift,
                DailyProductionReport.line == line,
                DailyProductionReport.model == model,
            )
        )
        if existing:
            report_id = existing.id
        else:
            row = DailyProductionReport(
                company_name=company,
                report_date=report_date,
                shift=shift,
                line=line,
                model=model,
                hour_slots_json=hour_slots_json,
            )
            session.add(row)
            session.commit()
            report_id = row.id
    return RedirectResponse(f"/production/daily-report/{report_id}/edit", status_code=303)


@app.get("/production/daily-report/{report_id}/edit", response_class=HTMLResponse)
def dpr_edit(request: Request, report_id: int):
    with _session() as session:
        report = session.get(DailyProductionReport, report_id)
        if not report:
            raise HTTPException(404, "report not found")
        # Submitted reports are locked — the user gets bounced to the
        # read-only summary instead.
        if (report.status or "").strip().lower() == "submitted":
            return RedirectResponse(
                f"/production/daily-report/{report_id}",
                status_code=303,
            )
        hour_slots = dpr.load_hour_slots(report.hour_slots_json, report.shift)
        production_rows, rejection_groups, rework_rows = _dpr_rows_for_report(session, report)
        sections_rows = {
            "production": production_rows,
            "rejection": [r for _, rows in rejection_groups for r in rows],
            "rework": rework_rows,
        }
        cells = _dpr_load_cells(session, report, hour_slots, sections_rows)
        idle_events = sorted(report.idle_events, key=lambda e: e.ordinal)
        totals = _dpr_compute_totals(cells, hour_slots, production_rows)
        hour_models = _dpr_load_hour_models(report)
        hour_model_chain = _dpr_resolve_hour_model_chain(hour_models, report.model or "", hour_slots)
        # Distinct models in the order they first ran during the shift.
        # Used to display "VEDA 3B | New Model | …" beside the editable
        # primary-model input on the edit page.
        distinct_models: list[str] = []
        for _k, _l in hour_slots:
            m = hour_model_chain.get(_k, "")
            if m and m not in distinct_models:
                distinct_models.append(m)
        if not distinct_models and (report.model or "").strip():
            distinct_models = [report.model.strip()]
        config_empty = not (production_rows or rejection_groups or rework_rows)
        default_doc_no = _dpr_doc_no_for_display(session, report)
        rd = report.report_date
        if isinstance(rd, str):
            from datetime import date as _date
            try:
                rd = _date.fromisoformat(rd)
            except ValueError:
                rd = _date.today()
        default_rev_date = rd.strftime("%d/%m/%Y")
        shift_rows = _shifts_load(session, report.company_name)
    return templates.TemplateResponse(
        request,
        "production_edit.html",
        {
            "report": report,
            "hour_slots": hour_slots,
            "cumulative_key": dpr.CUMULATIVE_KEY,
            "production_rows": production_rows,
            "rejection_groups": rejection_groups,
            "rework_rows": rework_rows,
            "cells": cells,
            "idle_events": idle_events,
            "totals": totals,
            "hour_models": hour_models,
            "hour_model_chain": hour_model_chain,
            "distinct_models": distinct_models,
            "shift_options": [(s.key, s.name) for s in shift_rows],
            "line_options": dpr.LINE_OPTIONS,
            "config_empty": config_empty,
            "default_doc_no": default_doc_no,
            "default_rev_date": default_rev_date,
        },
    )


_OLD_AUTO_DOC_PATTERN = "/PRD-ENTRY/"


def _dpr_doc_no_for_display(session: Session, report: DailyProductionReport) -> str:
    """Return the doc_no to show on edit/summary pages.

    If the stored value looks like an earlier auto-generated format (we used
    'PRD-ENTRY' as the segment marker), regenerate using the current format
    so the user doesn't have to manually clear stale auto-numbers."""
    stored = (report.doc_no or "").strip()
    if stored and _OLD_AUTO_DOC_PATTERN not in stored:
        return stored
    return _dpr_default_doc_no(session, report)


def _dpr_default_doc_no(session: Session, report: DailyProductionReport) -> str:
    """Generate a doc-no like AAPL/26-27/PRD-ENTRY/DD-MM-YYYY/LINE<n>[_seq].

    The company prefix is the initials of the company name (Avinash
    Appliances Private Limited → AAPL). FY is computed from report_date
    (Apr–Mar Indian FY). Sequential suffix is appended (_2, _3, ...) when
    multiple DPR records exist for the same (company, date, line) — the
    first such report has no suffix."""
    co = report.company_name or ""
    # Strip trailing parenthetical content (e.g. "(April 26 - March 27)") so
    # FY/date suffixes don't pollute the initials.
    import re as _re
    co_clean = _re.sub(r"\s*\([^)]*\)\s*$", "", co).strip()
    initials = "".join(w[0] for w in co_clean.split() if w and w[0].isalpha()).upper()[:6] or "CO"
    rd = report.report_date
    if isinstance(rd, str):
        from datetime import date as _date
        try:
            d = _date.fromisoformat(rd)
        except ValueError:
            d = _date.today()
    else:
        d = rd
    fy_start = d.year if d.month >= 4 else d.year - 1
    fy = f"{fy_start % 100:02d}-{(fy_start + 1) % 100:02d}"
    same_day_reports = list(session.scalars(
        select(DailyProductionReport)
        .where(
            DailyProductionReport.company_name == report.company_name,
            DailyProductionReport.report_date == report.report_date,
            DailyProductionReport.line == report.line,
        )
        .order_by(DailyProductionReport.id)
    ))
    seq = next((i for i, r in enumerate(same_day_reports, start=1) if r.id == report.id), 1)
    date_short = f"{d.day:02d}-{d.month}-{d.year % 100:02d}"
    return f"{initials}/PRD/{date_short}/{report.shift}/LINE-{report.line}_{seq}"


def _dpr_parse_float(raw) -> float:
    try:
        s = (raw or "").strip() if isinstance(raw, str) else raw
        if s in (None, "", "-"):
            return 0.0
        return float(s)
    except (TypeError, ValueError):
        return 0.0


@app.post("/production/daily-report/{report_id}/save")
async def dpr_save(request: Request, report_id: int):
    form = await request.form()
    action = (form.get("action") or "save").strip()
    with _session() as session:
        report = session.get(DailyProductionReport, report_id)
        if not report:
            raise HTTPException(404, "report not found")
        # Once a report has been pushed to Tally (status == 'submitted') it is
        # frozen — any edit/save/finalize request is rejected so the books
        # match what was posted.
        if (report.status or "").strip().lower() == "submitted":
            raise HTTPException(
                403,
                "This report has already been posted to Tally and cannot be edited.",
            )

        # ---- Meta fields ----
        report.model = (form.get("model") or "").strip()
        report.narration = (form.get("narration") or "").strip() or None
        report.rework_cleared_qty = _dpr_parse_float(form.get("rework_cleared_qty"))
        report.supervisor_name = (form.get("supervisor_name") or "").strip() or None
        report.incharge_name = (form.get("incharge_name") or "").strip() or None
        report.head_name = (form.get("head_name") or "").strip() or None
        report.doc_no = (form.get("doc_no") or "").strip() or None
        report.rev_no = (form.get("rev_no") or "").strip() or None
        report.rev_date = (form.get("rev_date") or "").strip() or None

        # ---- Editable hour labels ----
        hour_slots = dpr.load_hour_slots(report.hour_slots_json, report.shift)
        new_slots: list[tuple[str, str]] = []
        for key, default_label in hour_slots:
            new_label = (form.get(f"hour_label__{key}") or default_label).strip() or default_label
            new_slots.append((key, new_label))
        import json as _json
        report.hour_slots_json = _json.dumps([{"key": k, "label": l} for k, l in new_slots])

        # ---- Grid cells ----
        # Replace all cells atomically: simpler than diffing and avoids leaving
        # ghost values when a row is removed from the static row list.
        for existing in list(report.cells):
            session.delete(existing)
        session.flush()
        prod_rows, rej_groups, rw_rows = _dpr_rows_for_report(session, report)
        row_keys_by_section = {
            "production": [r.key for r in prod_rows],
            "rejection": [r.key for _, rows in rej_groups for r in rows],
            "rework": [r.key for r in rw_rows],
        }
        hour_keys = [dpr.CUMULATIVE_KEY] + [k for k, _ in new_slots]
        for section, row_keys in row_keys_by_section.items():
            for row_key in row_keys:
                for hour_key in hour_keys:
                    v = _dpr_parse_float(form.get(f"cell__{section}__{row_key}__{hour_key}"))
                    if v:
                        session.add(DPRHourlyCell(
                            report_id=report.id,
                            section=section,
                            row_key=row_key,
                            hour_key=hour_key,
                            value=v,
                        ))

        # ---- Per-hour model overrides ----
        # Only persist *change points* — hours where the entered model
        # differs from the previous resolved model. Forward-fill is a render
        # concern; storing every echoed cell would clutter the table and
        # break the "blank = inherit from previous" mental model.
        for existing in list(report.hour_models):
            session.delete(existing)
        session.flush()
        primary_model = (report.model or "").strip()
        prev_model = primary_model
        for key, _label in new_slots:
            raw = (form.get(f"hour_model__{key}") or "").strip()
            if raw and raw != prev_model:
                session.add(DPRHourModel(
                    report_id=report.id,
                    hour_key=key,
                    model=raw,
                ))
                prev_model = raw

        # ---- Idle events ----
        for existing in list(report.idle_events):
            session.delete(existing)
        session.flush()
        idle_indices = sorted({
            int(k.split("__")[1]) for k in form.keys()
            if k.startswith("idle__") and k.split("__")[1].isdigit()
        })
        for ordinal, idx in enumerate(idle_indices):
            desc = (form.get(f"idle__{idx}__description") or "").strip()
            machine = (form.get(f"idle__{idx}__machine") or "").strip()
            if not desc and not machine:
                continue  # Skip empty idle rows the user added but never filled.
            session.add(DPRIdleEvent(
                report_id=report.id,
                ordinal=ordinal,
                machine=machine or None,
                description=desc or None,
                from_time=(form.get(f"idle__{idx}__from") or "").strip() or None,
                to_time=(form.get(f"idle__{idx}__to") or "").strip() or None,
                time_loss_min=_dpr_parse_float(form.get(f"idle__{idx}__time_loss")),
                attended_by=(form.get(f"idle__{idx}__attended_by") or "").strip() or None,
                remarks=(form.get(f"idle__{idx}__remarks") or "").strip() or None,
            ))

        # ---- Per-hour rejection-allocation check ----
        # Mark-ready and View-summary both require the same allocation to be
        # closed out that the hour-entry page enforces. Save draft skips the
        # check so partial work-in-progress isn't blocked. We commit the
        # draft so the user's edits aren't lost, but refuse to advance the
        # status / navigate away when any hour has uncounted glasses.
        validation_errors: list[str] = []
        if action in ("submit", "view"):
            stage_chain = _dpr_build_stage_chain(prod_rows, rej_groups)

            def _form_cell(section: str, row_key: str | None, hkey: str) -> float:
                if not row_key:
                    return 0.0
                return _dpr_parse_float(form.get(f"cell__{section}__{row_key}__{hkey}"))

            for hkey, hlabel in new_slots:
                for s in stage_chain:
                    if not s.get("validate_count"):
                        continue
                    input_val = _form_cell("production", s.get("effective_input_row_key"), hkey)
                    output_val = _form_cell("production", s.get("output_row_key"), hkey)
                    filled = sum(
                        _form_cell("rejection", rk, hkey)
                        for rk in s.get("rejection_row_keys", [])
                    )
                    if input_val == 0 and output_val == 0 and filled == 0:
                        continue
                    diff = (input_val - output_val) - filled
                    if abs(diff) >= 0.001:
                        if diff > 0:
                            validation_errors.append(
                                f"{hlabel} · {s['label']}: {diff:.0f} uncounted "
                                f"(input {input_val:.0f} − output {output_val:.0f} = "
                                f"{(input_val - output_val):.0f}, rejection adds up to {filled:.0f})"
                            )
                        else:
                            validation_errors.append(
                                f"{hlabel} · {s['label']}: {(-diff):.0f} over expected "
                                f"(input {input_val:.0f} − output {output_val:.0f} = "
                                f"{(input_val - output_val):.0f}, rejection adds up to {filled:.0f})"
                            )

        # Status state machine:
        #   draft → saved (via the "Save" button, action=submit). Saved means
        #     "ready to post but not yet pushed to Tally" — still editable.
        #   saved → submitted ONLY happens when the SJ-LINE voucher is
        #     successfully posted to Tally (line_publish_submit).
        if action == "submit" and not validation_errors:
            report.status = "saved"
            report.submitted_at = datetime.utcnow()
        report.updated_at = datetime.utcnow()
        session.commit()

        if validation_errors:
            from urllib.parse import urlencode
            qs = urlencode({"validation_error": "\n".join(validation_errors)})
            return RedirectResponse(
                f"/production/daily-report/{report_id}/edit?{qs}",
                status_code=303,
            )

        target = f"/production/daily-report/{report_id}"
        # ``save`` keeps the user on the edit page; ``submit`` (now meaning
        # finalize-as-saved) and ``view`` jump to the summary.
        if action == "save":
            target += "/edit"
    return RedirectResponse(target, status_code=303)


@app.get("/production/daily-report/{report_id}/entry", response_class=HTMLResponse)
def dpr_entry(request: Request, report_id: int, hour: str | None = None):
    with _session() as session:
        report = session.get(DailyProductionReport, report_id)
        if not report:
            raise HTTPException(404, "report not found")
        hour_slots = dpr.load_hour_slots(report.hour_slots_json, report.shift)
        valid_keys = [k for k, _ in hour_slots]
        if not valid_keys:
            raise HTTPException(400, "report has no hour slots configured")
        selected_hour = hour if hour in valid_keys else valid_keys[0]
        selected_label = next(lbl for k, lbl in hour_slots if k == selected_hour)
        production_rows, rejection_groups, rework_rows = _dpr_rows_for_report(session, report)
        sections_rows = {
            "production": production_rows,
            "rejection": [r for _, rows in rejection_groups for r in rows],
            "rework": rework_rows,
        }
        cells = _dpr_load_cells(session, report, hour_slots, sections_rows)
        config_empty = not (production_rows or rejection_groups or rework_rows)
        prefilled_keys = {
            "rejection": {
                r.key
                for _, rows in rejection_groups
                for r in rows
                if cells["rejection"][r.key].get(selected_hour)
            },
            "rework": {
                r.key for r in rework_rows
                if cells["rework"][r.key].get(selected_hour)
            },
        }
        stage_chain = _dpr_build_stage_chain(production_rows, rejection_groups)
        hour_models = _dpr_load_hour_models(report)
        hour_model_chain = _dpr_resolve_hour_model_chain(hour_models, report.model or "", hour_slots)
        selected_model = hour_model_chain.get(selected_hour, "")
        all_specs = list(session.scalars(
            select(ProductionModelSpec).order_by(
                ProductionModelSpec.company_name, ProductionModelSpec.model
            )
        ))
        status_map = _model_missing_status(session, [s.model for s in all_specs])
        models_by_customer: dict[str, list[str]] = {}
        spec_customer_by_model: dict[str, str] = {}
        for s in all_specs:
            if status_map.get(s.model, "ok") != "ok":
                continue  # missing BLANK or DRILLED stock item in Tally → can't post voucher
            cust = s.company_name or ""
            models_by_customer.setdefault(cust, []).append(s.model)
            spec_customer_by_model[s.model] = cust
        selected_customer = spec_customer_by_model.get(selected_model, "")
    return templates.TemplateResponse(
        request,
        "production_entry.html",
        {
            "report": report,
            "hour_slots": hour_slots,
            "selected_hour": selected_hour,
            "selected_label": selected_label,
            "selected_model": selected_model,
            "selected_customer": selected_customer,
            "models_by_customer": models_by_customer,
            "production_rows": production_rows,
            "rejection_groups": rejection_groups,
            "rework_rows": rework_rows,
            "cells": cells,
            "prefilled_keys": prefilled_keys,
            "stage_chain": stage_chain,
            "config_empty": config_empty,
        },
    )


def _dpr_build_stage_chain(
    production_rows: list[_DprRow],
    rejection_groups: list[tuple[str, list[_DprRow]]],
) -> list[dict]:
    """Derive a per-stage I/O chain from the report's row catalog.

    For each stage encountered (in production-row order):
      - input_row_key: the production row whose label ends in 'Input', if any
      - output_row_key: the other production row for this stage (treated as
        the stage's output) — last seen wins if multiple
      - When a stage has no explicit input row, its effective input is the
        previous stage's output (no WIP between stages, per user spec).
      - rejection_row_keys: rejection rows whose stage matches.
    """
    stages_seen: list[str] = []
    by_stage: dict[str, dict] = {}
    for r in production_rows:
        s = by_stage.setdefault(r.stage, {
            "key": r.stage,
            "label": r.stage.replace("_", " ").title(),
            "input_row_key": None,
            "input_row_label": None,
            "output_row_key": None,
            "output_row_label": None,
            "rejection_row_keys": [],
            "rejection_rows": [],
            "validate_count": True,
        })
        if r.stage not in stages_seen:
            stages_seen.append(r.stage)
        # Prefer the explicit role flag; fall back to label heuristic for
        # rows that haven't been migrated yet.
        is_input = (r.role == "input") if r.role else ("input" in (r.label or "").lower())
        if is_input:
            s["input_row_key"] = r.key
            s["input_row_label"] = r.label
        else:
            s["output_row_key"] = r.key
            s["output_row_label"] = r.label
            # Output row drives the stage's allocation check.
            s["validate_count"] = bool(r.validate_count)
            # Use this row's label to refine the stage label (e.g. "Single Edger")
            lbl = r.label
            for suffix in (" Output", " (Incl Rework)"):
                if lbl.endswith(suffix):
                    lbl = lbl[: -len(suffix)]
            s["label"] = lbl

    rej_by_stage: dict[str, list[_DprRow]] = {}
    for _g, rows in rejection_groups:
        for r in rows:
            rej_by_stage.setdefault(r.stage, []).append(r)
    for s_key, rows in rej_by_stage.items():
        if s_key in by_stage:
            by_stage[s_key]["rejection_row_keys"] = [r.key for r in rows]
            by_stage[s_key]["rejection_rows"] = [{"key": r.key, "label": r.label} for r in rows]

    chain: list[dict] = []
    prev_output_key = None
    for s_key in stages_seen:
        s = by_stage[s_key]
        # Effective input source: explicit input row OR previous stage's output
        s["effective_input_row_key"] = s["input_row_key"] or prev_output_key
        chain.append(s)
        if s["output_row_key"]:
            prev_output_key = s["output_row_key"]
    return chain


@app.post("/production/daily-report/{report_id}/entry/save")
async def dpr_entry_save(request: Request, report_id: int):
    form = await request.form()
    hour_key = (form.get("hour") or "").strip()
    with _session() as session:
        report = session.get(DailyProductionReport, report_id)
        if not report:
            raise HTTPException(404, "report not found")
        hour_slots = dpr.load_hour_slots(report.hour_slots_json, report.shift)
        valid_keys = {k for k, _ in hour_slots}
        if hour_key not in valid_keys:
            raise HTTPException(400, "invalid hour slot")

        # ---- Upsert cells for this hour only ----
        # Delete existing cells for the selected hour, then re-insert from form.
        # Other hours' cells and cumulative cells are left untouched.
        for existing in list(report.cells):
            if existing.hour_key == hour_key:
                session.delete(existing)
        session.flush()

        prod_rows, rej_groups, rw_rows = _dpr_rows_for_report(session, report)
        row_keys_by_section = {
            "production": [r.key for r in prod_rows],
            "rejection": [r.key for _, rows in rej_groups for r in rows],
            "rework": [r.key for r in rw_rows],
        }
        for section, row_keys in row_keys_by_section.items():
            for row_key in row_keys:
                v = _dpr_parse_float(form.get(f"cell__{section}__{row_key}"))
                if v:
                    session.add(DPRHourlyCell(
                        report_id=report.id,
                        section=section,
                        row_key=row_key,
                        hour_key=hour_key,
                        value=v,
                    ))

        # ---- Per-hour model override for this hour ----
        # Compute the model that *would* be in effect at this hour from the
        # previous change points alone (excluding any existing row for this
        # hour). Persist only if the user's value differs.
        primary_model = (report.model or "").strip()
        existing_overrides = {hm.hour_key: hm.model for hm in report.hour_models if hm.hour_key != hour_key}
        prev_model = primary_model
        for k, _l in hour_slots:
            if k == hour_key:
                break
            ov = (existing_overrides.get(k) or "").strip()
            if ov:
                prev_model = ov
        raw_model = (form.get("hour_model") or "").strip()
        for existing in list(report.hour_models):
            if existing.hour_key == hour_key:
                session.delete(existing)
        session.flush()
        if raw_model and raw_model != prev_model:
            session.add(DPRHourModel(
                report_id=report.id,
                hour_key=hour_key,
                model=raw_model,
            ))

        # ---- Append idle events (do NOT delete existing) ----
        existing_max = max((e.ordinal for e in report.idle_events), default=-1)
        idle_indices = sorted({
            int(k.split("__")[1]) for k in form.keys()
            if k.startswith("idle_new__") and len(k.split("__")) >= 3 and k.split("__")[1].isdigit()
        })
        next_ord = existing_max + 1
        for idx in idle_indices:
            desc = (form.get(f"idle_new__{idx}__description") or "").strip()
            machine = (form.get(f"idle_new__{idx}__machine") or "").strip()
            if not desc and not machine:
                continue
            session.add(DPRIdleEvent(
                report_id=report.id,
                ordinal=next_ord,
                machine=machine or None,
                description=desc or None,
                from_time=(form.get(f"idle_new__{idx}__from") or "").strip() or None,
                to_time=(form.get(f"idle_new__{idx}__to") or "").strip() or None,
                time_loss_min=_dpr_parse_float(form.get(f"idle_new__{idx}__time_loss")),
                attended_by=(form.get(f"idle_new__{idx}__attended_by") or "").strip() or None,
                remarks=(form.get(f"idle_new__{idx}__remarks") or "").strip() or None,
            ))
            next_ord += 1

        report.updated_at = datetime.utcnow()
        session.commit()
    return RedirectResponse(f"/production/daily-report/{report_id}/edit", status_code=303)


@app.get("/production/daily-report/{report_id}/import/template")
def dpr_import_template(report_id: int):
    """Return an Excel template pre-filled with this report's structure
    (sections, rows, hour columns) and live formulas for row/col totals.
    A hidden first column carries `section:row_key` markers so we can map
    the upload back to DB cells without relying on label matches."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    with _session() as session:
        report = session.get(DailyProductionReport, report_id)
        if not report:
            raise HTTPException(404, "report not found")
        hour_slots = dpr.load_hour_slots(report.hour_slots_json, report.shift)
        production_rows, rejection_groups, rework_rows = _dpr_rows_for_report(session, report)
        sections_rows = {
            "production": production_rows,
            "rejection": [r for _, rows in rejection_groups for r in rows],
            "rework": rework_rows,
        }
        cells = _dpr_load_cells(session, report, hour_slots, sections_rows)
        idle_events_snapshot = [
            (e.machine, e.description, e.from_time, e.to_time, e.time_loss_min, e.attended_by, e.remarks)
            for e in sorted(report.idle_events, key=lambda x: x.ordinal)
        ]
        report_meta = {
            "company_name": report.company_name,
            "report_date": str(report.report_date),
            "shift": report.shift,
            "line": report.line,
            "model": report.model or "",
        }

    wb = Workbook()
    ws = wb.active
    ws.title = "DPR"

    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="DBEAFE")
    grp_fill = PatternFill("solid", fgColor="EDE9FE")
    cum_fill = PatternFill("solid", fgColor="FEF3C7")
    total_fill = PatternFill("solid", fgColor="E5E7EB")

    # Meta block
    meta = [
        ("Company", report_meta["company_name"]),
        ("Date", report_meta["report_date"]),
        ("Shift", report_meta["shift"]),
        ("Line", str(report_meta["line"])),
        ("Model", report_meta["model"]),
    ]
    for i, (k, v) in enumerate(meta, start=1):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)

    # Column layout: A=marker (hidden), B=Description, C=Cumulative,
    # D..(D+N-1)=hour slots
    n_hours = len(hour_slots)
    col_marker = 1
    col_desc = 2
    col_cum = 3
    col_first_hour = 4
    col_last_hour = col_first_hour + n_hours - 1
    col_row_total = col_last_hour  # alias kept for column-width loop bound
    last_col_letter = get_column_letter(col_last_hour)

    header_row = 7
    ws.cell(row=header_row, column=col_marker, value="_marker").font = Font(bold=True, italic=True, color="888888")
    ws.cell(row=header_row, column=col_desc, value="Description").font = Font(bold=True)
    ws.cell(row=header_row, column=col_cum, value="Cumulative").font = Font(bold=True)
    ws.cell(row=header_row, column=col_cum).fill = cum_fill
    for i, (_key, label) in enumerate(hour_slots):
        c = ws.cell(row=header_row, column=col_first_hour + i, value=label)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for cc in range(1, col_last_hour + 1):
        ws.cell(row=header_row, column=cc).fill = head_fill
        ws.cell(row=header_row, column=cc).border = border

    cur_row = header_row + 1

    def write_section(title: str, rows_list, section_key: str, groups=None):
        nonlocal cur_row
        # Section heading row spans whole grid
        c = ws.cell(row=cur_row, column=col_desc, value=title)
        c.font = Font(bold=True, size=12, color="1E3A8A")
        cur_row += 1

        body_first = cur_row

        if groups:
            for group_name, group_rows in groups:
                gc = ws.cell(row=cur_row, column=col_desc, value=group_name)
                gc.font = Font(bold=True, color="4C1D95")
                gc.fill = grp_fill
                for cc in range(1, col_last_hour + 1):
                    ws.cell(row=cur_row, column=cc).fill = grp_fill
                cur_row += 1
                for r in group_rows:
                    _write_data_row(section_key, r)
        else:
            for r in (rows_list or []):
                _write_data_row(section_key, r)

        body_last = cur_row - 1
        if body_last >= body_first:
            # Footer column totals (only when there is at least one data row)
            ws.cell(row=cur_row, column=col_desc, value=f"{title} Total").font = Font(bold=True)
            ws.cell(row=cur_row, column=col_cum, value=f"=SUM({get_column_letter(col_cum)}{body_first}:{get_column_letter(col_cum)}{body_last})")
            for i in range(n_hours):
                col_letter = get_column_letter(col_first_hour + i)
                ws.cell(row=cur_row, column=col_first_hour + i, value=f"=SUM({col_letter}{body_first}:{col_letter}{body_last})")
            for cc in range(1, col_last_hour + 1):
                ws.cell(row=cur_row, column=cc).fill = total_fill
                ws.cell(row=cur_row, column=cc).font = Font(bold=True)
                ws.cell(row=cur_row, column=cc).border = border
            cur_row += 1
        cur_row += 1  # blank spacer

    def _write_data_row(section_key: str, r):
        nonlocal cur_row
        ws.cell(row=cur_row, column=col_marker, value=f"{section_key}:{r.key}")
        ws.cell(row=cur_row, column=col_desc, value=r.label)
        # Cumulative — pre-fill existing value (no formula here so user can override)
        existing_cum = cells[section_key][r.key].get(dpr.CUMULATIVE_KEY) or 0
        cum_cell = ws.cell(row=cur_row, column=col_cum, value=(existing_cum if existing_cum else None))
        cum_cell.fill = cum_fill
        # Hour cells — pre-fill existing values
        for i, (hkey, _label) in enumerate(hour_slots):
            v = cells[section_key][r.key].get(hkey) or 0
            ws.cell(row=cur_row, column=col_first_hour + i, value=(v if v else None))
        for cc in range(1, col_last_hour + 1):
            ws.cell(row=cur_row, column=cc).border = border
        cur_row += 1

    write_section("1 · Production", production_rows, "production")
    write_section("2 · Rejection", None, "rejection", groups=rejection_groups)
    write_section("3 · Rework", rework_rows, "rework")

    # Column widths
    ws.column_dimensions[get_column_letter(col_marker)].hidden = True
    ws.column_dimensions[get_column_letter(col_desc)].width = 32
    ws.column_dimensions[get_column_letter(col_cum)].width = 12
    for i in range(n_hours):
        ws.column_dimensions[get_column_letter(col_first_hour + i)].width = 12
    ws.row_dimensions[header_row].height = 32

    # Idle events sheet
    ws2 = wb.create_sheet("Idle Time")
    headers = ["Machine", "Description", "From (HH:MM)", "To (HH:MM)", "Time Loss (min)", "Attended By", "Remarks"]
    for i, h in enumerate(headers, start=1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = head_fill
        c.border = border
    for idx, (machine, desc, frm, to, loss, attended, remarks) in enumerate(idle_events_snapshot, start=2):
        ws2.cell(row=idx, column=1, value=machine)
        ws2.cell(row=idx, column=2, value=desc)
        ws2.cell(row=idx, column=3, value=frm)
        ws2.cell(row=idx, column=4, value=to)
        ws2.cell(row=idx, column=5, value=loss)
        ws2.cell(row=idx, column=6, value=attended)
        ws2.cell(row=idx, column=7, value=remarks)
    # Time-loss formula hint for new rows (rows 2..50)
    for r in range(2, 51):
        cell = ws2.cell(row=r, column=5)
        if cell.value in (None, 0):
            cell.value = f'=IF(AND(ISNUMBER(--C{r}),ISNUMBER(--D{r})),(MOD((TIMEVALUE(D{r})-TIMEVALUE(C{r}))+1,1))*1440,"")'
    for col, w in zip("ABCDEFG", [16, 32, 14, 14, 16, 18, 32]):
        ws2.column_dimensions[col].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"dpr_{report.report_date}_{report.shift}_L{report.line}_template.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.post("/production/daily-report/{report_id}/import")
async def dpr_import(report_id: int, file: UploadFile = File(...)):
    """Parse an uploaded DPR Excel and replace the report's cells + idle
    events. Validates that the workbook has the marker column we wrote in
    the template, then reads computed values (data_only=True) so any
    user-typed formulas resolve to numbers."""
    from io import BytesIO
    from openpyxl import load_workbook

    raw = await file.read()
    if not raw:
        raise HTTPException(400, "uploaded file is empty")
    try:
        wb = load_workbook(BytesIO(raw), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"could not read Excel: {e}")

    if "DPR" not in wb.sheetnames:
        raise HTTPException(400, "missing 'DPR' sheet — please use the downloaded template")
    ws = wb["DPR"]

    with _session() as session:
        report = session.get(DailyProductionReport, report_id)
        if not report:
            raise HTTPException(404, "report not found")
        hour_slots = dpr.load_hour_slots(report.hour_slots_json, report.shift)
        n_hours = len(hour_slots)
        prod_rows, rej_groups, rw_rows = _dpr_rows_for_report(session, report)
        valid_keys = {
            "production": {r.key for r in prod_rows},
            "rejection": {r.key for _, rows in rej_groups for r in rows},
            "rework": {r.key for r in rw_rows},
        }

        col_marker, col_cum, col_first_hour = 1, 3, 4

        # Find header row by locating "_marker" in column A
        header_row = None
        for row in range(1, 25):
            v = ws.cell(row=row, column=col_marker).value
            if isinstance(v, str) and v.strip() == "_marker":
                header_row = row
                break
        if header_row is None:
            raise HTTPException(
                400,
                "could not find marker column — please use the downloaded template (do not delete column A or the header row)",
            )

        # Collect (section, row_key, row_index) pairs
        parsed: list[tuple[str, str, int]] = []
        max_row = ws.max_row or 0
        for r in range(header_row + 1, max_row + 1):
            mv = ws.cell(row=r, column=col_marker).value
            if not isinstance(mv, str) or ":" not in mv:
                continue
            section, row_key = mv.split(":", 1)
            section = section.strip()
            row_key = row_key.strip()
            if section not in valid_keys or row_key not in valid_keys[section]:
                continue
            parsed.append((section, row_key, r))

        if not parsed:
            raise HTTPException(400, "no recognised data rows — make sure the marker column was preserved")

        def _num(v):
            if v in (None, ""):
                return 0.0
            try:
                return float(v)
            except (TypeError, ValueError):
                return 0.0

        # Replace all cells atomically
        for existing in list(report.cells):
            session.delete(existing)
        session.flush()

        rows_imported = 0
        for section, row_key, r in parsed:
            cum = _num(ws.cell(row=r, column=col_cum).value)
            if cum:
                session.add(DPRHourlyCell(
                    report_id=report.id, section=section, row_key=row_key,
                    hour_key=dpr.CUMULATIVE_KEY, value=cum,
                ))
                rows_imported += 1
            for i, (hkey, _label) in enumerate(hour_slots):
                v = _num(ws.cell(row=r, column=col_first_hour + i).value)
                if v:
                    session.add(DPRHourlyCell(
                        report_id=report.id, section=section, row_key=row_key,
                        hour_key=hkey, value=v,
                    ))
                    rows_imported += 1

        # Idle Time sheet (optional) — replace if present
        if "Idle Time" in wb.sheetnames:
            ws2 = wb["Idle Time"]
            for existing in list(report.idle_events):
                session.delete(existing)
            session.flush()
            ordinal = 0
            for r in range(2, (ws2.max_row or 1) + 1):
                machine = ws2.cell(row=r, column=1).value
                desc = ws2.cell(row=r, column=2).value
                if not (machine or desc):
                    continue
                session.add(DPRIdleEvent(
                    report_id=report.id,
                    ordinal=ordinal,
                    machine=str(machine).strip() if machine else None,
                    description=str(desc).strip() if desc else None,
                    from_time=str(ws2.cell(row=r, column=3).value).strip() if ws2.cell(row=r, column=3).value else None,
                    to_time=str(ws2.cell(row=r, column=4).value).strip() if ws2.cell(row=r, column=4).value else None,
                    time_loss_min=_num(ws2.cell(row=r, column=5).value),
                    attended_by=str(ws2.cell(row=r, column=6).value).strip() if ws2.cell(row=r, column=6).value else None,
                    remarks=str(ws2.cell(row=r, column=7).value).strip() if ws2.cell(row=r, column=7).value else None,
                ))
                ordinal += 1

        report.updated_at = datetime.utcnow()
        session.commit()

    return RedirectResponse(f"/production/daily-report/{report_id}/edit", status_code=303)


@app.get("/production/daily-report/{report_id}", response_class=HTMLResponse)
def dpr_summary(request: Request, report_id: int):
    with _session() as session:
        report = session.get(DailyProductionReport, report_id)
        if not report:
            raise HTTPException(404, "report not found")
        hour_slots = dpr.load_hour_slots(report.hour_slots_json, report.shift)
        production_rows, rejection_groups, rework_rows = _dpr_rows_for_report(session, report)
        sections_rows = {
            "production": production_rows,
            "rejection": [r for _, rows in rejection_groups for r in rows],
            "rework": rework_rows,
        }
        cells = _dpr_load_cells(session, report, hour_slots, sections_rows)
        idle_events = sorted(report.idle_events, key=lambda e: e.ordinal)
        totals = _dpr_compute_totals(cells, hour_slots, production_rows)
        hour_models = _dpr_load_hour_models(report)
        hour_model_chain = _dpr_resolve_hour_model_chain(hour_models, report.model or "", hour_slots)
        production_by_model = _dpr_production_by_model(
            cells, hour_slots, production_rows, hour_models, report.model or ""
        )
        # Distinct models in the order they first appear during the shift,
        # used for the paper-meta "Model" cell ("Renz3b | Renz5").
        seen_models: list[str] = []
        for _k, _l in hour_slots:
            m = hour_model_chain.get(_k, "")
            if m and m not in seen_models:
                seen_models.append(m)
        if not seen_models and (report.model or "").strip():
            seen_models = [report.model.strip()]
        models_label = " | ".join(seen_models) if seen_models else (report.model or "—")
        default_doc_no = _dpr_doc_no_for_display(session, report)
        rd = report.report_date
        if isinstance(rd, str):
            from datetime import date as _date
            try:
                rd = _date.fromisoformat(rd)
            except ValueError:
                rd = _date.today()
        default_rev_date = rd.strftime("%d/%m/%Y")
        company_logo = _logo_for_company(report.company_name or "")
        import re as _re
        company_clean = _re.sub(r"\s*\([^)]*\)\s*$", "", report.company_name or "").strip()
    return templates.TemplateResponse(
        request,
        "production_summary.html",
        {
            "report": report,
            "hour_slots": hour_slots,
            "cumulative_key": dpr.CUMULATIVE_KEY,
            "production_rows": production_rows,
            "rejection_groups": rejection_groups,
            "rework_rows": rework_rows,
            "cells": cells,
            "idle_events": idle_events,
            "totals": totals,
            "hour_models": hour_models,
            "hour_model_chain": hour_model_chain,
            "models_label": models_label,
            "production_by_model": production_by_model,
            "default_doc_no": default_doc_no,
            "default_rev_date": default_rev_date,
            "company_logo": company_logo,
            "company_clean": company_clean,
        },
    )


@app.post("/production/daily-report/delete-all-drafts")
async def dpr_delete_all_drafts(
    date: str | None = Form(None),
    date_from: str | None = Form(None),
    date_to: str | None = Form(None),
    redirect_to: str | None = Form(None),
):
    """Bulk-delete every draft DPR under the configured Tally company. The
    delete is scoped to whatever date filter the caller is showing — either an
    exact ``date`` or a ``date_from`` / ``date_to`` range (either bound is
    optional). Submitted reports are never touched."""
    company = _tally_company_name()
    with _session() as session:
        q = select(DailyProductionReport).where(
            DailyProductionReport.company_name == company,
            DailyProductionReport.status == "draft",
        )
        if date:
            q = q.where(DailyProductionReport.report_date == date)
        if date_from:
            q = q.where(DailyProductionReport.report_date >= date_from)
        if date_to:
            q = q.where(DailyProductionReport.report_date <= date_to)
        drafts = list(session.scalars(q))
        for r in drafts:
            session.delete(r)
        session.commit()
    if redirect_to:
        return RedirectResponse(redirect_to, status_code=303)
    from urllib.parse import urlencode
    qs = urlencode({"date": date} if date else {})
    target = "/production/daily-report" + (f"?{qs}" if qs else "")
    return RedirectResponse(target, status_code=303)


@app.post("/production/daily-report/{report_id}/delete")
async def dpr_delete(report_id: int):
    """Delete a Daily Production Report. Only drafts can be deleted — once a
    report is submitted, the row stays around so the published voucher can be
    reconciled against it."""
    with _session() as session:
        report = session.get(DailyProductionReport, report_id)
        if report is None:
            return RedirectResponse("/production/daily-report", status_code=303)
        if (report.status or "").strip().lower() != "draft":
            raise HTTPException(400, "Only draft reports can be deleted.")
        session.delete(report)
        session.commit()
    return RedirectResponse("/production/daily-report", status_code=303)


# ===========================================================================
# Process configuration (Daily Production Report row catalog)
# ===========================================================================

import re as _re_pp


def _stage_slugify(label: str) -> str:
    s = _re_pp.sub(r"[^a-z0-9]+", "_", label.lower()).strip("_")
    return s or "stage"


def _ps_load(session: Session, company: str) -> list[ProcessStage]:
    """Return active stages for a company, seeding from the static defaults on
    first use so the page is never empty."""
    rows = list(session.scalars(
        select(ProcessStage)
        .where(ProcessStage.company_name == company)
        .order_by(ProcessStage.sort_order, ProcessStage.id)
    ))
    if rows:
        return rows
    for idx, (key, label) in enumerate(dpr.STAGES):
        session.add(ProcessStage(
            company_name=company, key=key, label=label,
            sort_order=(idx + 1) * 10, active=True,
        ))
    session.commit()
    return list(session.scalars(
        select(ProcessStage)
        .where(ProcessStage.company_name == company)
        .order_by(ProcessStage.sort_order, ProcessStage.id)
    ))


def _ps_options(stages: list[ProcessStage]) -> list[tuple[str, str]]:
    return [(s.key, s.label) for s in stages if s.active]


def _ps_labels(stages: list[ProcessStage]) -> dict[str, str]:
    # Include inactive too so historical references still render with a label.
    return {s.key: s.label for s in stages}


def _ps_active_keys(session: Session, company: str) -> set[str]:
    return {s.key for s in _ps_load(session, company) if s.active}


def _shift_slug(label: str, taken: set[str]) -> str:
    """Generate a slot key from its label. Numeric labels (e.g. '8', '12:30')
    become 'h08', 'h1230'; alphabetic labels (e.g. 'OT') become a slug.
    Appends '_2', '_3', ... if collision."""
    raw = _re_pp.sub(r"[^a-z0-9]+", "", label.lower()) or "slot"
    if raw[0].isdigit():
        # zero-pad two digits if it's a single hour, keep as-is otherwise
        base = "h" + (raw.zfill(2) if len(raw) == 1 else raw)
    else:
        base = raw
    candidate = base
    n = 2
    while candidate in taken:
        candidate = f"{base}_{n}"
        n += 1
    return candidate


def _shifts_load(session: Session, company: str) -> list[Shift]:
    """Return the company's shift list, seeding from `dpr.SHIFT_OPTIONS` on
    first use so the page is never empty. Sorted by sort_order, then id."""
    rows = list(session.scalars(
        select(Shift)
        .where(Shift.company_name == company)
        .order_by(Shift.sort_order, Shift.id)
    ))
    if rows:
        return rows
    for idx, key in enumerate(dpr.SHIFT_OPTIONS):
        session.add(Shift(
            company_name=company, key=key, name=key, sort_order=(idx + 1) * 10,
        ))
    session.commit()
    return list(session.scalars(
        select(Shift)
        .where(Shift.company_name == company)
        .order_by(Shift.sort_order, Shift.id)
    ))


def _shift_keys(session: Session, company: str) -> list[str]:
    return [s.key for s in _shifts_load(session, company)]


def _shift_key_in_use(session: Session, company: str, key: str) -> int:
    """Count of DPR reports already pointing at this shift key — used to gate
    delete/rename of the shift itself."""
    from sqlalchemy import func
    return session.scalar(
        select(func.count(DailyProductionReport.id)).where(
            DailyProductionReport.company_name == company,
            DailyProductionReport.shift == key,
        )
    ) or 0


def _shift_make_key(name: str, taken: set[str]) -> str:
    raw = _re_pp.sub(r"[^A-Za-z0-9]+", "", name.upper()) or "SHIFT"
    candidate = raw[:20]
    n = 2
    while candidate in taken:
        suffix = f"_{n}"
        candidate = (raw[: 20 - len(suffix)] + suffix)
        n += 1
    return candidate


_TIME_RE = __import__("re").compile(r"^([01]?\d|2[0-3]):([0-5]\d)$")


def _normalize_time(raw: str | None) -> str | None:
    """Accept 'H:MM' or 'HH:MM' (24h), return canonical 'HH:MM' or None."""
    s = (raw or "").strip()
    if not s:
        return None
    m = _TIME_RE.match(s)
    if not m:
        raise HTTPException(400, f"invalid time {raw!r}; use HH:MM (24h)")
    return f"{int(m.group(1)):02d}:{m.group(2)}"


def _fmt_time_12h(hhmm: str) -> str:
    """'09:00' -> '9:00', '14:30' -> '2:30' (drops leading zero, 12h clock)."""
    h, mm = hhmm.split(":")
    h12 = int(h) % 12 or 12
    return f"{h12}:{mm}"


def _shift_label_from_times(from_t: str | None, to_t: str | None) -> str:
    """Derive a DPR header label from the From/To pair entered by the user.
    Convention: a slot's [from, to] is its actual interval. A boundary slot
    has only one side set — `to_time` alone means "Before X" (the slot ends
    at X with no defined start), `from_time` alone means "After X"."""
    if from_t and to_t:
        return f"{_fmt_time_12h(from_t)} - {_fmt_time_12h(to_t)}"
    if to_t and not from_t:
        return f"Before {_fmt_time_12h(to_t)}"
    if from_t and not to_t:
        return f"After {_fmt_time_12h(from_t)}"
    raise HTTPException(400, "From or To is required")


def _shift_load(session: Session, company: str, shift: str) -> list[ShiftPresetSlot]:
    """Return slots for (company, shift), seeding from the static SHIFT_PRESETS
    on first use so the page is never empty."""
    rows = list(session.scalars(
        select(ShiftPresetSlot)
        .where(ShiftPresetSlot.company_name == company, ShiftPresetSlot.shift == shift)
        .order_by(ShiftPresetSlot.sort_order, ShiftPresetSlot.id)
    ))
    if rows:
        return rows
    seed = dpr.SHIFT_PRESETS.get(shift) or dpr.HOUR_SLOTS
    for idx, (key, label) in enumerate(seed):
        session.add(ShiftPresetSlot(
            company_name=company, shift=shift, key=key, label=label,
            sort_order=(idx + 1) * 10,
        ))
    session.commit()
    return list(session.scalars(
        select(ShiftPresetSlot)
        .where(ShiftPresetSlot.company_name == company, ShiftPresetSlot.shift == shift)
        .order_by(ShiftPresetSlot.sort_order, ShiftPresetSlot.id)
    ))


def _pp_resolve_company(session: Session, requested: str | None) -> str:
    companies = _list_companies(session) or sorted(
        c for c in session.scalars(select(Voucher.company_name).distinct()) if c
    )
    return requested or (companies[0] if companies else "")


def _pp_load(
    session: Session, company: str, line: str, *, include_inactive: bool = False
) -> dict[str, list[ProductionProcess]]:
    q = (
        select(ProductionProcess)
        .where(ProductionProcess.company_name == company, ProductionProcess.line == line)
        .order_by(ProductionProcess.section, ProductionProcess.sort_order, ProductionProcess.id)
    )
    if not include_inactive:
        q = q.where(ProductionProcess.active.is_(True))
    out: dict[str, list[ProductionProcess]] = {s: [] for s, _ in dpr.SECTION_OPTIONS}
    for row in session.scalars(q):
        out.setdefault(row.section, []).append(row)
    return out


# ---------------------------------------------------------------------------
# Production model specs — physical dimensions used for SJ-LINE scrap weight.
# ---------------------------------------------------------------------------

def _parse_spec_value(raw: str | None) -> float:
    if raw is None or str(raw).strip() == "":
        return 0.0
    try:
        return float(str(raw).strip())
    except ValueError:
        raise HTTPException(400, f"invalid number: {raw!r}")


_HOLE_HEADER_RE = __import__("re").compile(r"^\s*(\d+(?:\.\d+)?)\s*mm\s*$", __import__("re").IGNORECASE)


def _detect_header_row(rows: list[list[str]]) -> int:
    """Find the row index whose cells include 'model' and 'length' headers.

    Handles workbooks with merged title rows above the real header (e.g.
    "DRILL BITS AND SEGMENTS" spanning the size columns). Falls back to row 0
    so the existing flat-CSV path keeps working."""
    for i, row in enumerate(rows[:8]):
        joined = " | ".join(str(c or "").strip().lower() for c in row)
        if "model" in joined and ("length" in joined or "l (mm)" in joined):
            return i
    return 0


def _spec_columns_from_header(header: list[str]) -> tuple[dict[str, int], list[tuple[int, float]]]:
    """Return ((named-column index map), [(col_index, hole_diameter_mm), ...]).

    The named columns capture the base spec fields (model + L/W/thickness +
    optional notes). The list of hole-columns captures every header that
    matches `<number> MM`, with the parsed diameter — these become rows in
    ProductionModelHole when their cell value is > 0."""
    norm = [str(h or "").strip().lower() for h in header]
    def find(*needles: str) -> int | None:
        for i, h in enumerate(norm):
            if all(n in h for n in needles):
                return i
        return None
    cols: dict[str, int | None] = {
        "customer": find("customer") if find("customer") is not None else find("brand"),
        "model": find("model"),
        "length": find("length") if find("length") is not None else find("l", "mm"),
        "width": find("width") if find("width") is not None else find("w", "mm"),
        "thickness": find("thick"),
        "notes": find("notes") if find("notes") is not None else find("remark"),
        "blank_price": find("blank", "price"),
        "drilled_price": find("drilled", "price"),
        "printed_price": find("printed", "price"),
    }
    missing = [k for k in ("model", "length", "width", "thickness") if cols[k] is None]
    if missing:
        raise HTTPException(
            400,
            f"upload missing required columns: {', '.join(missing)}. "
            f"Got headers: {header!r}",
        )
    hole_cols: list[tuple[int, float]] = []
    skip_idx = {v for v in cols.values() if v is not None}
    for i, raw in enumerate(header):
        if i in skip_idx:
            continue
        m = _HOLE_HEADER_RE.match(str(raw or ""))
        if m:
            try:
                hole_cols.append((i, float(m.group(1))))
            except ValueError:
                continue
    return ({k: v for k, v in cols.items() if v is not None}, hole_cols)


def _tally_company_name() -> str:
    """The Tally company whose stock-group / stock-item master we cross-check
    model specs against. Driven by ``TALLY_COMPANY_NAME`` in the env, with a
    sensible default for the current shop."""
    import os as _os
    return (_os.getenv("TALLY_COMPANY_NAME") or "Avinash Appliances Private Limited (April 26 - March 27)").strip()


_BLANK_GLASS_PREFIX = "BLANK GLASS - "
_DRILLED_GLASS_PREFIX = "DRILLED GLASS - "


def _model_stock_presence(session: Session) -> tuple[set[str], set[str]]:
    """Scan Tally's stock items under ``TALLY_COMPANY_NAME`` for entries
    named ``BLANK GLASS - <model>`` and ``DRILLED GLASS - <model>``. Returns
    two sets of bare model names (uppercased for case-insensitive matching):
    one for the BLANK side, one for the DRILLED side."""
    company = _tally_company_name()
    if not company:
        return set(), set()
    rows = session.scalars(
        select(StockItem.name).where(StockItem.company_name == company)
    )
    blank_prefix_lower = _BLANK_GLASS_PREFIX.lower()
    drilled_prefix_lower = _DRILLED_GLASS_PREFIX.lower()
    blank: set[str] = set()
    drilled: set[str] = set()
    for name in rows:
        if not name:
            continue
        s = str(name).strip()
        low = s.lower()
        if low.startswith(blank_prefix_lower):
            tail = s[len(_BLANK_GLASS_PREFIX):].strip().upper()
            if tail:
                blank.add(tail)
        elif low.startswith(drilled_prefix_lower):
            tail = s[len(_DRILLED_GLASS_PREFIX):].strip().upper()
            if tail:
                drilled.add(tail)
    return blank, drilled


def _model_missing_status(
    session: Session, models: list[str]
) -> dict[str, str]:
    """Per-model readiness against Tally's BLANK / DRILLED stock items
    *and* the per-spec price columns. A model is ``ok`` only when both
    stock items exist AND both blank_price + drilled_price are populated;
    anything else maps to a tag the UI can render.

    Status values (precedence: stock-missing first, then price-missing):
      - ``ok``               — both items exist AND both prices set
      - ``blank_missing``    — DRILLED found, BLANK missing             → BG-NF
      - ``drilled_missing``  — BLANK found, DRILLED missing             → DG-NF
      - ``both_missing``     — neither stock item found                  → red bg, no tag
      - ``no_price``         — both stock items found, but a price is 0/None → NP
    """
    blank_set, drilled_set = _model_stock_presence(session)
    # Price lookup keyed by uppercased model name for case-insensitive match.
    price_rows = session.execute(
        select(
            ProductionModelSpec.model,
            ProductionModelSpec.blank_price,
            ProductionModelSpec.drilled_price,
        )
    ).all()
    prices: dict[str, tuple[float | None, float | None]] = {
        (m or "").strip().upper(): (bp, dp) for m, bp, dp in price_rows
    }
    out: dict[str, str] = {}
    for m in models:
        key = (m or "").strip().upper()
        has_b = key in blank_set
        has_d = key in drilled_set
        if not has_b and not has_d:
            out[m] = "both_missing"
        elif not has_b:
            out[m] = "blank_missing"
        elif not has_d:
            out[m] = "drilled_missing"
        else:
            bp, dp = prices.get(key, (None, None))
            if not bp or not dp:
                out[m] = "no_price"
            else:
                out[m] = "ok"
    return out


@app.get("/production/models", response_class=HTMLResponse)
def model_specs_index(request: Request, company: str | None = None):
    """Flat grid of every model across every company.

    The `company` query-param is still accepted as the default target for
    the upload form, but the grid itself is no longer filtered — Customer
    is rendered as the first column and the user can narrow it via the
    Excel-style filter dropdown on that column header."""
    from sqlalchemy.orm import selectinload
    with _session() as session:
        companies = _list_companies(session) or sorted(
            c for c in session.scalars(select(Voucher.company_name).distinct()) if c
        )
        upload_company = _pp_resolve_company(session, company)
        specs = list(session.scalars(
            select(ProductionModelSpec)
            .options(selectinload(ProductionModelSpec.holes))
            .order_by(ProductionModelSpec.company_name, ProductionModelSpec.model)
        ))
        model_status = _model_missing_status(session, [s.model for s in specs])
        last_sync_at = max((s.updated_at for s in specs), default=None)
    diameter_set: set[float] = set()
    counts: dict[int, dict[float, int]] = {}
    for s in specs:
        per: dict[float, int] = {}
        for h in s.holes:
            diameter_set.add(h.diameter_mm)
            per[h.diameter_mm] = h.count
        counts[s.id] = per
    all_diameters = sorted(diameter_set)
    customers = sorted({s.company_name for s in specs if s.company_name})
    return templates.TemplateResponse(
        request,
        "model_specs.html",
        {
            "companies": companies,
            "customers": customers,
            "upload_company": upload_company,
            "specs": specs,
            "all_diameters": all_diameters,
            "counts": counts,
            "model_status": model_status,
            "tally_company_name": _tally_company_name(),
            "last_sync_at": last_sync_at,
            "onedrive_configured": onedrive_client.is_configured(),
            "onedrive_signed_in": onedrive_client.has_token(),
            "ms_worksheet_default": onedrive_client.default_worksheet_name(),
            "ms_table_default": onedrive_client.default_table_name(),
            "edit_unlocked": _models_edit_unlocked(request),
            "edit_lock_enabled": bool((__import__("os").getenv("MODEL_SPECS_EDIT_PASSWORD") or "").strip()),
        },
    )


def _models_edit_password() -> str:
    import os as _os
    return (_os.getenv("MODEL_SPECS_EDIT_PASSWORD") or "").strip()


def _models_edit_unlocked(request: Request) -> bool:
    """Legacy session-cookie check, retained so the page-level read can still
    show/hide the (now obsolete) Enable-Edit pill. Action-level enforcement
    happens in ``_check_form_password`` per save/delete request."""
    pw = _models_edit_password()
    if not pw:
        return True
    return request.cookies.get("models_edit") == "1"


def _check_form_password(form) -> bool:
    """Strict per-action gate: the password posted with the form must match
    ``MODEL_SPECS_EDIT_PASSWORD``. If the env var is unset, the lock is off."""
    expected = _models_edit_password()
    if not expected:
        return True
    supplied = (form.get("edit_password") or "").strip()
    return supplied == expected


@app.post("/production/models/verify-password")
async def model_specs_verify_password(request: Request):
    """JSON endpoint used by the Model Specs page to verify the password
    before opening Add / Edit / Delete dialogs. Returns ``{"ok": bool}``."""
    form = await request.form()
    expected = _models_edit_password()
    supplied = (form.get("password") or "").strip()
    if not expected:
        return JSONResponse({"ok": True, "lock_disabled": True})
    return JSONResponse({"ok": supplied == expected})


@app.post("/production/models/unlock")
async def model_specs_unlock(request: Request):
    import os as _os
    form = await request.form()
    pw_supplied = (form.get("password") or "").strip()
    pw_expected = (_os.getenv("MODEL_SPECS_EDIT_PASSWORD") or "").strip()
    company = (form.get("company") or "").strip()
    from urllib.parse import urlencode
    if not pw_expected:
        # No password configured — nothing to unlock; redirect with note.
        qs = urlencode({"company": company, "msg": "Edit lock is disabled (no MODEL_SPECS_EDIT_PASSWORD set)."})
        return RedirectResponse(f"/production/models?{qs}", status_code=303)
    if pw_supplied != pw_expected:
        qs = urlencode({"company": company, "msg": "Wrong password."})
        return RedirectResponse(f"/production/models?{qs}", status_code=303)
    qs = urlencode({"company": company, "msg": "Editing enabled."})
    resp = RedirectResponse(f"/production/models?{qs}", status_code=303)
    # Session cookie (no expires/max-age) — cleared when browser closes.
    resp.set_cookie("models_edit", "1", httponly=True, samesite="lax")
    return resp


@app.post("/production/models/lock")
async def model_specs_lock(request: Request):
    form = await request.form()
    company = (form.get("company") or "").strip()
    from urllib.parse import urlencode
    qs = urlencode({"company": company, "msg": "Editing locked."})
    resp = RedirectResponse(f"/production/models?{qs}", status_code=303)
    resp.delete_cookie("models_edit")
    return resp


@app.post("/production/models/save")
async def model_specs_save(request: Request):
    """Upsert one model's base spec. Optional ``hole_diameter_mm`` /
    ``hole_count`` form arrays let the same form add multiple hole rows in
    one POST. The form may carry ``spec_id`` to identify an existing row
    for in-place edit (Customer / Model rename); without it, lookup falls
    back to (company, model) and creates a new spec when no match exists.
    Holes are wiped and re-written when ``replace_holes=1`` is present."""
    form = await request.form()
    if not _check_form_password(form):
        raise HTTPException(403, "Wrong or missing password — model specs edit is locked.")
    company = (form.get("company") or "").strip()
    model = (form.get("model") or "").strip()
    if not company or not model:
        raise HTTPException(400, "company and model are required")
    raw_spec_id = (form.get("spec_id") or "").strip()
    spec_id = int(raw_spec_id) if raw_spec_id.isdigit() else None
    diameters = form.getlist("hole_diameter_mm")
    counts = form.getlist("hole_count")
    replace_holes = form.get("replace_holes") in ("1", "true", "on")
    with _session() as session:
        spec = None
        if spec_id is not None:
            spec = session.get(ProductionModelSpec, spec_id)
        if spec is None:
            spec = session.scalars(
                select(ProductionModelSpec).where(
                    ProductionModelSpec.company_name == company,
                    ProductionModelSpec.model == model,
                )
            ).one_or_none()
        is_new = spec is None
        if is_new:
            # Block creating a second row that collides with an existing one.
            clash = session.scalars(
                select(ProductionModelSpec).where(
                    ProductionModelSpec.company_name == company,
                    ProductionModelSpec.model == model,
                )
            ).one_or_none()
            if clash is not None:
                from urllib.parse import urlencode
                qs = urlencode({"company": company, "msg": f"A spec for {company} / {model} already exists — edit that row instead."})
                return RedirectResponse(f"/production/models?{qs}", status_code=303)
            spec = ProductionModelSpec(company_name=company, model=model, length_mm=0, width_mm=0, thickness_mm=0)
            session.add(spec)
        else:
            # Allow Customer / Model rename, but block renaming into an existing pair.
            if (spec.company_name, spec.model) != (company, model):
                clash = session.scalars(
                    select(ProductionModelSpec).where(
                        ProductionModelSpec.company_name == company,
                        ProductionModelSpec.model == model,
                        ProductionModelSpec.id != spec.id,
                    )
                ).one_or_none()
                if clash is not None:
                    from urllib.parse import urlencode
                    qs = urlencode({"company": company, "msg": f"Cannot rename to {company} / {model} — another spec already uses that pair."})
                    return RedirectResponse(f"/production/models?{qs}", status_code=303)
            spec.company_name = company
            spec.model = model
        spec.length_mm = _parse_spec_value(form.get("length_mm"))
        spec.width_mm = _parse_spec_value(form.get("width_mm"))
        spec.thickness_mm = _parse_spec_value(form.get("thickness_mm"))
        spec.notes = (form.get("notes") or "").strip() or None
        session.flush()  # ensure spec.id

        if is_new or replace_holes:
            for h in list(spec.holes):
                session.delete(h)
            session.flush()
        # Merge inline hole rows (from the Add form's + button), summing
        # duplicates so the same diameter twice doesn't violate the unique
        # constraint.
        merged: dict[float, int] = {}
        for d_raw, c_raw in zip(diameters, counts):
            try:
                d = float(str(d_raw).strip())
                c = int(float(str(c_raw).strip()))
            except (TypeError, ValueError):
                continue
            if d <= 0 or c <= 0:
                continue
            merged[d] = merged.get(d, 0) + c
        for d, c in merged.items():
            existing = session.scalars(
                select(ProductionModelHole).where(
                    ProductionModelHole.spec_id == spec.id,
                    ProductionModelHole.diameter_mm == d,
                )
            ).one_or_none()
            if existing is None:
                session.add(ProductionModelHole(spec_id=spec.id, diameter_mm=d, count=c))
            else:
                existing.count = c
        session.commit()
    from urllib.parse import urlencode
    return RedirectResponse(f"/production/models?{urlencode({'company': company})}", status_code=303)


@app.post("/production/models/holes/save")
async def model_holes_save(request: Request):
    form = await request.form()
    if not _check_form_password(form):
        raise HTTPException(403, "Wrong or missing password — model specs edit is locked.")
    spec_id = int(form.get("spec_id") or 0)
    company = (form.get("company") or "").strip()
    diameter = _parse_spec_value(form.get("diameter_mm"))
    count = int(_parse_spec_value(form.get("count")))
    if spec_id <= 0 or diameter <= 0 or count <= 0:
        raise HTTPException(400, "spec_id, diameter_mm > 0 and count > 0 are required")
    with _session() as session:
        existing = session.scalars(
            select(ProductionModelHole).where(
                ProductionModelHole.spec_id == spec_id,
                ProductionModelHole.diameter_mm == diameter,
            )
        ).one_or_none()
        if existing is None:
            session.add(ProductionModelHole(spec_id=spec_id, diameter_mm=diameter, count=count))
        else:
            existing.count = count
        session.commit()
    from urllib.parse import urlencode
    return RedirectResponse(f"/production/models?{urlencode({'company': company})}", status_code=303)


@app.post("/production/models/holes/delete")
async def model_holes_delete(request: Request):
    form = await request.form()
    if not _check_form_password(form):
        raise HTTPException(403, "Wrong or missing password — model specs edit is locked.")
    hole_id = int(form.get("hole_id") or 0)
    company = (form.get("company") or "").strip()
    with _session() as session:
        h = session.get(ProductionModelHole, hole_id)
        if h is not None:
            session.delete(h)
            session.commit()
    from urllib.parse import urlencode
    return RedirectResponse(f"/production/models?{urlencode({'company': company})}", status_code=303)


@app.post("/production/models/delete")
async def model_specs_delete(request: Request):
    form = await request.form()
    if not _check_form_password(form):
        raise HTTPException(403, "Wrong or missing password — model specs edit is locked.")
    spec_id = int(form.get("spec_id") or 0)
    company = (form.get("company") or "").strip()
    with _session() as session:
        spec = session.get(ProductionModelSpec, spec_id)
        if spec is not None:
            session.delete(spec)
            session.commit()
    from urllib.parse import urlencode
    return RedirectResponse(f"/production/models?{urlencode({'company': company})}", status_code=303)


@app.post("/production/models/upload")
async def model_specs_upload(
    request: Request,
    worksheet_name: str = Form(""),
    table_name: str = Form(""),
    edit_password: str = Form(""),
    file: UploadFile = File(...),
):
    """Bulk-upsert specs from an .xlsx or .csv export of the OneDrive sheet.

    Optional ``worksheet_name`` picks a specific sheet (default: the active
    sheet). Optional ``table_name`` (xlsx only) restricts parsing to a named
    Excel table on that sheet, so anything outside the table — title rows,
    summary blocks — is ignored. The sheet's "Customer" column is required
    per-row; rows without one are skipped and reported as errors."""
    expected_pw = _models_edit_password()
    if expected_pw and (edit_password or "").strip() != expected_pw:
        raise HTTPException(403, "Wrong or missing password — model specs upload is locked.")
    raw = await file.read()
    name = (file.filename or "").lower()
    ws_arg = worksheet_name.strip()
    tbl_arg = table_name.strip()
    rows: list[list[str]]
    if name.endswith(".csv"):
        import csv, io
        text = raw.decode("utf-8-sig", errors="replace")
        rows = list(csv.reader(io.StringIO(text)))
    elif name.endswith(".xlsx") or name.endswith(".xls"):
        try:
            from openpyxl import load_workbook  # type: ignore
            from openpyxl.utils import range_boundaries  # type: ignore
        except ImportError:
            raise HTTPException(500, "openpyxl is not installed; install it or upload a CSV instead")
        import io
        # ``read_only=True`` doesn't expose worksheet.tables, so we open in
        # normal mode when a table_name is given.
        wb = load_workbook(filename=io.BytesIO(raw), data_only=True, read_only=not bool(tbl_arg))
        if ws_arg:
            if ws_arg not in wb.sheetnames:
                raise HTTPException(400, f"worksheet {ws_arg!r} not found in workbook (sheets: {wb.sheetnames})")
            ws = wb[ws_arg]
        else:
            ws = wb.active
        if tbl_arg:
            tables = getattr(ws, "tables", {}) or {}
            if tbl_arg not in tables:
                raise HTTPException(400, f"table {tbl_arg!r} not found on sheet {ws.title!r} (tables: {list(tables.keys())})")
            ref = tables[tbl_arg].ref if hasattr(tables[tbl_arg], "ref") else tables[tbl_arg]
            min_col, min_row, max_col, max_row = range_boundaries(ref)
            rows = []
            for r in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
                rows.append([("" if c is None else str(c)) for c in r])
        else:
            rows = [[("" if c is None else str(c)) for c in r] for r in ws.iter_rows(values_only=True)]
    else:
        raise HTTPException(400, "upload must be .xlsx or .csv")
    counts = _apply_master_data_rows(rows)
    msg_parts = [f"inserted={counts['inserted']}", f"updated={counts['updated']}", f"holes={counts['holes_written']}"]
    if counts["skipped"]:
        msg_parts.append(f"skipped={counts['skipped']}")
    if counts["errors"]:
        msg_parts.append(f"errors={len(counts['errors'])}: {counts['errors'][0]!r}")
    from urllib.parse import urlencode
    qs = urlencode({"msg": " ".join(msg_parts)})
    return RedirectResponse(f"/production/models?{qs}", status_code=303)


def _apply_master_data_rows(rows: list[list[str]]) -> dict:
    """Upsert ProductionModelSpec rows from a header+data matrix.

    Used by both the manual xlsx/csv upload path and the OneDrive sync. Each
    row must carry a Customer value — rows without one are skipped and
    surfaced as errors so nothing slips through with a wrong owner."""
    rows = [r for r in rows if any((c or "").strip() for c in r)]
    if not rows:
        raise HTTPException(400, "no data rows found")
    header_idx = _detect_header_row(rows)
    header = rows[header_idx]
    data_rows = rows[header_idx + 1:]
    cols, hole_cols = _spec_columns_from_header(header)
    inserted = updated = skipped = holes_written = 0
    errors: list[str] = []
    with _session() as session:
        for idx, row in enumerate(data_rows, start=header_idx + 2):
            try:
                model_idx = cols.get("model")
                model = (row[model_idx] or "").strip() if model_idx is not None and len(row) > model_idx else ""
                if not model:
                    skipped += 1
                    continue
                def cell(key: str) -> str:
                    i = cols.get(key)
                    return row[i] if i is not None and i < len(row) else ""
                length = _parse_spec_value(cell("length"))
                width = _parse_spec_value(cell("width"))
                thick = _parse_spec_value(cell("thickness"))
                if length <= 0 or width <= 0 or thick <= 0:
                    errors.append(f"row {idx}: missing length/width/thickness for {model}")
                    continue
                notes = (cell("notes") or "").strip() or None
                def price(key: str) -> float | None:
                    raw = (cell(key) or "").strip()
                    if not raw:
                        return None
                    try:
                        return float(raw.replace(",", ""))
                    except ValueError:
                        return None
                blank_price = price("blank_price")
                drilled_price = price("drilled_price")
                printed_price = price("printed_price")
                row_customer = (cell("customer") or "").strip()
                if not row_customer:
                    errors.append(f"row {idx}: missing Customer for model {model!r} — row skipped")
                    skipped += 1
                    continue
                # Look up by model name globally — a model belongs to exactly
                # one customer in this domain, so duplicates across customers
                # shouldn't exist. If one is found, we update its customer too.
                spec = session.scalars(
                    select(ProductionModelSpec).where(ProductionModelSpec.model == model)
                ).one_or_none()
                if spec is None:
                    spec = ProductionModelSpec(
                        company_name=row_customer, model=model,
                        length_mm=length, width_mm=width, thickness_mm=thick,
                        blank_price=blank_price, drilled_price=drilled_price,
                        printed_price=printed_price, notes=notes,
                    )
                    session.add(spec)
                    session.flush()  # get spec.id
                    inserted += 1
                else:
                    spec.company_name = row_customer
                    spec.length_mm = length
                    spec.width_mm = width
                    spec.thickness_mm = thick
                    spec.blank_price = blank_price
                    spec.drilled_price = drilled_price
                    spec.printed_price = printed_price
                    if notes is not None:
                        spec.notes = notes
                    # Replace existing holes with the uploaded set so a column
                    # going to 0 actually removes that diameter.
                    for h in list(spec.holes):
                        session.delete(h)
                    session.flush()
                    updated += 1
                for col_i, diameter in hole_cols:
                    raw_count = row[col_i] if col_i < len(row) else ""
                    try:
                        count_val = int(_parse_spec_value(raw_count))
                    except HTTPException:
                        count_val = 0
                    if count_val > 0:
                        session.add(ProductionModelHole(
                            spec_id=spec.id, diameter_mm=diameter, count=count_val,
                        ))
                        holes_written += 1
            except HTTPException:
                raise
            except Exception as exc:
                errors.append(f"row {idx}: {type(exc).__name__}: {exc}")
        session.commit()
    return {
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "holes_written": holes_written,
        "errors": errors,
    }


@app.get("/oauth/microsoft/login")
def oauth_ms_login():
    if not onedrive_client.is_configured():
        raise HTTPException(400, "Microsoft Graph env vars not set in .env (MS_CLIENT_ID / MS_TENANT_ID / MS_CLIENT_SECRET / MS_WORKBOOK_PATH)")
    return RedirectResponse(onedrive_client.build_auth_url(), status_code=303)


@app.get("/oauth/microsoft/callback")
def oauth_ms_callback(code: str = "", state: str = "", error: str = "", error_description: str = ""):
    if error:
        raise HTTPException(400, f"Microsoft sign-in returned error: {error} — {error_description}")
    if not code:
        raise HTTPException(400, "missing authorization code")
    try:
        onedrive_client.exchange_code(code, state)
    except onedrive_client.OneDriveError as exc:
        raise HTTPException(400, str(exc))
    return RedirectResponse("/production/models?msg=OneDrive+connected", status_code=303)


@app.post("/oauth/microsoft/signout")
def oauth_ms_signout():
    onedrive_client.sign_out()
    return RedirectResponse("/production/models?msg=Signed+out+of+OneDrive", status_code=303)


@app.post("/production/models/sync")
def model_specs_sync(
    worksheet_name: str = Form(""),
    table_name: str = Form(""),
    edit_password: str = Form(""),
):
    """Pull a named Excel table live from OneDrive and upsert the specs."""
    expected_pw = _models_edit_password()
    if expected_pw and (edit_password or "").strip() != expected_pw:
        raise HTTPException(403, "Wrong or missing password — model specs sync is locked.")
    if not onedrive_client.is_configured():
        raise HTTPException(400, "Microsoft Graph env vars not set in .env")
    if not onedrive_client.has_token():
        return RedirectResponse("/oauth/microsoft/login", status_code=303)
    try:
        rows = onedrive_client.fetch_master_table_rows(
            worksheet_name=worksheet_name or None,
            table_name=table_name or None,
        )
    except onedrive_client.OneDriveError as exc:
        from urllib.parse import urlencode
        qs = urlencode({"msg": f"OneDrive sync failed: {exc}"})
        return RedirectResponse(f"/production/models?{qs}", status_code=303)
    counts = _apply_master_data_rows(rows)
    msg_parts = [
        "OneDrive sync:",
        f"inserted={counts['inserted']}",
        f"updated={counts['updated']}",
        f"holes={counts['holes_written']}",
    ]
    if counts["skipped"]:
        msg_parts.append(f"skipped={counts['skipped']}")
    if counts["errors"]:
        msg_parts.append(f"errors={len(counts['errors'])}: {counts['errors'][0]!r}")
    from urllib.parse import urlencode
    qs = urlencode({"msg": " ".join(msg_parts)})
    return RedirectResponse(f"/production/models?{qs}", status_code=303)


# ---------------------------------------------------------------------------
# Consolidated daily SJ - LINE voucher publish flow.
# ---------------------------------------------------------------------------

def _ldvp_remote_id(company: str, report_date: str) -> str:
    return f"sjline-{company}-{report_date}".replace(" ", "_")


def _ldvp_latest(session: Session, company: str, report_date: str) -> LineDailyVoucherPost | None:
    rows = list(session.scalars(
        select(LineDailyVoucherPost)
        .where(
            LineDailyVoucherPost.company_name == company,
            LineDailyVoucherPost.report_date == report_date,
        )
        .order_by(LineDailyVoucherPost.created_at.desc())
    ))
    return rows[0] if rows else None


@app.get("/production/line-voucher", response_class=HTMLResponse)
def line_publish_preview(
    request: Request,
    company: str,
    date: str,
    scrap_rate: float | None = None,
):
    """Preview-and-confirm page for the consolidated SJ - LINE voucher."""
    rate = scrap_rate if scrap_rate is not None else line_voucher.DEFAULT_SCRAP_RATE
    with _session() as session:
        result = line_voucher.build_sj_line_voucher(
            session, company, date,
            scrap_rate=rate,
            remote_id=_ldvp_remote_id(company, date),
        )
        latest = _ldvp_latest(session, company, date)
    # Split entries for display in two columns matching the Tally SJ layout.
    consume = [e for e in result.voucher["inventory_entries"] if e.get("direction") == "out"]
    produce = [e for e in result.voucher["inventory_entries"] if e.get("direction") == "in"]
    return templates.TemplateResponse(
        request,
        "line_publish.html",
        {
            "company": company,
            "report_date": date,
            "scrap_rate": rate,
            "result": result,
            "consume": consume,
            "produce": produce,
            "latest": latest,
        },
    )


@app.post("/production/line-voucher")
async def line_publish_submit(request: Request):
    form = await request.form()
    company = (form.get("company") or "").strip()
    date_ = (form.get("date") or "").strip()
    try:
        rate = float(form.get("scrap_rate") or line_voucher.DEFAULT_SCRAP_RATE)
    except ValueError:
        raise HTTPException(400, "scrap_rate must be a number")
    if not company or not date_:
        raise HTTPException(400, "company and date are required")
    force_repost = (form.get("force_repost") or "").strip() in ("1", "true", "on", "yes")
    remote_id = _ldvp_remote_id(company, date_)
    with _session() as session:
        # Idempotency gate: if the SJ - LINE voucher for this (company, date)
        # has already been posted to Tally, refuse to post a second time
        # unless the user has explicitly opted into ``force_repost`` (which
        # they only get from the "Force re-post" button after acknowledging
        # they've deleted the prior voucher in Tally).
        already_posted = session.scalars(
            select(LineDailyVoucherPost).where(
                LineDailyVoucherPost.company_name == company,
                LineDailyVoucherPost.report_date == date_,
                LineDailyVoucherPost.status == "posted",
            )
        ).first()
        if already_posted is not None and not force_repost:
            from urllib.parse import urlencode
            qs = urlencode({
                "company": company,
                "date": date_,
                "msg": (
                    f"Voucher already posted on "
                    f"{already_posted.posted_at.strftime('%Y-%m-%d %H:%M') if already_posted.posted_at else 'previously'} "
                    f"(#{already_posted.tally_voucher_number or already_posted.tally_master_id}). "
                    "Re-posting is blocked to avoid duplicate vouchers in Tally."
                ),
            })
            return RedirectResponse(
                f"/production/line-voucher?{qs}",
                status_code=303,
            )
        result = line_voucher.build_sj_line_voucher(
            session, company, date_, scrap_rate=rate, remote_id=remote_id,
        )
        if not result.voucher["inventory_entries"]:
            raise HTTPException(400, "No production data to post for this date")
        post = LineDailyVoucherPost(
            company_name=company,
            report_date=date_,
            remote_id=f"{remote_id}-{int(datetime.utcnow().timestamp())}",
            status="pending",
            scrap_rate=rate,
        )
        session.add(post)
        session.commit()
        post_id = post.id

        try:
            client = _tally_client()
            api_result = client.import_voucher(company, result.voucher, dry_run=False)
        except Exception as exc:
            post.status = "failed"
            post.tally_error = f"{type(exc).__name__}: {exc}"
            session.commit()
            from urllib.parse import urlencode
            return RedirectResponse(
                f"/production/line-voucher?{urlencode({'company': company, 'date': date_})}",
                status_code=303,
            )

        if api_result.get("ok"):
            post.status = "posted"
            post.posted_at = datetime.utcnow()
            vch_id = api_result.get("last_vch_id")
            post.tally_master_id = str(vch_id) if vch_id is not None else None
            vch_no = None
            if vch_id is not None:
                try:
                    vch_no = client.fetch_voucher_number_by_master_id(company, vch_id)
                except Exception:
                    vch_no = None
            post.tally_voucher_number = vch_no or (str(vch_id) if vch_id is not None else None)
            # Mark every DPR row that contributed to this day's voucher as
            # ``submitted`` so it locks against further edits and so the
            # Recent reports list distinguishes posted-to-Tally from saved.
            now = datetime.utcnow()
            for r in session.scalars(
                select(DailyProductionReport).where(
                    DailyProductionReport.company_name == company,
                    DailyProductionReport.report_date == date_,
                )
            ):
                r.status = "submitted"
                r.submitted_at = r.submitted_at or now
                r.updated_at = now
        else:
            post.status = "failed"
            parts = []
            if api_result.get("line_error"):
                parts.append(f"LINEERROR: {api_result['line_error']}")
            if api_result.get("exception"):
                parts.append(f"EXCEPTIONS: {api_result['exception']}")
            parts.append(
                f"created={api_result.get('created')} altered={api_result.get('altered')} "
                f"errors={api_result.get('errors')}"
            )
            post.tally_error = " | ".join(parts)
        session.commit()
    from urllib.parse import urlencode
    return RedirectResponse(
        f"/production/line-voucher?{urlencode({'company': company, 'date': date_})}",
        status_code=303,
    )


@app.get("/production/processes", response_class=HTMLResponse)
def pp_index(request: Request, company: str | None = None):
    with _session() as session:
        companies = _list_companies(session) or sorted(
            c for c in session.scalars(select(Voucher.company_name).distinct()) if c
        )
        company = _pp_resolve_company(session, company)
        # Per-line summary: count rows per section.
        line_summaries = []
        for line in dpr.LINE_OPTIONS:
            counts = {sec: 0 for sec, _ in dpr.SECTION_OPTIONS}
            inactive = 0
            for row in session.scalars(
                select(ProductionProcess).where(
                    ProductionProcess.company_name == company,
                    ProductionProcess.line == line,
                )
            ):
                if row.active:
                    counts[row.section] = counts.get(row.section, 0) + 1
                else:
                    inactive += 1
            total = sum(counts.values())
            line_summaries.append({
                "line": line,
                "counts": counts,
                "total": total,
                "inactive": inactive,
            })
        catalog_counts = {s: 0 for s, _ in dpr.SECTION_OPTIONS}
        for e in session.scalars(
            select(ProcessCatalogEntry).where(ProcessCatalogEntry.company_name == company)
        ):
            catalog_counts[e.section] = catalog_counts.get(e.section, 0) + 1
        catalog_total = sum(catalog_counts.values())
        shift_rows = _shifts_load(session, company)
        shift_counts = {s.key: 0 for s in shift_rows}
        for s in session.scalars(
            select(ShiftPresetSlot).where(ShiftPresetSlot.company_name == company)
        ):
            shift_counts[s.shift] = shift_counts.get(s.shift, 0) + 1
    return templates.TemplateResponse(
        request,
        "process_config_index.html",
        {
            "companies": companies,
            "selected_company": company,
            "line_summaries": line_summaries,
            "section_options": dpr.SECTION_OPTIONS,
            "catalog_counts": catalog_counts,
            "catalog_total": catalog_total,
            "shift_options": [(s.key, s.name) for s in shift_rows],
            "shift_counts": shift_counts,
        },
    )


@app.get("/production/processes/{line}", response_class=HTMLResponse)
def pp_line(
    request: Request,
    line: str,
    company: str | None = None,
    show_inactive: int = 0,
    edit: int | None = None,
):
    if line not in dpr.LINE_OPTIONS:
        raise HTTPException(404, "unknown line")
    with _session() as session:
        company = _pp_resolve_company(session, company)
        stages = _ps_load(session, company)
        sections = _pp_load(session, company, line, include_inactive=bool(show_inactive))
        edit_row = session.get(ProductionProcess, edit) if edit else None
        if edit_row and (edit_row.company_name != company or edit_row.line != line):
            edit_row = None
        catalog_by_section: dict[str, list[ProcessCatalogEntry]] = {s: [] for s, _ in dpr.SECTION_OPTIONS}
        for e in session.scalars(
            select(ProcessCatalogEntry)
            .where(ProcessCatalogEntry.company_name == company)
            .order_by(ProcessCatalogEntry.section, ProcessCatalogEntry.group_label, ProcessCatalogEntry.label)
        ):
            catalog_by_section.setdefault(e.section, []).append(e)
    return templates.TemplateResponse(
        request,
        "process_config_line.html",
        {
            "company": company,
            "line": line,
            "sections": sections,
            "section_options": dpr.SECTION_OPTIONS,
            "stage_options": _ps_options(stages),
            "stage_labels": _ps_labels(stages),
            "show_inactive": bool(show_inactive),
            "edit_row": edit_row,
            "catalog_by_section": catalog_by_section,
        },
    )


@app.get("/production/shift-presets", response_class=HTMLResponse)
def sp_index(request: Request, company: str | None = None):
    from types import SimpleNamespace
    with _session() as session:
        company = _pp_resolve_company(session, company)
        # Seed both tables up-front so subsequent reads don't trigger
        # commits that would expire already-loaded rows.
        shift_rows_orm = _shifts_load(session, company)
        slot_rows_by_key: dict[str, list] = {
            s.key: _shift_load(session, company, s.key) for s in shift_rows_orm
        }
        usage = {s.key: _shift_key_in_use(session, company, s.key) for s in shift_rows_orm}
        # Materialize plain holders so the template doesn't dereference ORM
        # attributes after the session closes.
        shift_rows = [
            SimpleNamespace(key=s.key, name=s.name) for s in shift_rows_orm
        ]
        slots = {
            key: [
                SimpleNamespace(
                    id=r.id, key=r.key, label=r.label,
                    from_time=r.from_time, to_time=r.to_time,
                )
                for r in rows
            ]
            for key, rows in slot_rows_by_key.items()
        }
    return templates.TemplateResponse(
        request,
        "shift_presets.html",
        {
            "company": company,
            "shifts": slots,
            "shift_rows": shift_rows,
            "shift_usage": usage,
        },
    )


@app.post("/production/shifts/add")
async def shift_add(request: Request):
    form = await request.form()
    company = (form.get("company") or "").strip()
    name = (form.get("name") or "").strip()
    if not (company and name):
        raise HTTPException(400, "company and name required")
    with _session() as session:
        existing = _shifts_load(session, company)
        taken = {s.key for s in existing}
        key = _shift_make_key(name, taken)
        max_so = max((s.sort_order for s in existing), default=0)
        session.add(Shift(
            company_name=company, key=key, name=name, sort_order=max_so + 10,
        ))
        session.commit()
    return RedirectResponse(
        f"/production/shift-presets?company={company}#sec-{key}", status_code=303
    )


@app.post("/production/shifts/save-all")
async def shifts_save_all(request: Request):
    """Persist every shift's display name in one transaction. Form fields
    `shift_key` and `shift_name` are parallel lists in document order. Empty
    names are silently skipped (treat as no-op)."""
    form = await request.form()
    company = (form.get("company") or "").strip()
    if not company:
        raise HTTPException(400, "company required")
    keys = form.getlist("shift_key")
    names = form.getlist("shift_name")
    if len(keys) != len(names):
        raise HTTPException(400, "form arrays mismatched")
    with _session() as session:
        rows_by_key = {
            r.key: r for r in session.scalars(
                select(Shift).where(Shift.company_name == company)
            )
        }
        for k, n in zip(keys, names):
            k = (k or "").strip()
            n = (n or "").strip()
            if not (k and n):
                continue
            row = rows_by_key.get(k)
            if not row:
                continue
            row.name = n
        session.commit()
    return RedirectResponse(
        f"/production/shift-presets?company={company}", status_code=303
    )


@app.post("/production/shifts/{shift}/rename")
async def shift_rename(shift: str, request: Request):
    form = await request.form()
    company = (form.get("company") or "").strip()
    name = (form.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "name required")
    with _session() as session:
        row = session.scalar(
            select(Shift).where(Shift.company_name == company, Shift.key == shift)
        )
        if not row:
            raise HTTPException(404, "shift not found")
        row.name = name
        session.commit()
    return RedirectResponse(
        f"/production/shift-presets?company={company}#sec-{shift}", status_code=303
    )


@app.post("/production/shifts/{shift}/move")
async def shift_move(shift: str, request: Request):
    form = await request.form()
    company = (form.get("company") or "").strip()
    direction = (form.get("dir") or "").strip()
    if direction not in {"up", "down"}:
        raise HTTPException(400, "dir must be up or down")
    with _session() as session:
        peers = _shifts_load(session, company)
        try:
            idx = next(i for i, r in enumerate(peers) if r.key == shift)
        except StopIteration:
            raise HTTPException(404, "shift not found")
        target_idx = idx - 1 if direction == "up" else idx + 1
        if 0 <= target_idx < len(peers):
            row = peers[idx]
            other = peers[target_idx]
            row.sort_order, other.sort_order = other.sort_order, row.sort_order
            session.commit()
    return RedirectResponse(
        f"/production/shift-presets?company={company}#sec-{shift}", status_code=303
    )


@app.post("/production/shifts/{shift}/delete")
async def shift_delete(shift: str, request: Request):
    form = await request.form()
    company = (form.get("company") or "").strip()
    with _session() as session:
        if _shift_key_in_use(session, company, shift) > 0:
            raise HTTPException(
                400,
                f"shift {shift!r} is referenced by existing DPR reports — "
                "rename instead, or delete those reports first",
            )
        row = session.scalar(
            select(Shift).where(Shift.company_name == company, Shift.key == shift)
        )
        if row:
            # Cascade: drop the shift's hour-slot presets too. Existing DPR
            # reports keep their stored hour_slots_json, so no data is lost.
            for s in session.scalars(
                select(ShiftPresetSlot).where(
                    ShiftPresetSlot.company_name == company,
                    ShiftPresetSlot.shift == shift,
                )
            ):
                session.delete(s)
            session.delete(row)
            session.commit()
    return RedirectResponse(
        f"/production/shift-presets?company={company}", status_code=303
    )


@app.post("/production/shift-presets/{shift}/add")
async def sp_add(shift: str, request: Request):
    form = await request.form()
    company = (form.get("company") or "").strip()
    from_t = _normalize_time(form.get("from_time"))
    to_t = _normalize_time(form.get("to_time"))
    if not company:
        raise HTTPException(400, "company required")
    label = _shift_label_from_times(from_t, to_t)
    with _session() as session:
        if shift not in _shift_keys(session, company):
            raise HTTPException(404, "unknown shift")
        existing = _shift_load(session, company, shift)
        taken = {s.key for s in existing}
        key = _shift_slug(label, taken)
        max_so = max((s.sort_order for s in existing), default=0)
        session.add(ShiftPresetSlot(
            company_name=company, shift=shift, key=key, label=label,
            from_time=from_t, to_time=to_t,
            sort_order=max_so + 10,
        ))
        session.commit()
    return RedirectResponse(
        f"/production/shift-presets?company={company}#sec-{shift}", status_code=303
    )


@app.post("/production/shift-presets/{shift}/save-all")
async def sp_save_all(shift: str, request: Request):
    """Persist every row's From/To in one transaction. Form fields `slot_id`,
    `from_time`, `to_time` are parallel lists in document order. Rows whose
    From and To are both blank are silently skipped (treat as no-op)."""
    form = await request.form()
    company = (form.get("company") or "").strip()
    if not company:
        raise HTTPException(400, "company required")
    slot_ids = form.getlist("slot_id")
    from_raws = form.getlist("from_time")
    to_raws = form.getlist("to_time")
    if not (len(slot_ids) == len(from_raws) == len(to_raws)):
        raise HTTPException(400, "form arrays mismatched")
    with _session() as session:
        for sid_raw, ft_raw, tt_raw in zip(slot_ids, from_raws, to_raws):
            try:
                sid = int(sid_raw)
            except (TypeError, ValueError):
                continue
            row = session.get(ShiftPresetSlot, sid)
            if not row or row.company_name != company or row.shift != shift:
                continue
            from_t = _normalize_time(ft_raw)
            to_t = _normalize_time(tt_raw)
            if not (from_t or to_t):
                continue
            row.from_time = from_t
            row.to_time = to_t
            row.label = _shift_label_from_times(from_t, to_t)
        session.commit()
    return RedirectResponse(
        f"/production/shift-presets?company={company}#sec-{shift}", status_code=303
    )


@app.post("/production/shift-presets/{shift}/{slot_id}/edit")
async def sp_edit(shift: str, slot_id: int, request: Request):
    form = await request.form()
    company = (form.get("company") or "").strip()
    from_t = _normalize_time(form.get("from_time"))
    to_t = _normalize_time(form.get("to_time"))
    label = _shift_label_from_times(from_t, to_t)
    with _session() as session:
        row = session.get(ShiftPresetSlot, slot_id)
        if not row or row.company_name != company or row.shift != shift:
            raise HTTPException(404, "slot not found")
        row.label = label
        row.from_time = from_t
        row.to_time = to_t
        session.commit()
    return RedirectResponse(
        f"/production/shift-presets?company={company}#sec-{shift}", status_code=303
    )


@app.post("/production/shift-presets/{shift}/{slot_id}/move")
async def sp_move(shift: str, slot_id: int, request: Request):
    form = await request.form()
    company = (form.get("company") or "").strip()
    direction = (form.get("dir") or "").strip()
    if direction not in {"up", "down"}:
        raise HTTPException(400, "dir must be up or down")
    with _session() as session:
        row = session.get(ShiftPresetSlot, slot_id)
        if not row or row.company_name != company or row.shift != shift:
            raise HTTPException(404, "slot not found")
        peers = list(session.scalars(
            select(ShiftPresetSlot)
            .where(
                ShiftPresetSlot.company_name == company,
                ShiftPresetSlot.shift == shift,
            )
            .order_by(ShiftPresetSlot.sort_order, ShiftPresetSlot.id)
        ))
        try:
            idx = next(i for i, r in enumerate(peers) if r.id == row.id)
        except StopIteration:
            raise HTTPException(500, "peer index not found")
        target_idx = idx - 1 if direction == "up" else idx + 1
        if 0 <= target_idx < len(peers):
            other = peers[target_idx]
            row.sort_order, other.sort_order = other.sort_order, row.sort_order
            session.commit()
    return RedirectResponse(
        f"/production/shift-presets?company={company}#sec-{shift}", status_code=303
    )


@app.post("/production/shift-presets/{shift}/{slot_id}/delete")
async def sp_delete(shift: str, slot_id: int, request: Request):
    form = await request.form()
    company = (form.get("company") or "").strip()
    with _session() as session:
        row = session.get(ShiftPresetSlot, slot_id)
        if row and row.company_name == company and row.shift == shift:
            session.delete(row)
            session.commit()
    return RedirectResponse(
        f"/production/shift-presets?company={company}#sec-{shift}", status_code=303
    )


@app.post("/production/shift-presets/{shift}/reset")
async def sp_reset(shift: str, request: Request):
    """Wipe all slots for (company, shift) and re-seed from the static
    SHIFT_PRESETS defaults. Useful when the user wants a clean reset."""
    form = await request.form()
    company = (form.get("company") or "").strip()
    with _session() as session:
        if shift not in _shift_keys(session, company):
            raise HTTPException(404, "unknown shift")
        for r in session.scalars(
            select(ShiftPresetSlot).where(
                ShiftPresetSlot.company_name == company, ShiftPresetSlot.shift == shift,
            )
        ):
            session.delete(r)
        session.commit()
        _shift_load(session, company, shift)
    return RedirectResponse(
        f"/production/shift-presets?company={company}#sec-{shift}", status_code=303
    )


@app.get("/production/process-catalog", response_class=HTMLResponse)
def pp_catalog(request: Request, company: str | None = None, edit: int | None = None, synced: int | None = None):
    with _session() as session:
        company = _pp_resolve_company(session, company)
        edit_entry = None
        if edit is not None:
            cand = session.get(ProcessCatalogEntry, edit)
            if cand and cand.company_name == company:
                edit_entry = cand
        stages = _ps_load(session, company)
        # Usage counts for safe-delete on stages.
        stage_usage: dict[str, int] = {s.key: 0 for s in stages}
        for e in session.scalars(
            select(ProcessCatalogEntry).where(ProcessCatalogEntry.company_name == company)
        ):
            stage_usage[e.stage] = stage_usage.get(e.stage, 0) + 1
        for r in session.scalars(
            select(ProductionProcess).where(ProductionProcess.company_name == company)
        ):
            stage_usage[r.stage] = stage_usage.get(r.stage, 0) + 1
        entries = list(session.scalars(
            select(ProcessCatalogEntry)
            .where(ProcessCatalogEntry.company_name == company)
            .order_by(ProcessCatalogEntry.section, ProcessCatalogEntry.group_label, ProcessCatalogEntry.label)
        ))
        sections: dict[str, list[ProcessCatalogEntry]] = {s: [] for s, _ in dpr.SECTION_OPTIONS}
        for e in entries:
            sections.setdefault(e.section, []).append(e)
    return templates.TemplateResponse(
        request,
        "process_catalog.html",
        {
            "company": company,
            "sections": sections,
            "section_options": dpr.SECTION_OPTIONS,
            "stages": stages,
            "stage_options": _ps_options(stages),
            "stage_labels": _ps_labels(stages),
            "stage_usage": stage_usage,
            "edit_entry": edit_entry,
            "synced": synced,
        },
    )


@app.post("/production/process-catalog/stage/add")
async def pp_stage_add(request: Request):
    form = await request.form()
    company = (form.get("company") or "").strip()
    label = (form.get("label") or "").strip()
    if not (company and label):
        raise HTTPException(400, "company and label required")
    key = _stage_slugify(label)
    with _session() as session:
        existing = session.scalar(
            select(ProcessStage).where(
                ProcessStage.company_name == company, ProcessStage.key == key,
            )
        )
        if existing:
            # Reactivate if previously deactivated; otherwise no-op.
            if not existing.active:
                existing.active = True
                session.commit()
        else:
            max_so = session.scalar(
                select(ProcessStage.sort_order)
                .where(ProcessStage.company_name == company)
                .order_by(ProcessStage.sort_order.desc())
            ) or 0
            session.add(ProcessStage(
                company_name=company, key=key, label=label,
                sort_order=max_so + 10, active=True,
            ))
            session.commit()
    return RedirectResponse(
        f"/production/process-catalog?company={company}#sec-stages", status_code=303
    )


@app.post("/production/process-catalog/stage/{stage_id}/edit")
async def pp_stage_edit(stage_id: int, request: Request):
    form = await request.form()
    company = (form.get("company") or "").strip()
    label = (form.get("label") or "").strip()
    if not label:
        raise HTTPException(400, "label required")
    with _session() as session:
        row = session.get(ProcessStage, stage_id)
        if not row or row.company_name != company:
            raise HTTPException(404, "stage not found")
        row.label = label
        session.commit()
    return RedirectResponse(
        f"/production/process-catalog?company={company}#sec-stages", status_code=303
    )


@app.post("/production/process-catalog/stage/{stage_id}/delete")
async def pp_stage_delete(stage_id: int, request: Request):
    """Soft delete (deactivate) — keeps existing references resolvable to a
    label even after the stage is hidden from new pickers. Hard delete is
    intentionally not offered to avoid breaking historical reports."""
    form = await request.form()
    company = (form.get("company") or "").strip()
    with _session() as session:
        row = session.get(ProcessStage, stage_id)
        if not row or row.company_name != company:
            raise HTTPException(404, "stage not found")
        row.active = False
        session.commit()
    return RedirectResponse(
        f"/production/process-catalog?company={company}#sec-stages", status_code=303
    )


@app.post("/production/process-catalog/stage/{stage_id}/reactivate")
async def pp_stage_reactivate(stage_id: int, request: Request):
    form = await request.form()
    company = (form.get("company") or "").strip()
    with _session() as session:
        row = session.get(ProcessStage, stage_id)
        if not row or row.company_name != company:
            raise HTTPException(404, "stage not found")
        row.active = True
        session.commit()
    return RedirectResponse(
        f"/production/process-catalog?company={company}#sec-stages", status_code=303
    )


@app.post("/production/process-catalog/seed")
async def pp_catalog_seed(request: Request):
    """Bulk-load the Line 4 paper-form defaults as catalog entries.
    Skips entries that already exist (matched by section + label + group_label)."""
    form = await request.form()
    company = (form.get("company") or "").strip()
    if not company:
        raise HTTPException(400, "company required")
    with _session() as session:
        existing = {
            (e.section, e.label, e.group_label or "")
            for e in session.scalars(
                select(ProcessCatalogEntry).where(ProcessCatalogEntry.company_name == company)
            )
        }
        for section, seeds in dpr.DEFAULT_SEEDS.items():
            for label, stage, group_label in seeds:
                key = (section, label, group_label or "")
                if key in existing:
                    continue
                session.add(ProcessCatalogEntry(
                    company_name=company,
                    section=section,
                    label=label,
                    stage=stage,
                    group_label=group_label,
                ))
                existing.add(key)
        session.commit()
    return RedirectResponse(
        f"/production/process-catalog?company={company}", status_code=303
    )


@app.post("/production/process-catalog/add")
async def pp_catalog_add(request: Request):
    form = await request.form()
    company = (form.get("company") or "").strip()
    section = (form.get("section") or "").strip()
    label = (form.get("label") or "").strip()
    stage = (form.get("stage") or "").strip()
    group_label = (form.get("group_label") or "").strip() or None
    role = (form.get("role") or "").strip().lower() or None
    validate_count = form.get("validate_count") in ("1", "on", "true")
    if role not in {None, "input", "output"}:
        raise HTTPException(400, "role must be input or output")
    if not (company and section and label and stage):
        raise HTTPException(400, "company, section, label and stage required")
    if section not in dict(dpr.SECTION_OPTIONS):
        raise HTTPException(400, "invalid section")
    if section == "production" and role is None:
        role = "output"
    if section != "production":
        role = None
    with _session() as session:
        if stage not in _ps_active_keys(session, company):
            raise HTTPException(400, "invalid stage")
        session.add(ProcessCatalogEntry(
            company_name=company,
            section=section,
            label=label,
            stage=stage,
            group_label=group_label,
            role=role,
            validate_count=validate_count if section == "production" else True,
        ))
        session.commit()
    return RedirectResponse(
        f"/production/process-catalog?company={company}#sec-{section}", status_code=303
    )


@app.post("/production/process-catalog/sync-flags")
async def pp_catalog_sync_flags(request: Request):
    """Push role + validate_count from catalog entries to matching line
    rows (matched by company + section + label + stage). Existing line rows
    keep their sort_order, group_label and active flag — only the two flag
    columns are overwritten. Doesn't add new rows or remove orphans."""
    form = await request.form()
    company = (form.get("company") or "").strip()
    if not company:
        raise HTTPException(400, "company required")
    updated = 0
    with _session() as session:
        catalog = list(session.scalars(
            select(ProcessCatalogEntry).where(ProcessCatalogEntry.company_name == company)
        ))
        index = {(c.section, c.label, c.stage): c for c in catalog}
        for row in session.scalars(
            select(ProductionProcess).where(ProductionProcess.company_name == company)
        ):
            cat = index.get((row.section, row.label, row.stage))
            if not cat:
                continue
            changed = False
            if row.section == "production":
                if row.role != cat.role:
                    row.role = cat.role
                    changed = True
                if bool(row.validate_count) != bool(cat.validate_count):
                    row.validate_count = bool(cat.validate_count)
                    changed = True
            if changed:
                updated += 1
        session.commit()
    return RedirectResponse(
        f"/production/process-catalog?company={company}&synced={updated}",
        status_code=303,
    )


@app.post("/production/process-catalog/{entry_id}/edit")
async def pp_catalog_edit(entry_id: int, request: Request):
    form = await request.form()
    company = (form.get("company") or "").strip()
    label = (form.get("label") or "").strip()
    stage = (form.get("stage") or "").strip()
    group_label = (form.get("group_label") or "").strip() or None
    role = (form.get("role") or "").strip().lower() or None
    if role not in {None, "input", "output"}:
        raise HTTPException(400, "role must be input or output")
    if not (label and stage):
        raise HTTPException(400, "label and stage required")
    with _session() as session:
        if stage not in _ps_active_keys(session, company):
            raise HTTPException(400, "invalid stage")
        row = session.get(ProcessCatalogEntry, entry_id)
        if not row or row.company_name != company:
            raise HTTPException(404, "catalog entry not found")
        row.label = label
        row.stage = stage
        row.group_label = group_label if row.section == "rejection" else None
        if row.section == "production":
            row.role = role or "output"
            row.validate_count = form.get("validate_count") in ("1", "on", "true")
        else:
            row.role = None
        section = row.section
        session.commit()
    return RedirectResponse(
        f"/production/process-catalog?company={company}#sec-{section}",
        status_code=303,
    )


@app.post("/production/process-catalog/{entry_id}/delete")
async def pp_catalog_delete(entry_id: int, request: Request):
    form = await request.form()
    company = (form.get("company") or "").strip()
    with _session() as session:
        row = session.get(ProcessCatalogEntry, entry_id)
        section = row.section if row else ""
        if row:
            session.delete(row)
            session.commit()
    return RedirectResponse(
        f"/production/process-catalog?company={company}#sec-{section}", status_code=303
    )


def _pp_cascade_from_production(
    session: Session, company: str, line: str, stage_keys: list[str]
) -> int:
    """Auto-import every Rejection + Rework catalog entry that matches one of
    `stage_keys` into this line, skipping labels already present in their
    section. Returns the number of rows added.

    Single source of truth for the production-stage cascade — called by
    apply-catalog (single stage just added), seed (all production stages),
    and refresh-from-catalog (all currently active production stages)."""
    if not stage_keys:
        return 0
    added = 0
    for cascade_section in ("rejection", "rework"):
        # Dedup by (stage, label, group) — same label can legitimately appear
        # under multiple groups (e.g. "Chip off / Breakage" exists under both
        # Single Edger and Auto Corner). Label-only dedup would skip the
        # second occurrence and lose those entries.
        existing_keys = {
            (r.stage, r.label, r.group_label or "") for r in session.scalars(
                select(ProductionProcess).where(
                    ProductionProcess.company_name == company,
                    ProductionProcess.line == line,
                    ProductionProcess.section == cascade_section,
                )
            )
        }
        cas_max_so = session.scalar(
            select(ProductionProcess.sort_order)
            .where(
                ProductionProcess.company_name == company,
                ProductionProcess.line == line,
                ProductionProcess.section == cascade_section,
            )
            .order_by(ProductionProcess.sort_order.desc())
        ) or 0
        for cat in session.scalars(
            select(ProcessCatalogEntry)
            .where(
                ProcessCatalogEntry.company_name == company,
                ProcessCatalogEntry.section == cascade_section,
                ProcessCatalogEntry.stage.in_(stage_keys),
            )
            .order_by(ProcessCatalogEntry.stage, ProcessCatalogEntry.group_label, ProcessCatalogEntry.label)
        ):
            key = (cat.stage, cat.label, cat.group_label or "")
            if key in existing_keys:
                continue
            cas_max_so += 10
            session.add(ProductionProcess(
                company_name=company,
                line=line,
                section=cascade_section,
                stage=cat.stage,
                label=cat.label,
                group_label=cat.group_label,
                sort_order=cas_max_so,
                active=True,
            ))
            existing_keys.add(key)
            added += 1
    return added


@app.post("/production/processes/{line}/apply-catalog")
async def pp_apply_catalog(line: str, request: Request):
    """Copy a catalog entry's (label, stage, group) into a new ProductionProcess
    row on this line. Catalog and line row are independent after this."""
    if line not in dpr.LINE_OPTIONS:
        raise HTTPException(404, "unknown line")
    form = await request.form()
    company = (form.get("company") or "").strip()
    entry_id_raw = (form.get("entry_id") or "").strip()
    if not entry_id_raw.isdigit():
        raise HTTPException(400, "entry_id required")
    with _session() as session:
        entry = session.get(ProcessCatalogEntry, int(entry_id_raw))
        if not entry or entry.company_name != company:
            raise HTTPException(404, "catalog entry not found")
        max_so = session.scalar(
            select(ProductionProcess.sort_order)
            .where(
                ProductionProcess.company_name == company,
                ProductionProcess.line == line,
                ProductionProcess.section == entry.section,
            )
            .order_by(ProductionProcess.sort_order.desc())
        ) or 0
        section = entry.section
        session.add(ProductionProcess(
            company_name=company,
            line=line,
            section=entry.section,
            stage=entry.stage,
            label=entry.label,
            group_label=entry.group_label,
            role=entry.role,
            validate_count=bool(getattr(entry, "validate_count", True)),
            sort_order=max_so + 10,
            active=True,
        ))

        if entry.section == "production":
            session.flush()  # so the new row is visible to the cascade query
            _pp_cascade_from_production(session, company, line, [entry.stage])

        session.commit()
    return RedirectResponse(
        f"/production/processes/{line}?company={company}#sec-{section}", status_code=303
    )


@app.post("/production/processes/{line}/clear")
async def pp_clear(line: str, request: Request):
    """Hard-delete every process row on this line (all sections, active +
    inactive). Existing DPR cells that referenced these row IDs become
    orphans — they remain in the DB but won't render in any matrix."""
    if line not in dpr.LINE_OPTIONS:
        raise HTTPException(404, "unknown line")
    form = await request.form()
    company = (form.get("company") or "").strip()
    with _session() as session:
        for r in session.scalars(
            select(ProductionProcess).where(
                ProductionProcess.company_name == company,
                ProductionProcess.line == line,
            )
        ):
            session.delete(r)
        session.commit()
    return RedirectResponse(
        f"/production/processes/{line}?company={company}", status_code=303
    )


@app.post("/production/processes/{line}/seed")
async def pp_seed(line: str, request: Request):
    if line not in dpr.LINE_OPTIONS:
        raise HTTPException(404, "unknown line")
    form = await request.form()
    company = (form.get("company") or "").strip()
    section = (form.get("section") or "").strip()
    if section not in dict(dpr.SECTION_OPTIONS):
        raise HTTPException(400, "invalid section")
    seeds = dpr.DEFAULT_SEEDS.get(section, [])
    with _session() as session:
        existing = session.scalars(
            select(ProductionProcess).where(
                ProductionProcess.company_name == company,
                ProductionProcess.line == line,
                ProductionProcess.section == section,
            )
        ).first()
        if existing:
            raise HTTPException(400, "section already has rows; clear them first")
        seeded_stages: list[str] = []
        for idx, (label, stage, group_label) in enumerate(seeds):
            session.add(ProductionProcess(
                company_name=company,
                line=line,
                section=section,
                stage=stage,
                label=label,
                group_label=group_label,
                sort_order=(idx + 1) * 10,
                active=True,
            ))
            if stage not in seeded_stages:
                seeded_stages.append(stage)
        # Mirror apply-catalog cascade: production-section seed pulls in any
        # matching rejection / rework catalog entries automatically.
        if section == "production" and seeded_stages:
            session.flush()
            _pp_cascade_from_production(session, company, line, seeded_stages)
        session.commit()
    return RedirectResponse(
        f"/production/processes/{line}?company={company}", status_code=303
    )


@app.post("/production/processes/{line}/refresh-cascade")
async def pp_refresh_cascade(line: str, request: Request):
    """Re-run the production-stage cascade against the current state of the
    catalog. Pulls in catalog entries (rejection/rework) that match any stage
    used by an active production row on this line and aren't already present.
    Doesn't remove or reactivate anything — purely additive."""
    if line not in dpr.LINE_OPTIONS:
        raise HTTPException(404, "unknown line")
    form = await request.form()
    company = (form.get("company") or "").strip()
    with _session() as session:
        active_prod_stages = sorted({
            r.stage for r in session.scalars(
                select(ProductionProcess).where(
                    ProductionProcess.company_name == company,
                    ProductionProcess.line == line,
                    ProductionProcess.section == "production",
                    ProductionProcess.active.is_(True),
                )
            )
        })
        _pp_cascade_from_production(session, company, line, active_prod_stages)
        session.commit()
    return RedirectResponse(
        f"/production/processes/{line}?company={company}", status_code=303
    )


@app.post("/production/processes/{line}/add")
async def pp_add(line: str, request: Request):
    if line not in dpr.LINE_OPTIONS:
        raise HTTPException(404, "unknown line")
    form = await request.form()
    company = (form.get("company") or "").strip()
    section = (form.get("section") or "").strip()
    label = (form.get("label") or "").strip()
    stage = (form.get("stage") or "").strip()
    group_label = (form.get("group_label") or "").strip() or None
    role = (form.get("role") or "").strip().lower() or None
    validate_count = form.get("validate_count") in ("1", "on", "true")
    if role not in {None, "input", "output"}:
        raise HTTPException(400, "role must be input or output")
    if not (company and section and label and stage):
        raise HTTPException(400, "company, section, label and stage required")
    if section not in dict(dpr.SECTION_OPTIONS):
        raise HTTPException(400, "invalid section")
    if section == "production" and role is None:
        # Default new production rows to output if the form omitted role.
        role = "output"
    if section != "production":
        role = None
    with _session() as session:
        if stage not in _ps_active_keys(session, company):
            raise HTTPException(400, "invalid stage")
        max_so = session.scalar(
            select(ProductionProcess.sort_order)
            .where(
                ProductionProcess.company_name == company,
                ProductionProcess.line == line,
                ProductionProcess.section == section,
            )
            .order_by(ProductionProcess.sort_order.desc())
        ) or 0
        session.add(ProductionProcess(
            company_name=company,
            line=line,
            section=section,
            stage=stage,
            label=label,
            group_label=group_label,
            role=role,
            validate_count=validate_count if section == "production" else True,
            sort_order=max_so + 10,
            active=True,
        ))
        session.commit()
    return RedirectResponse(
        f"/production/processes/{line}?company={company}#sec-{section}", status_code=303
    )


@app.post("/production/processes/{line}/{process_id}/edit")
async def pp_edit(line: str, process_id: int, request: Request):
    form = await request.form()
    company = (form.get("company") or "").strip()
    label = (form.get("label") or "").strip()
    stage = (form.get("stage") or "").strip()
    group_label = (form.get("group_label") or "").strip() or None
    role = (form.get("role") or "").strip().lower() or None
    if role not in {None, "input", "output"}:
        raise HTTPException(400, "role must be input or output")
    if not (label and stage):
        raise HTTPException(400, "label and stage required")
    with _session() as session:
        if stage not in _ps_active_keys(session, company):
            raise HTTPException(400, "invalid stage")
        row = session.get(ProductionProcess, process_id)
        if not row or row.line != line:
            raise HTTPException(404, "row not found")
        row.label = label
        row.stage = stage
        row.group_label = group_label
        row.role = role if row.section == "production" else None
        if row.section == "production":
            row.validate_count = form.get("validate_count") in ("1", "on", "true")
        section = row.section
        session.commit()
    return RedirectResponse(
        f"/production/processes/{line}?company={company}#sec-{section}",
        status_code=303,
    )


@app.post("/production/processes/{line}/{process_id}/move")
async def pp_move(line: str, process_id: int, request: Request):
    form = await request.form()
    company = (form.get("company") or "").strip()
    direction = (form.get("dir") or "").strip()
    if direction not in {"up", "down"}:
        raise HTTPException(400, "dir must be up or down")
    with _session() as session:
        row = session.get(ProductionProcess, process_id)
        if not row or row.line != line:
            raise HTTPException(404, "row not found")
        # Find the neighbour in the same line+section ordered by sort_order.
        peers = list(session.scalars(
            select(ProductionProcess)
            .where(
                ProductionProcess.company_name == row.company_name,
                ProductionProcess.line == row.line,
                ProductionProcess.section == row.section,
            )
            .order_by(ProductionProcess.sort_order, ProductionProcess.id)
        ))
        try:
            idx = next(i for i, r in enumerate(peers) if r.id == row.id)
        except StopIteration:
            raise HTTPException(500, "peer index not found")
        target_idx = idx - 1 if direction == "up" else idx + 1
        section = row.section
        if 0 <= target_idx < len(peers):
            other = peers[target_idx]
            row.sort_order, other.sort_order = other.sort_order, row.sort_order
            session.commit()
    return RedirectResponse(
        f"/production/processes/{line}?company={company}#sec-{section}",
        status_code=303,
    )


@app.post("/production/processes/{line}/{process_id}/toggle")
async def pp_toggle(line: str, process_id: int, request: Request):
    form = await request.form()
    company = (form.get("company") or "").strip()
    show_inactive = (form.get("show_inactive") or "").strip()
    with _session() as session:
        row = session.get(ProductionProcess, process_id)
        if not row or row.line != line:
            raise HTTPException(404, "row not found")
        was_active = row.active
        row.active = not row.active
        section = row.section
        stage_key = row.stage

        # Cascade: deactivating a Production row "removes" the stage from this
        # line's hourly-output catalog. If no other active production row on
        # this line still carries that stage, the rejection / rework rows
        # tagged with the stage become orphans — deactivate them too. Inverse
        # cascade (re-activating production) is intentionally NOT done; users
        # can manually reactivate the rejection/rework rows from the inactive
        # list if they want them back.
        if was_active and section == "production":
            other_active = session.scalar(
                select(ProductionProcess.id).where(
                    ProductionProcess.company_name == row.company_name,
                    ProductionProcess.line == row.line,
                    ProductionProcess.section == "production",
                    ProductionProcess.stage == stage_key,
                    ProductionProcess.active.is_(True),
                    ProductionProcess.id != row.id,
                )
            )
            if not other_active:
                for cas in session.scalars(
                    select(ProductionProcess).where(
                        ProductionProcess.company_name == row.company_name,
                        ProductionProcess.line == row.line,
                        ProductionProcess.section.in_(["rejection", "rework"]),
                        ProductionProcess.stage == stage_key,
                        ProductionProcess.active.is_(True),
                    )
                ):
                    cas.active = False

        session.commit()
    qs = f"company={company}"
    if show_inactive:
        qs += "&show_inactive=1"
    return RedirectResponse(
        f"/production/processes/{line}?{qs}#sec-{section}", status_code=303
    )


# ---------------------------------------------------------------------------
# Company aliases (short codes used in the consumable-import spreadsheet)
# ---------------------------------------------------------------------------

class CompanyAliasIn(BaseModel):
    code: str
    company_name: str


class CompanyAliasesPayload(BaseModel):
    aliases: list[CompanyAliasIn]


@app.get("/api/company-aliases")
def api_list_company_aliases():
    with _session() as session:
        rows = list(session.scalars(select(CompanyAlias).order_by(CompanyAlias.code)))
        return {
            "aliases": [{"code": r.code, "company_name": r.company_name} for r in rows],
            "companies": _list_companies(session),
        }


@app.put("/api/company-aliases")
def api_put_company_aliases(payload: CompanyAliasesPayload):
    with _session() as session:
        # Replace-all semantics: simpler to reason about than partial upsert.
        for r in list(session.scalars(select(CompanyAlias))):
            session.delete(r)
        session.flush()
        seen: set[str] = set()
        for a in payload.aliases:
            code = (a.code or "").strip().upper()
            name = (a.company_name or "").strip()
            if not code or not name or code in seen:
                continue
            seen.add(code)
            session.add(CompanyAlias(code=code, company_name=name))
        session.commit()
        return api_list_company_aliases()


# ---------------------------------------------------------------------------
# Consumable import (Excel sheet/table → draft ProductionEntries)
# ---------------------------------------------------------------------------

def _parse_dmy(raw) -> str | None:
    """Parse a date header into ISO YYYY-MM-DD. Accepts datetime objects (Excel
    auto-converted dates) and strings like '1-4-2026', '01/04/26', '1/4/2026'."""
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date().isoformat()
    if isinstance(raw, date):
        return raw.isoformat()
    s = str(raw).strip()
    if not s:
        return None
    for sep in ("-", "/", "."):
        if sep in s:
            parts = s.split(sep)
            if len(parts) == 3:
                try:
                    d, m, y = (int(p) for p in parts)
                except ValueError:
                    return None
                if y < 100:
                    y += 2000
                try:
                    return date(y, m, d).isoformat()
                except ValueError:
                    return None
    return None


def _build_item_voucher_type_map(session: Session, company: str) -> dict[str, str]:
    """For each consume-stock-item under `company`, find the policy/voucher_type
    that owns its group (walking ancestrally). Returns {item_name → voucher_type}.

    If multiple policies claim an item, the alphabetically-first voucher_type wins
    (deterministic) — operators should keep policies non-overlapping anyway.
    """
    from ..policy import _descendant_group_names
    policies = list_policies(session, company)
    item_to_vt: dict[str, str] = {}
    items = list(
        session.scalars(
            select(StockItem).where(StockItem.company_name == company)
        )
    )
    for policy in sorted(policies, key=lambda p: p.voucher_type):
        consume_groups = [g.stock_group for g in policy.groups if g.role == "consume"]
        if not consume_groups:
            continue
        expanded = _descendant_group_names(session, company, consume_groups)
        for it in items:
            if it.parent in expanded and it.name not in item_to_vt:
                item_to_vt[it.name] = policy.voucher_type
    return item_to_vt


def _parse_consumable_workbook(
    raw: bytes, sheet_name: str | None, table_name: str | None
) -> tuple[list[str], list[list]]:
    """Read the workbook and return (header_row, data_rows). data_rows are
    raw cell values (no normalisation), aligned with the header.
    """
    from io import BytesIO
    from openpyxl import load_workbook

    try:
        wb = load_workbook(BytesIO(raw), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"could not read Excel: {e}")

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise HTTPException(
                400,
                f"sheet '{sheet_name}' not found. Sheets in file: {', '.join(wb.sheetnames)}",
            )
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    if table_name:
        # openpyxl exposes named tables (ListObjects) on each worksheet.
        tables = getattr(ws, "tables", None) or {}
        if table_name not in tables:
            available = ", ".join(tables.keys()) or "(none)"
            raise HTTPException(
                400,
                f"table '{table_name}' not found in sheet '{ws.title}'. Tables: {available}",
            )
        ref = tables[table_name].ref if hasattr(tables[table_name], "ref") else tables[table_name]
        cells = ws[ref]
    else:
        cells = list(ws.iter_rows())

    rows = [[c.value for c in row] for row in cells]
    if len(rows) < 2:
        raise HTTPException(400, "no data rows found in selected range")
    header = [str(v).strip() if v is not None else "" for v in rows[0]]
    return header, rows[1:]


def _resolve_consumable_columns(header: list[str]) -> tuple[int, int, list[tuple[int, str]]]:
    """Given the header row, return (item_col, company_col, [(idx, iso_date), ...])."""
    item_col = company_col = -1
    for i, h in enumerate(header):
        norm = h.lower().replace("_", " ").strip()
        if norm in ("item name", "item", "stock item", "stock item name"):
            item_col = i
        elif norm in ("company", "co", "company code"):
            company_col = i
    if item_col < 0:
        raise HTTPException(400, "missing required column 'ITEM NAME'")
    if company_col < 0:
        raise HTTPException(400, "missing required column 'Company'")
    date_cols: list[tuple[int, str]] = []
    for i, h in enumerate(header):
        if i in (item_col, company_col):
            continue
        iso = _parse_dmy(h)
        if iso:
            date_cols.append((i, iso))
    if not date_cols:
        raise HTTPException(
            400,
            "no date columns recognised — header dates must be in dd-mm-yyyy or dd/mm/yyyy form",
        )
    return item_col, company_col, date_cols


def _build_consumable_preview(
    session: Session,
    header: list[str],
    data_rows: list[list],
    from_date: str | None = None,
    to_date: str | None = None,
) -> dict:
    item_col, company_col, date_cols = _resolve_consumable_columns(header)
    if from_date:
        date_cols = [(i, d) for (i, d) in date_cols if d >= from_date]
    if to_date:
        date_cols = [(i, d) for (i, d) in date_cols if d <= to_date]
    if (from_date or to_date) and not date_cols:
        raise HTTPException(400, "no date columns fall within the selected From/To range")

    alias_map = {
        a.code.upper(): a.company_name
        for a in session.scalars(select(CompanyAlias))
    }
    known_companies = set(_list_companies(session))

    # Cache per-company item→voucher_type maps and item index
    vt_cache: dict[str, dict[str, str]] = {}
    item_cache: dict[str, set[str]] = {}

    def _vt_for(company: str) -> dict[str, str]:
        if company not in vt_cache:
            vt_cache[company] = _build_item_voucher_type_map(session, company)
        return vt_cache[company]

    def _items_for(company: str) -> set[str]:
        if company not in item_cache:
            item_cache[company] = set(
                session.scalars(
                    select(StockItem.name).where(StockItem.company_name == company)
                )
            )
        return item_cache[company]

    matched: list[dict] = []  # one row per (company, date, item) with qty>0
    errors: list[dict] = []   # unmatched items / unknown companies / orphan policies

    for r_idx, row in enumerate(data_rows, start=2):  # row 1 = header
        if not row or all(v in (None, "") for v in row):
            continue
        item_name = (str(row[item_col]).strip() if row[item_col] is not None else "")
        code = (str(row[company_col]).strip().upper() if row[company_col] is not None else "")
        if not item_name and not code:
            continue
        if not item_name:
            errors.append({"row": r_idx, "kind": "missing_item", "code": code})
            continue
        if not code:
            errors.append({"row": r_idx, "kind": "missing_company", "item": item_name})
            continue
        company = alias_map.get(code)
        if not company:
            errors.append({"row": r_idx, "kind": "unknown_alias", "code": code, "item": item_name})
            continue
        if company not in known_companies:
            errors.append({
                "row": r_idx, "kind": "alias_targets_unknown_company",
                "code": code, "item": item_name, "company": company,
            })
            continue
        items = _items_for(company)
        if item_name not in items:
            errors.append({
                "row": r_idx, "kind": "unknown_item",
                "code": code, "company": company, "item": item_name,
            })
            continue
        vt_map = _vt_for(company)
        voucher_type = vt_map.get(item_name)
        if not voucher_type:
            errors.append({
                "row": r_idx, "kind": "no_policy_for_item",
                "code": code, "company": company, "item": item_name,
            })
            continue
        for col_idx, iso_date in date_cols:
            v = row[col_idx] if col_idx < len(row) else None
            if v in (None, ""):
                continue
            try:
                qty = float(v)
            except (TypeError, ValueError):
                errors.append({
                    "row": r_idx, "kind": "non_numeric_qty",
                    "company": company, "item": item_name,
                    "date": iso_date, "raw": str(v),
                })
                continue
            if qty == 0:
                continue
            matched.append({
                "row": r_idx,
                "code": code,
                "company": company,
                "item": item_name,
                "date": iso_date,
                "qty": qty,
                "voucher_type": voucher_type,
            })

    # Group matched into entries: (company, date, voucher_type) → [lines]
    bundles: dict[tuple[str, str, str], list[dict]] = {}
    for m in matched:
        key = (m["company"], m["date"], m["voucher_type"])
        bundles.setdefault(key, []).append(m)
    grouped = [
        {
            "company": k[0],
            "date": k[1],
            "voucher_type": k[2],
            "lines": v,
            "total_qty": sum(x["qty"] for x in v),
        }
        for k, v in sorted(bundles.items())
    ]

    return {
        "matched_count": len(matched),
        "error_count": len(errors),
        "entry_count": len(grouped),
        "entries": grouped,
        "errors": errors,
    }


@app.post("/api/consumable-import/inspect")
async def api_consumable_import_inspect(
    file: UploadFile = File(...),
    sheet_name: str = Form(""),
    table_name: str = Form(""),
):
    """Return the structure of an uploaded workbook so the UI can offer
    sheet/table/date dropdowns. If sheet_name/table_name are provided, also
    returns the recognised date columns for that selection."""
    from io import BytesIO
    from openpyxl import load_workbook

    raw = await file.read()
    if not raw:
        raise HTTPException(400, "uploaded file is empty")
    try:
        wb = load_workbook(BytesIO(raw), data_only=True, read_only=False)
    except Exception as e:
        raise HTTPException(400, f"could not read Excel: {e}")

    sheets = list(wb.sheetnames)
    tables_by_sheet: dict[str, list[str]] = {}
    for sn in sheets:
        ws = wb[sn]
        tables_by_sheet[sn] = list((getattr(ws, "tables", None) or {}).keys())

    date_columns: list[dict] = []
    resolved_sheet = sheet_name or (sheets[0] if sheets else "")
    if resolved_sheet:
        try:
            header, _rows = _parse_consumable_workbook(
                raw, resolved_sheet, table_name or None
            )
            for i, h in enumerate(header):
                iso = _parse_dmy(h)
                if iso:
                    date_columns.append({"label": str(h), "date": iso})
        except HTTPException:
            # Header may be unparseable until a table is chosen — that's fine,
            # frontend just won't show date options yet.
            pass

    return {
        "sheets": sheets,
        "tables_by_sheet": tables_by_sheet,
        "date_columns": date_columns,
        "resolved_sheet": resolved_sheet,
    }


@app.post("/api/consumable-import/preview")
async def api_consumable_import_preview(
    file: UploadFile = File(...),
    sheet_name: str = Form(""),
    table_name: str = Form(""),
    from_date: str = Form(""),
    to_date: str = Form(""),
):
    raw = await file.read()
    if not raw:
        raise HTTPException(400, "uploaded file is empty")
    with _session() as session:
        header, data_rows = _parse_consumable_workbook(raw, sheet_name or None, table_name or None)
        return _build_consumable_preview(
            session, header, data_rows, from_date or None, to_date or None
        )


@app.post("/api/consumable-import/commit")
async def api_consumable_import_commit(
    file: UploadFile = File(...),
    sheet_name: str = Form(""),
    table_name: str = Form(""),
    from_date: str = Form(""),
    to_date: str = Form(""),
):
    raw = await file.read()
    if not raw:
        raise HTTPException(400, "uploaded file is empty")
    with _session() as session:
        header, data_rows = _parse_consumable_workbook(raw, sheet_name or None, table_name or None)
        preview = _build_consumable_preview(
            session, header, data_rows, from_date or None, to_date or None
        )

        # Cache per-company stock items for rate/uom defaults
        items_by_company: dict[str, dict[str, StockItem]] = {}

        created_ids: list[int] = []
        for bundle in preview["entries"]:
            company = bundle["company"]
            voucher_type = bundle["voucher_type"]
            entry_date = bundle["date"]
            policy = get_policy(session, company, voucher_type)
            if policy is None:
                # Should not happen — preview already filtered, but stay defensive.
                continue
            entry = ProductionEntry(
                remote_id="pending",
                company_name=company,
                entry_date=entry_date,
                voucher_type=voucher_type,
                status="draft",
                narration=f"Imported consumables · {entry_date}",
            )
            session.add(entry)
            session.flush()
            entry.remote_id = generate_remote_id(entry_date, voucher_type, entry.id)
            if company not in items_by_company:
                items_by_company[company] = {
                    it.name: it
                    for it in session.scalars(
                        select(StockItem).where(StockItem.company_name == company)
                    )
                }
            by_name = items_by_company[company]
            for line in bundle["lines"]:
                item = by_name.get(line["item"])
                uom = (item.closing_uom or item.base_units or "No.") if item else "No."
                rate = (item.closing_rate if item else 0.0) or 0.0
                qty = float(line["qty"])
                amount = round(qty * rate, 2)
                godown = default_godown_for_role(policy, "consume")
                session.add(
                    ProductionEntryLine(
                        entry_id=entry.id,
                        role="consume",
                        item_name=line["item"],
                        quantity=qty,
                        uom=uom,
                        rate=rate,
                        amount=amount,
                        godown=godown,
                        opening_stock_snapshot=item.closing_quantity if item else 0.0,
                        description=None,
                    )
                )
            created_ids.append(entry.id)

        session.commit()
        return {
            "created_entry_ids": created_ids,
            "created_count": len(created_ids),
            "error_count": preview["error_count"],
            "errors": preview["errors"],
        }
